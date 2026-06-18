import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from plotly import graph_objects as go
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import GradientBoostingRegressor
try:
    from xgboost import XGBRegressor
    _XGBOOST_AVAILABLE = True
except ImportError:
    _XGBOOST_AVAILABLE = False
import warnings
import re
import io as _io_mod
import os as _os_cfg

# ── Lève la limite de taille d'upload (par défaut 200 Mo) à 5 Go.
# On écrit la config AVANT que Streamlit n'initialise le serveur. Ainsi
# aucun fichier annexe n'est nécessaire : le partage du seul app.py suffit.
try:
    _cfg_dir = _os_cfg.path.join(_os_cfg.path.expanduser("~"), ".streamlit")
    _os_cfg.makedirs(_cfg_dir, exist_ok=True)
    _cfg_path = _os_cfg.path.join(_cfg_dir, "config.toml")
    _need_write = True
    if _os_cfg.path.exists(_cfg_path):
        with open(_cfg_path, "r", encoding="utf-8") as _f:
            if "maxUploadSize" in _f.read():
                _need_write = False
    if _need_write:
        with open(_cfg_path, "a", encoding="utf-8") as _f:
            _f.write("\n[server]\nmaxUploadSize = 5000\nmaxMessageSize = 5000\n")
except Exception:
    pass
# Variable d'environnement (prise en compte aussi par certains lancements)
_os_cfg.environ.setdefault("STREAMLIT_SERVER_MAX_UPLOAD_SIZE", "5000")

warnings.filterwarnings("ignore")

st.set_page_config(
    layout="wide",
    page_title="PricePackaging",
    page_icon="📦",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for _k, _v in {
    "run_analysis": False,
    "page": "home",
    "category": None,
    "selected_ref": None,
    "selected_component": None,
    "ref_search_input": "",
    "ref_components_summary": None,
    "ref_not_found": False,
    "browse_category": None,    # Catégorie en navigation par type (carton/sac/palette/cale)
    "browse_type": None,        # Type strict sélectionné
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap');
* { font-family:'DM Sans',sans-serif; box-sizing:border-box; }
html,body,[data-testid="stAppViewContainer"] { background:#f0f4f8; }
[data-testid="stMainBlockContainer"] { padding-top:2rem; }

/* ── Remplacer l'animation "Running" (spinner avec bicyclette/arrêt) de Streamlit
   par un simple badge "Loading…" plus discret et professionnel ── */
/* On garde le toolbar masqué (menu Deploy etc.) mais on AFFICHE le status widget
   pour que l'utilisateur voie quand Streamlit est en train de calculer. */
[data-testid="stToolbar"] { display:none !important; }

/* Style amélioré pour l'indicateur "Running..." natif Streamlit */
/* === INDICATEUR DE CHARGEMENT (visible en haut à droite) ===
   Quand Streamlit est en train de recalculer la page, on amplifie son
   widget natif "RUNNING..." pour qu'il soit IMPOSSIBLE à manquer. */
[data-testid="stStatusWidget"]{
  position:fixed !important;top:14px !important;right:18px !important;
  z-index:99999 !important;
  pointer-events:none;
}
[data-testid="stStatusWidget"] > div{
  background:linear-gradient(135deg,#1e40af,#3b82f6) !important;
  color:white !important;
  padding:10px 22px !important;border-radius:24px !important;
  font-size:14px !important;font-weight:800 !important;
  letter-spacing:1px !important;
  box-shadow:0 8px 28px rgba(59,130,246,0.55),0 0 0 2px rgba(59,130,246,0.2) !important;
  border:2px solid rgba(255,255,255,0.4) !important;
  animation:loadingBadgePulse 1.2s ease-in-out infinite !important;
  text-transform:uppercase !important;
}
[data-testid="stStatusWidget"] svg{
  fill:white !important;color:white !important;
  width:18px !important;height:18px !important;
  margin-right:6px !important;
}
@keyframes loadingBadgePulse{
  0%,100%{transform:scale(1);box-shadow:0 8px 28px rgba(59,130,246,0.55),0 0 0 2px rgba(59,130,246,0.2);}
  50%{transform:scale(1.06);box-shadow:0 12px 36px rgba(59,130,246,0.75),0 0 0 6px rgba(59,130,246,0.15);}
}

/* Bandeau de chargement plein écran semi-transparent (déclenché par classe sur body) */
.app-loading-overlay{position:fixed;top:0;left:0;right:0;height:4px;
  background:linear-gradient(90deg,transparent,#3b82f6,#10b981,#3b82f6,transparent);
  background-size:200% 100%;
  z-index:99998;display:none;
  animation:loadingBar 1.4s linear infinite;}
@keyframes loadingBar{0%{background-position:200% 0;}100%{background-position:-200% 0;}}
body:has([data-testid="stStatusWidget"]) .app-loading-overlay{display:block;}

/* ── Hero ── */
.home-hero{background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 100%);border-radius:20px;
  padding:52px 48px;text-align:center;margin-bottom:28px;
  box-shadow:0 8px 40px rgba(15,23,42,0.25);}
.home-hero h1{font-size:46px;font-weight:800;color:white;margin:0;letter-spacing:-2px;}
.home-hero p{font-size:15px;color:#94a3b8;margin-top:10px;}
.home-hero .accent{color:#38bdf8;}

/* ── Search ── */
.ref-search-box{background:white;border-radius:16px;padding:32px 36px;margin-bottom:24px;
  box-shadow:0 4px 24px rgba(0,0,0,0.07);border-top:4px solid #3b82f6;}
.ref-search-box h2{font-size:20px;font-weight:800;color:#0f172a;margin:0 0 6px 0;}
.ref-search-box p{font-size:13px;color:#64748b;margin:0 0 20px 0;}

/* ── Category cards ── */
.cat-card{background:white;border-radius:12px;padding:10px 10px 12px;
    box-shadow:0 2px 10px rgba(0,0,0,0.04);text-align:center;
    border:1px solid #f1f5f9;transition:box-shadow .15s,border-color .15s;
    min-height:130px;margin:4px;display:flex;flex-direction:column;align-items:center;gap:6px;}
.cat-card:hover{box-shadow:0 6px 20px rgba(59,130,246,0.10);border-color:#bfdbfe;}
.cat-icon{font-size:26px;margin-bottom:2px;}
.cat-label{font-size:13px;font-weight:800;color:#0f172a;margin:2px 0 2px;}
.cat-desc{font-size:10px;color:#94a3b8;margin:0;}
.cat-type-badge{font-size:11px;color:#1e3a5f;font-weight:700;font-family:'DM Sans',sans-serif;
  background:#e0f2fe;border:1px solid #7dd3fc;border-radius:6px;padding:3px 10px;margin:4px 0 2px;
  max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.cat-price-badge{display:inline-block;background:linear-gradient(135deg,#3b82f6,#2563eb);
  color:white;border-radius:30px;padding:5px 14px;font-size:15px;font-weight:800;
  margin:4px 0 2px;box-shadow:0 3px 10px rgba(59,130,246,0.3);}
.cat-price-badge.mod-color{background:linear-gradient(135deg,#7c3aed,#6d28d9);}
.cat-price-badge.transport-color{background:linear-gradient(135deg,#0ea5e9,#0284c7);}
.price-indicator{display:inline-block;font-size:10px;font-weight:700;padding:4px 10px;
  border-radius:20px;margin-top:4px;text-transform:uppercase;letter-spacing:0.5px;}
.price-good{background:#d1fae5;color:#065f46;}
.price-medium{background:#fef3c7;color:#92400e;}
.price-high{background:#fee2e2;color:#991b1b;}
.price-neutral{background:#e2e8f0;color:#475569;}
.cat-no-data{font-size:11px;color:#cbd5e1;font-style:italic;margin:6px 0;}

/* ── Dim-badges container : affichage élégant des dimensions fixes d'un type ── */
.dim-badges-container{margin:18px 0 14px 0;
  background:linear-gradient(135deg,#eff6ff 0%,#dbeafe 100%);
  border:1px solid #bfdbfe;border-radius:14px;padding:18px 22px;
  box-shadow:0 2px 10px rgba(59,130,246,0.08);
  position:relative;overflow:hidden;}
.dim-badges-container::before{content:"";position:absolute;top:0;left:0;right:0;height:3px;
  background:linear-gradient(90deg,#3b82f6,#8b5cf6,#06b6d4);}
.dim-badges-title{font-size:11px;font-weight:800;color:#1e3a8a;margin-bottom:10px;
  text-transform:uppercase;letter-spacing:1.2px;}
.dim-badges-row{display:flex;flex-wrap:wrap;gap:14px;}
.dim-badge{display:flex;flex-direction:column;align-items:flex-start;
  background:white;border:1px solid #bfdbfe;border-radius:10px;
  padding:10px 18px;box-shadow:0 1px 3px rgba(0,0,0,0.04);
  min-width:96px;position:relative;}
.dim-badge::after{content:"";position:absolute;left:0;top:8px;bottom:8px;width:3px;
  background:linear-gradient(180deg,#3b82f6,#1e40af);border-radius:0 3px 3px 0;}
.dim-badge-label{font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;
  letter-spacing:0.8px;padding-left:6px;}
.dim-badge-value{font-size:22px;color:#0f172a;font-weight:800;line-height:1.1;
  padding-left:6px;letter-spacing:-0.5px;margin-top:2px;}
.dim-badge-unit{font-size:11px;color:#64748b;font-weight:600;padding-left:6px;}

/* ── Browse-by-type section (2e rangée d'icônes) ─────────────────────── */
.browse-section{margin-top:36px;margin-bottom:8px;
  background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#0c4a6e 100%);
  border:1px solid #1e40af;border-radius:18px;padding:26px 28px;
  box-shadow:0 8px 28px rgba(15,23,42,0.25);position:relative;overflow:hidden;}
.browse-section::before{content:"";position:absolute;top:0;left:0;right:0;height:4px;
  background:linear-gradient(90deg,#38bdf8,#06b6d4,#8b5cf6);}
.browse-section::after{content:"";position:absolute;top:-40px;right:-40px;width:180px;height:180px;
  background:radial-gradient(circle,rgba(56,189,248,0.15) 0%,transparent 70%);border-radius:50%;}
.browse-section h3{font-size:20px;font-weight:800;color:white;margin:0 0 8px 0;
  letter-spacing:-0.4px;display:flex;align-items:center;gap:10px;position:relative;z-index:1;}
.browse-section p{font-size:13px;color:#94a3b8;margin:0 0 20px 0;position:relative;z-index:1;}
.browse-card{background:white;border-radius:14px;padding:16px 10px;
  box-shadow:0 4px 14px rgba(15,23,42,0.18);border:2px solid transparent;
  transition:all .25s ease;text-align:center;min-height:128px;
  display:flex;flex-direction:column;justify-content:center;align-items:center;gap:7px;
  cursor:pointer;position:relative;z-index:1;}
.browse-card:hover{border-color:#38bdf8;transform:translateY(-4px);
  box-shadow:0 12px 28px rgba(56,189,248,0.25);}
.browse-card .b-icon{font-size:36px;line-height:1;filter:drop-shadow(0 2px 4px rgba(0,0,0,0.1));}
.browse-card .b-label{font-size:14px;font-weight:800;color:#0f172a;letter-spacing:-0.2px;}
.browse-card .b-desc{font-size:10px;color:#64748b;font-style:italic;}
.browse-card .b-count{font-size:11px;font-weight:700;color:white;
  background:linear-gradient(135deg,#0ea5e9,#0284c7);border-radius:20px;padding:3px 12px;
  box-shadow:0 2px 6px rgba(8,145,178,0.3);}

/* ── Type-grid (page de sélection d'un type au sein d'une catégorie) ── */
.type-grid-header{background:linear-gradient(135deg,#1e40af 0%,#3b82f6 100%);
  color:white;border-radius:16px;padding:24px 28px;margin-bottom:20px;
  box-shadow:0 6px 20px rgba(59,130,246,0.25);position:relative;overflow:hidden;}
.type-grid-header::before{content:"";position:absolute;top:-30px;right:-30px;
  width:140px;height:140px;background:rgba(255,255,255,0.08);border-radius:50%;}
.type-grid-header h2{margin:0;font-size:24px;font-weight:900;letter-spacing:-0.5px;
  display:flex;align-items:center;gap:12px;}
.type-grid-header p{margin:6px 0 0 0;font-size:13px;opacity:0.85;font-weight:500;}
.type-card{background:white;border-radius:14px;padding:18px 14px;
  box-shadow:0 3px 12px rgba(15,23,42,0.06);border:2px solid #e2e8f0;
  transition:all .2s ease;text-align:center;min-height:170px;
  display:flex;flex-direction:column;justify-content:center;align-items:center;gap:8px;
  cursor:pointer;position:relative;overflow:hidden;}
.type-card::before{content:"";position:absolute;top:0;left:0;right:0;height:4px;
  background:linear-gradient(90deg,#3b82f6,#8b5cf6,#ec4899);opacity:0;
  transition:opacity .25s ease;}
.type-card:hover{border-color:#3b82f6;transform:translateY(-4px);
  box-shadow:0 12px 28px rgba(59,130,246,0.18);}
.type-card:hover::before{opacity:1;}
.type-card .t-svg{width:64px;height:64px;display:block;margin:0 auto 8px;}
.type-card .t-name{font-size:13px;font-weight:800;color:#0f172a;letter-spacing:-0.2px;
  line-height:1.3;min-height:34px;display:flex;align-items:center;justify-content:center;}
.type-card .t-stats{font-size:11px;color:#64748b;font-weight:600;
  background:#f1f5f9;border-radius:8px;padding:4px 10px;margin-top:4px;}
.type-card .t-price{font-size:12px;color:#059669;font-weight:800;margin-top:4px;}

/* ── Banners ── */
.ref-found-banner{background:linear-gradient(135deg,#f0fdf4,#dcfce7);
  border:2px solid #10b981;border-radius:14px;padding:14px 22px;margin-bottom:18px;
  display:flex;align-items:center;gap:12px;font-size:14px;font-weight:700;color:#065f46;}
.ref-not-found-banner{background:#fef2f2;border:2px solid #ef4444;border-radius:14px;
  padding:14px 22px;margin-bottom:18px;font-size:14px;font-weight:700;color:#991b1b;}

/* ── Ref summary bar ── */
.ref-summary{background:linear-gradient(135deg,#0f172a,#1e3a5f);border-radius:16px;
  padding:22px 28px;margin-bottom:20px;display:flex;align-items:center;gap:20px;
  box-shadow:0 4px 20px rgba(15,23,42,0.2);}
.ref-code{font-size:28px;font-weight:900;color:#38bdf8;font-family:'DM Mono',monospace;}
.ref-meta{color:#94a3b8;font-size:13px;}
.ref-meta strong{color:white;font-size:15px;display:block;margin-bottom:2px;}

/* ── Component rows ── */
.comp-table-wrap{background:white;border-radius:16px;padding:24px 28px;
  box-shadow:0 4px 24px rgba(0,0,0,0.07);margin-bottom:24px;}
.comp-table-wrap h2{font-size:18px;font-weight:800;color:#0f172a;margin:0 0 4px 0;}
.comp-table-wrap .sub{font-size:12px;color:#64748b;margin:0 0 18px 0;}
.comp-row{display:flex;align-items:flex-start;gap:12px;padding:14px 16px;
  border-radius:12px;margin-bottom:8px;border:1.5px solid #e2e8f0;background:#f8fafc;}
.c-icon{font-size:22px;flex-shrink:0;width:36px;text-align:center;margin-top:2px;}
.c-body{flex:1;min-width:0;}
.c-name{font-size:14px;font-weight:700;color:#0f172a;margin-bottom:2px;}
.c-meta{font-size:11px;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.6px;}
.c-dims{font-size:12px;color:#475569;margin-top:6px;padding:5px 10px;background:#f1f5f9;
  border-radius:8px;display:inline-block;}
.c-price{font-size:18px;font-weight:800;color:#0f172a;align-self:center;flex-shrink:0;}
.c-price-orig{font-size:10px;color:#94a3b8;font-weight:500;text-align:right;}

/* ── Section header ── */
.section-header{background:white;padding:22px 28px;border-radius:16px;margin-bottom:20px;
  box-shadow:0 2px 14px rgba(0,0,0,0.05);border-left:5px solid #3b82f6;}
.section-header h2{color:#0f172a;margin:0;font-size:21px;font-weight:800;}
.section-header p{color:#64748b;margin:6px 0 0 0;font-size:13px;}

/* ── Info panel — DESIGN AMÉLIORÉ : grille élégante, dégradés, hiérarchie claire ── */
.comp-info-panel{background:linear-gradient(135deg,#ffffff 0%,#f8fafc 100%);
  border-radius:18px;padding:26px 30px;
  box-shadow:0 4px 20px rgba(15,23,42,0.06),0 1px 3px rgba(15,23,42,0.04);
  margin-bottom:20px;border:1px solid #e2e8f0;
  position:relative;overflow:hidden;}
.comp-info-panel::before{content:"";position:absolute;top:0;left:0;right:0;height:4px;
  background:linear-gradient(90deg,#3b82f6,#8b5cf6,#06b6d4);}
.comp-info-panel h3{font-size:17px;font-weight:800;color:#0f172a;margin:0 0 20px 0;
  padding-bottom:14px;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:10px;
  letter-spacing:-0.2px;}
.comp-info-panel h3::before{content:"📋";font-size:18px;}
.info-grid{display:grid;grid-template-columns:repeat(4, minmax(0, 1fr));gap:12px;
  margin-bottom:6px;align-items:stretch;}
@media (max-width: 1100px){.info-grid{grid-template-columns:repeat(3, minmax(0, 1fr));}}
@media (max-width: 760px){.info-grid{grid-template-columns:repeat(2, minmax(0, 1fr));}}
.info-item{background:white;border-radius:12px;padding:14px 16px;
  border:1px solid #e2e8f0;
  display:flex;flex-direction:column;justify-content:center;
  min-height:88px;height:100%;
  transition:all .2s ease;
  position:relative;}
.info-item::after{content:"";position:absolute;left:0;top:14px;bottom:14px;width:3px;
  background:linear-gradient(180deg,#3b82f6,#8b5cf6);border-radius:0 3px 3px 0;
  opacity:.7;}
.info-item:hover{border-color:#3b82f6;box-shadow:0 4px 14px rgba(59,130,246,0.12);
  transform:translateY(-1px);}
.info-item .il{font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;
  letter-spacing:1px;margin-bottom:8px;line-height:1.2;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;padding-left:8px;}
.info-item .iv{font-size:19px;font-weight:800;color:#0f172a;line-height:1.15;
  display:flex;align-items:baseline;gap:6px;flex-wrap:wrap;
  word-break:break-word;padding-left:8px;letter-spacing:-0.3px;}
.info-item .iv-text{font-size:13px;font-weight:700;color:#1e293b;line-height:1.35;
  word-break:break-word;padding-left:8px;}
.info-item .iu{font-size:11px;color:#64748b;font-weight:600;}
/* Variant : item textuel (Type, Fournisseur, Site) — fond légèrement teinté */
.info-item.txt{background:linear-gradient(135deg,#f8fafc 0%,#eff6ff 100%);}
.info-item.txt::after{background:linear-gradient(180deg,#06b6d4,#0891b2);}
/* Variant : item numérique principal (Prix, Coût) — accent vert */
.info-item.main{background:linear-gradient(135deg,#f0fdf4 0%,#ecfdf5 100%);
  border-color:#bbf7d0;}
.info-item.main::after{background:linear-gradient(180deg,#10b981,#059669);}
.info-item.main .iv{color:#065f46;}

/* ── KPI blocks ── */
.kpi-blue  {background:linear-gradient(135deg,#3b82f6,#2563eb);padding:24px;border-radius:16px;
  color:white;text-align:center;box-shadow:0 8px 24px rgba(59,130,246,0.28);margin-bottom:16px;}
.kpi-green {background:linear-gradient(135deg,#10b981,#059669);padding:24px;border-radius:16px;
  color:white;text-align:center;box-shadow:0 8px 24px rgba(16,185,129,0.28);margin-bottom:16px;}
.kpi-orange{background:linear-gradient(135deg,#f59e0b,#d97706);padding:24px;border-radius:16px;
  color:white;text-align:center;box-shadow:0 8px 24px rgba(245,158,11,0.28);margin-bottom:16px;}
.kpi-red   {background:linear-gradient(135deg,#ef4444,#dc2626);padding:24px;border-radius:16px;
  color:white;text-align:center;box-shadow:0 8px 24px rgba(239,68,68,0.28);margin-bottom:16px;}
.kpi-purple{background:linear-gradient(135deg,#7c3aed,#6d28d9);padding:24px;border-radius:16px;
  color:white;text-align:center;box-shadow:0 8px 24px rgba(124,58,237,0.28);margin-bottom:16px;}
.kpi-value{font-size:42px;font-weight:900;margin:8px 0;letter-spacing:-1px;}
.kpi-label{font-size:11px;font-weight:700;opacity:.9;text-transform:uppercase;letter-spacing:1.5px;}
.kpi-status{font-size:13px;font-weight:700;margin-top:8px;text-transform:uppercase;letter-spacing:1px;}
.badge-pos{display:inline-block;padding:4px 13px;border-radius:20px;font-weight:700;
  font-size:11px;text-transform:uppercase;letter-spacing:0.8px;margin-top:6px;}
.badge-blue  {background:#dbeafe;color:#1e40af;}
.badge-orange{background:#fef3c7;color:#92400e;}

/* ── Result card ── */
.result-info-card{background:white;border-radius:14px;padding:18px 22px;margin:14px 0;
  box-shadow:0 2px 14px rgba(0,0,0,0.06);border-left:4px solid #3b82f6;}
.result-info-card h4{color:#64748b;font-size:10px;font-weight:700;margin:0 0 12px 0;
  text-transform:uppercase;letter-spacing:1.2px;}
.result-info-row{display:flex;gap:22px;flex-wrap:wrap;}
.ri-label{font-size:10px;color:#94a3b8;font-weight:700;text-transform:uppercase;
  letter-spacing:.8px;margin-bottom:3px;}
.ri-value{font-size:14px;font-weight:800;color:#0f172a;}

/* ── Stat cards ── */
.stat-card{background:white;padding:16px;border-radius:14px;text-align:center;
  box-shadow:0 2px 10px rgba(0,0,0,0.05);border-top:3px solid #3b82f6;}
.stat-card h4{color:#64748b;font-size:10px;font-weight:700;text-transform:uppercase;
  letter-spacing:1px;margin:0 0 6px 0;}
.stat-card .sub{font-size:9px;color:#94a3b8;margin:4px 0 0 0;}
.stat-value{color:#0f172a;font-size:22px;font-weight:800;margin:0;}

/* ── Table section ── */
.table-section{background:white;padding:20px;border-radius:16px;
  box-shadow:0 2px 14px rgba(0,0,0,0.05);margin:16px 0;}
.table-section h3{color:#0f172a;font-size:16px;font-weight:700;margin:0 0 12px 0;
  padding-bottom:10px;border-bottom:2px solid #f1f5f9;}

/* ── Info boxes ── */
.info-box{background:#eff6ff;border-left:4px solid #3b82f6;padding:13px 17px;
  border-radius:10px;color:#1e40af;font-weight:600;margin:12px 0;font-size:13px;}
.currency-note{background:#fefce8;border-left:4px solid #eab308;padding:10px 16px;
  border-radius:8px;color:#713f12;font-weight:600;margin:8px 0;font-size:12px;}
.warning-box{background:#fef2f2;border-left:4px solid #ef4444;padding:13px 17px;
  border-radius:10px;color:#991b1b;font-weight:600;margin:12px 0;font-size:13px;}

/* ── Config section ── */
.config-section{background:white;padding:24px;border-radius:16px;
  box-shadow:0 2px 14px rgba(0,0,0,0.05);}
.config-section h3{color:#0f172a;margin-bottom:14px;font-size:16px;font-weight:700;}
.dim-display{background:#0f172a;color:white;padding:16px;border-radius:12px;
  margin:12px 0;font-weight:600;text-align:center;font-size:14px;line-height:1.7;}

/* ── Streamlit overrides ── */
.stTextInput input,.stNumberInput input{
  border:2px solid #e2e8f0 !important;border-radius:10px !important;padding:9px !important;
  font-size:14px !important;background:white !important;color:#0f172a !important;font-weight:600 !important;}
.stTextInput input:focus,.stNumberInput input:focus{
  border-color:#3b82f6 !important;box-shadow:0 0 0 3px rgba(59,130,246,.15) !important;}
.stNumberInput label,.stSelectbox label,.stRadio>label{
  color:#0f172a !important;font-weight:700 !important;font-size:13px !important;}
.stSelectbox>div>div{border-radius:10px !important;background:white !important;
  border:2px solid #e2e8f0 !important;color:#0f172a !important;font-weight:600 !important;}
.stButton>button{background:#1e293b;color:white;border-radius:10px;font-weight:700;
  padding:10px 24px;border:none;font-size:13px;transition:all .2s ease;
  text-transform:uppercase;letter-spacing:.5px;}
.stButton>button:hover{background:#0f172a;transform:translateY(-1px);}
.stButton>button[kind="primary"]{background:linear-gradient(135deg,#3b82f6,#2563eb);
  box-shadow:0 4px 14px rgba(59,130,246,.35);}
.stButton>button[kind="primary"]:hover{box-shadow:0 6px 20px rgba(59,130,246,.5);}
.stTabs [role="tablist"] button{color:#64748b !important;font-weight:600 !important;font-size:14px !important;}
.stTabs [role="tablist"] button[aria-selected="true"]{
  color:#3b82f6 !important;border-bottom-color:#3b82f6 !important;}
</style>

<!-- Badge "Loading" personnalisé (remplace l'animation Streamlit) -->
<div class="app-loading-badge">Loading</div>
<script>
(function() {
    // Surveille l'apparition de l'indicateur "running" Streamlit pour activer notre badge
    const observer = new MutationObserver(() => {
        const status = document.querySelector('[data-testid="stStatusWidget"]');
        if (status && status.textContent && status.textContent.toLowerCase().includes('running')) {
            document.body.classList.add('app-running');
        } else {
            document.body.classList.remove('app-running');
        }
    });
    observer.observe(document.body, {childList:true, subtree:true});
})();
</script>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ═════════════════════════════════════════════════════════════════════════════

DATA_PATH       = "C:/Users/p134959/Downloads/EPDS_emballage (copie).xlsx"
DATA_PATH_CURRENCY = "C:/Users/p134959/Copie de EPDS.xlsx"  # Fichier d'appoint pour combler les currency manquantes

COL_DESIGNATION = "frpackagingdesignation packaging"
COL_PRICE       = "unitprice packaging"
COL_TOTAL_COST  = "total_cost"
COL_LENGTH      = "length packaging"
COL_WIDTH       = "width packaging"
COL_HEIGHT      = "height packaging"
COL_WEIGHT      = "weight packaging"
COL_THICKNESS   = "thickness packaging"
COL_MATERIAL    = "materialcode packaging"  # Code matière du composant
COL_PARTS_QTY   = "partsqty packaging"  # Quantité de pièces : prix_unitaire = prix / partsqty
COL_PACKAGING_QTY = "packagingqty packaging"  # Quantité packaging utilisée par référence
COL_ECON_DATE   = "economic_valid_date"  # Date de validation économique (moment où le document a été validé)
COL_LABOUR_SECONDS = "total_labour_seconds"  # Secondes de main d'œuvre totale (pour MOD)
COL_PACKAGING_LEVEL = "packaginglevel packaging"  # Niveau de packaging (4 = pack final)
COL_SUPPLIER    = "suppliername"
COL_SUPPLIER_ACCOUNT = "supplier_account"  # Fallback : utilisé quand suppliername est inconnu/vide
COL_REF         = "Reference_code"
COL_LABOUR      = "labour_rate_hour"
COL_LABOUR_COST = "labour_cost"
COL_SHIPPING    = "shipping_charge"
COL_CURRENCY    = "currency"
COL_MFG_SITE    = "manufacturing_site"
COL_DISPATCH    = "dispatch_site"
# Site client (destination). Pour le transport on compare à dimensions de
# site d'enlèvement (= dispatch_site) ET site client identiques.
# Le nom exact de la colonne est auto-détecté au chargement (voir load_data).
COL_CLIENT_SITE = "client_site"
# Pays du fournisseur (fabrication et expédition) — codes ISO type ES/TR/FR
COL_COUNTRY_FAB      = "Supplier_country_fab"
COL_COUNTRY_DISPATCH = "Supplier_country_dispatch"

# Codes pays → nom complet (pour affichage lisible)
COUNTRY_NAMES = {
    "FR": "France", "ES": "Espagne", "TR": "Turquie", "DE": "Allemagne",
    "IT": "Italie", "PT": "Portugal", "RO": "Roumanie", "MA": "Maroc",
    "PL": "Pologne", "CZ": "Tchéquie", "SK": "Slovaquie", "SI": "Slovénie",
    "GB": "Royaume-Uni", "UK": "Royaume-Uni", "BE": "Belgique", "NL": "Pays-Bas",
    "RU": "Russie", "IN": "Inde", "CN": "Chine", "JP": "Japon", "KR": "Corée du Sud",
    "BR": "Brésil", "AR": "Argentine", "US": "États-Unis", "MX": "Mexique",
    "TN": "Tunisie", "DZ": "Algérie", "HU": "Hongrie", "AT": "Autriche",
    "CH": "Suisse", "SE": "Suède", "RS": "Serbie", "BG": "Bulgarie", "UA": "Ukraine",
    # Élargissement (Europe + monde) pour couvrir tous les codes du parc
    "IE": "Irlande", "DK": "Danemark", "FI": "Finlande", "NO": "Norvège",
    "GR": "Grèce", "HR": "Croatie", "LT": "Lituanie", "LV": "Lettonie",
    "EE": "Estonie", "LU": "Luxembourg", "MT": "Malte", "CY": "Chypre",
    "IS": "Islande", "MK": "Macédoine du Nord", "BA": "Bosnie-Herzégovine",
    "AL": "Albanie", "ME": "Monténégro", "MD": "Moldavie", "BY": "Biélorussie",
    "CA": "Canada", "CL": "Chili", "CO": "Colombie", "PE": "Pérou",
    "ZA": "Afrique du Sud", "EG": "Égypte", "NG": "Nigéria", "KE": "Kenya",
    "AE": "Émirats arabes unis", "SA": "Arabie saoudite", "IL": "Israël",
    "TH": "Thaïlande", "VN": "Viêt Nam", "ID": "Indonésie", "MY": "Malaisie",
    "SG": "Singapour", "PH": "Philippines", "TW": "Taïwan", "HK": "Hong Kong",
    "AU": "Australie", "NZ": "Nouvelle-Zélande", "PK": "Pakistan", "BD": "Bangladesh",
}

def country_full_name(code: str) -> str:
    """Renvoie le nom complet du pays depuis son code ISO (ES → Espagne)."""
    if not code:
        return ""
    c = str(code).strip().upper()
    return COUNTRY_NAMES.get(c, c)  # si inconnu, on renvoie le code tel quel


# Drapeaux emoji (codes ISO → emoji régional)
COUNTRY_FLAGS = {
    "FR":"🇫🇷","ES":"🇪🇸","TR":"🇹🇷","DE":"🇩🇪","IT":"🇮🇹","PT":"🇵🇹","RO":"🇷🇴",
    "MA":"🇲🇦","PL":"🇵🇱","CZ":"🇨🇿","SK":"🇸🇰","SI":"🇸🇮","GB":"🇬🇧","UK":"🇬🇧",
    "BE":"🇧🇪","NL":"🇳🇱","RU":"🇷🇺","IN":"🇮🇳","CN":"🇨🇳","JP":"🇯🇵","KR":"🇰🇷",
    "BR":"🇧🇷","AR":"🇦🇷","US":"🇺🇸","MX":"🇲🇽","TN":"🇹🇳","DZ":"🇩🇿","HU":"🇭🇺",
    "AT":"🇦🇹","CH":"🇨🇭","SE":"🇸🇪","RS":"🇷🇸","BG":"🇧🇬","UA":"🇺🇦",
}

def country_flag(code: str) -> str:
    """Emoji drapeau depuis le code ISO. Génère le drapeau dynamiquement à
    partir des deux lettres (symboles indicateurs régionaux Unicode), ce qui
    couvre TOUS les pays — pas seulement une liste figée."""
    if not code:
        return "🏳️"
    c = str(code).strip().upper()
    # Drapeau explicite si présent (UK→GB déjà géré)
    if c in COUNTRY_FLAGS:
        return COUNTRY_FLAGS[c]
    # Génération Unicode : 2 lettres A–Z → indicateurs régionaux
    if len(c) == 2 and c.isalpha():
        try:
            return chr(0x1F1E6 + (ord(c[0]) - 65)) + chr(0x1F1E6 + (ord(c[1]) - 65))
        except Exception:
            return "🏳️"
    return "🏳️"


def country_from_site(site_value: str) -> str:
    """
    Extrait le code pays (2 lettres) depuis un libellé de site qui se termine
    par le code pays. Ex :
      "00013498 - 28 - GEODIS PLATE-FORME LOGISTIQUE - LIMAY - FR" → "FR"
      "00027104 - 01 - VISSCHER CARAVELLE POLAND ... - GRUDZIADZ - PL" → "PL"
    On prend le dernier segment séparé par '-' s'il fait exactement 2 lettres.
    """
    if not site_value:
        return ""
    s = str(site_value).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    # Découpe par tiret et on cherche le dernier segment de 2 lettres
    parts = [p.strip() for p in s.split("-") if p.strip()]
    for seg in reversed(parts):
        if len(seg) == 2 and seg.isalpha():
            return seg.upper()
    # Fallback : 2 derniers caractères alphabétiques
    tail = s.replace(" ", "")[-2:]
    if tail.isalpha():
        return tail.upper()
    return ""


def parse_econ_date(series_or_val):
    """
    Parse robuste et VECTORISÉ de la date de validation économique. Gère :
      - les dates texte (ISO, FR jj/mm/aaaa, etc.)
      - les numéros de série Excel (ex. 45366 → date réelle)
      - les valeurs déjà au format datetime
      - les colonnes de TYPE MIXTE (texte + nombres)
    Préserve l'INDEX ORIGINAL. Renvoie une Series datetime (ou un Timestamp
    si on passe une valeur unique). Vectorisé → rapide même sur 100k+ lignes.
    """
    single = not isinstance(series_or_val, pd.Series)
    s = pd.Series([series_or_val]) if single else series_or_val

    result = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    # 0) NUMÉROS DE SÉRIE EXCEL d'abord (sinon to_datetime les lit comme des ns).
    #    On repère les valeurs purement numériques dans la plage des dates Excel.
    nums_all = pd.to_numeric(s, errors="coerce")
    is_excel = nums_all.notna() & (nums_all > 32000) & (nums_all < 80000)
    if is_excel.any():
        excel_dates = pd.to_datetime(nums_all[is_excel], unit="D",
                                     origin="1899-12-30", errors="coerce")
        result.loc[excel_dates.index] = excel_dates

    # 1) Le RESTE en datetime texte (FR jj/mm/aaaa en priorité).
    rest = s[~is_excel]
    if not rest.empty:
        parsed = pd.to_datetime(rest, errors="coerce", dayfirst=True)
        # 1b) NaT restants → format ISO (aaaa-mm-jj)
        mask_iso = parsed.isna()
        if mask_iso.any():
            iso_try = pd.to_datetime(rest[mask_iso], errors="coerce", dayfirst=False)
            good_iso = iso_try.dropna()
            if not good_iso.empty:
                parsed.loc[good_iso.index] = good_iso
        good = parsed.dropna()
        if not good.empty:
            result.loc[good.index] = good

    return result.iloc[0] if single else result


FEATURE_COLS_DIM = [
    "length packaging","width packaging","height packaging",
    "weight packaging","Surface_mm2",
    "ratio_LW","ratio_LH","ratio_WH","density","surface_vol_ratio",
]

# Mots-clés de classification — STRICTS
CATEGORY_KEYWORDS = {
    "carton":    ["carton","boite","boîte","caisse","cloche","coiffe","bouclier",
                  "pare-brise","pare brises","pare brise","galia"],
    "sac":       ["sac","sachet","bulles","vci","plastique vci"],
    "palette":   ["palette","kit carton palette"],
    "cale":      ["cale","insert"],
    "film":      ["film","filme","etirable","étirable","retractable","rétractable",
                  "film bulle","film bulles","stretch","houssage","housse plastique"],
    "mod":       ["mod","main d'oeuvre","main d'œuvre","main d oeuvre"],
    "transport": ["transport","shipping","expéd","expedition"],
}

# Mots-clés exclusifs : si présents, on évite la classification carton
NON_CARTON_KEYWORDS = ["etiquette","étiquette","label","autocollant","sticker"]

CATEGORY_ICONS = {
    "carton":"📦","sac":"🛍️","palette":"🪵",
    "cale":"🧩","film":"🎞️","mod":"🏭","transport":"🚚","autre":"🔹",
}
CAT_LABELS = {
    "carton":"📦 Carton","sac":"🛍️ Sac","palette":"🪵 Palette",
    "cale":"🧩 Cale","film":"🎞️ Film","mod":"🏭 Taux MOD","transport":"🚚 Transport","autre":"🔹 Autre",
}
HOME_CATEGORIES = [
    ("carton",   "📦","Carton",   "Boites, boucliers, Galia…"),
    ("sac",      "🛍️","Sac",      "Sac, bulles, VCI…"),
    ("palette",  "🪵","Palette",  "Bois, plastique, NIMP15…"),
    ("cale",     "🧩","Cale",     "Cales, inserts…"),
    ("mod",      "🏭","MOD",      "Main d'œuvre directe"),
    ("transport","🚚","Transport","Frais d'expédition"),
]

_INVALID_DESIG = {"","nan","none","inconnu","unknown","-","n/a","nd","null","#n/a"}

FX_TO_EUR: dict[str, float] = {
    "EUR":1.0,"USD":0.92,"GBP":1.17,"CHF":1.05,"CAD":0.68,"JPY":0.0061,
    "CNY":0.127,"MXN":0.054,"BRL":0.18,"INR":0.011,"PLN":0.23,"CZK":0.041,
    "HUF":0.0026,"RON":0.20,"SEK":0.087,"NOK":0.086,"DKK":0.134,"TRY":0.020,
    "ZAR":0.049,"AUD":0.60,"NZD":0.55,"SGD":0.68,"HKD":0.118,"KRW":0.00067,
    "THB":0.026,"MAD":0.092,"TND":0.29,"DZD":0.0068,"RUB":0.011,
}


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def is_valid_desig(val) -> bool:
    if not isinstance(val, str):
        return False
    return val.strip().lower() not in _INVALID_DESIG and len(val.strip()) > 0


def clean_supplier(col: pd.Series) -> pd.Series:
    return col.fillna("Inconnu").astype(str).str.strip().replace(["nan","None",""], "Inconnu")


def _sup_match(series: pd.Series, selected: str) -> pd.Series:
    """Matching TOLÉRANT d'un nom de fournisseur — insensible à la casse/aux
    espaces, et par 'contient' dans les deux sens. Ainsi 'EUROSTYLE' retrouve
    'EUROSTYLE SYSTEMS TANGER S A' et inversement."""
    if selected is None or selected == "— Tous —":
        return pd.Series([True] * len(series), index=series.index)
    sel_norm = str(selected).strip().upper()
    s_norm = series.astype(str).str.strip().str.upper()
    contains = s_norm.str.contains(re.escape(sel_norm), na=False)
    both_ways = pd.Series([(sel_norm in v) or (v in sel_norm) for v in s_norm],
                          index=series.index)
    return (s_norm == sel_norm) | contains | both_ways


def _safe_float(val, default: float = 0.0) -> float:
    try:
        v = float(val)
        return v if np.isfinite(v) else default
    except Exception:
        return default


def normalize_currency(code: str) -> str:
    """
    Normalise un code de devise. Pour les valeurs non identifiables
    (vide, None, NaN, codes inconnus comme 'X' ou un nombre), on retourne UNKNOWN
    et NON PAS EUR par défaut — pour éviter de fausser les comparaisons en mélangeant
    des prix de devises différentes (ex: TRY traité comme EUR).
    """
    if not isinstance(code, str):
        return "UNKNOWN"
    c = code.strip().upper()
    if c in ("", "NAN", "NONE", "NULL"):
        return "UNKNOWN"
    sym_map = {"€":"EUR","$":"USD","£":"GBP","¥":"JPY","₹":"INR","₩":"KRW"}
    if c in sym_map:
        return sym_map[c]
    return c


def to_eur(value: float, currency: str) -> tuple[float, bool]:
    """Convertit une valeur en EUR. Retourne (NaN, False) si la devise est
    UNKNOWN — la ligne est ainsi exclue des comparaisons numériques
    (médiane, min, max) au lieu d'introduire un biais en supposant EUR."""
    cur = normalize_currency(currency)
    if cur == "UNKNOWN":
        return float("nan"), False
    if cur == "EUR":
        return value, False
    rate = FX_TO_EUR.get(cur)
    if rate is None:
        # Devise normalisée mais sans taux de change → on ne convertit pas et on exclut
        return float("nan"), False
    return round(value * rate, 4), (cur != "EUR")


def add_derived(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    L   = df[COL_LENGTH].fillna(0)
    W   = df[COL_WIDTH].fillna(0)
    H   = df[COL_HEIGHT].fillna(0)
    Wt  = df[COL_WEIGHT].fillna(0)
    vol = L * W * H + 1e-9
    df["Surface_mm2"]       = 2*(L*W + L*H + W*H)
    df["Volume_m3"]         = vol / 1e9
    df["weight_kg"]         = Wt / 1000
    df["ratio_LW"]          = L / (W + 1e-9)
    df["ratio_LH"]          = L / (H + 1e-9)
    df["ratio_WH"]          = W / (H + 1e-9)
    df["density"]           = Wt / vol
    df["surface_vol_ratio"] = df["Surface_mm2"] / vol
    return df


def scroll_top():
    st.components.v1.html(
        "<script>window.parent.document.querySelector('section.main').scrollTo(0,0);</script>",
        height=0,
    )


def go_page(page: str, **kwargs):
    st.session_state.page = page
    st.session_state.run_analysis = False
    for k, v in kwargs.items():
        st.session_state[k] = v


# ═════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION & TYPE DETECTION (CORRIGÉ)
# ═════════════════════════════════════════════════════════════════════════════

def classify_component(desig: str) -> str:
    """
    Classifie un composant. CORRIGÉ : exclut explicitement 'étiquette' du carton.
    """
    if not isinstance(desig, str) or not desig.strip():
        return "autre"
    d = desig.lower()

    # GARDE-FOU 1 : si c'est une étiquette / label, ce n'est PAS un carton
    if any(k in d for k in NON_CARTON_KEYWORDS):
        return "autre"

    # GARDE-FOU 2 : "film bulles" doit être classé FILM (pas sac), donc on teste
    # le film en priorité s'il y a le mot "film".
    if "film" in d:
        return "film"

    # Classification standard
    for cat, kws in CATEGORY_KEYWORDS.items():
        if any(k in d for k in kws):
            return cat
    return "autre"


def get_clean_type_label(desig: str, cat: str) -> str:
    """
    Retourne un libellé propre et court correspondant au sous-type réel du composant.
    Évite d'afficher des désignations longues ou incohérentes.
    """
    if not isinstance(desig, str) or not desig.strip():
        return CAT_LABELS.get(cat, cat).split(" ", 1)[-1] if cat in CAT_LABELS else "—"
    d = desig.lower()

    if cat == "carton":
        if "boite" in d or "boîte" in d:    return "Boîte carton"
        if "caisse" in d:                   return "Caisse carton"
        if "galia" in d:                    return "Carton Galia"
        if "rabats" in d:                   return "Carton à rabats"
        if "coiffe" in d:                   return "Coiffe"
        if "cloche" in d:                   return "Carton Cloche"
        if "bouclier" in d:                 return "Bouclier"
        if "pare" in d and "brise" in d:    return "Pare-brise"
        return "Carton"

    if cat == "sac":
        if "bulles" in d:    return "Sac bulles"
        if "vci" in d:      return "Sac VCI"
        if "sachet" in d:   return "Sachet"
        return "Sac"

    if cat == "film":
        if "bulle" in d:                        return "Film bulles"
        if "etirable" in d or "étirable" in d:  return "Film étirable"
        if "retractable" in d or "rétractable" in d: return "Film rétractable"
        if "stretch" in d:                      return "Film étirable"
        if "housse" in d or "houssage" in d:    return "Housse"
        return "Film"

    if cat == "palette":
        if "nimp" in d or "nimp15" in d: return "Palette NIMP15"
        if "kit" in d:                   return "Kit palette"
        if "plastique" in d:             return "Palette plastique"
        if "bois" in d:                  return "Palette bois"
        return "Palette"

    if cat == "cale":
        if "insert" in d:  return "Insert"
        return "Cale"

    if cat == "mod":      return "Taux MOD"
    if cat == "transport":return "Transport"

    # Fallback : tronquer la désignation
    return desig if len(desig) <= 24 else desig[:22] + "…"


# ═════════════════════════════════════════════════════════════════════════════
# TYPES STRICTS — pour filtrage exact tableau & courbe
# ═════════════════════════════════════════════════════════════════════════════

def get_strict_carton_type(desig: str) -> str:
    """
    Retourne le TYPE STRICT d'un carton (pour filtrage exact).
    Aligné sur la liste des types réels de la base.
    Renvoie "" si non identifié (le composant sera alors filtré sur sa désignation brute).
    """
    if not isinstance(desig, str) or not desig.strip():
        return ""
    d_low = desig.strip().lower()

    # Exclure étiquettes / labels
    if any(k in d_low for k in NON_CARTON_KEYWORDS):
        return ""

    # ── Carton à rabats - Type Galia GXX (G03..G16, etc.)
    m = re.search(r'rabats.*galia\s*([gG]\s*\d{1,3})', d_low)
    if m:
        gid = re.sub(r'\s+', '', m.group(1)).upper()
        return f"Carton à rabats - Type Galia {gid}"

    # ── Carton Cloche - Type Galia GXX / 1A / 2A / G40
    m = re.search(r'cloche.*galia\s*([gG]\s*\d{1,3}|\d+\s*[aA])', d_low)
    if m:
        gid = re.sub(r'\s+', '', m.group(1)).upper()
        return f"Carton Cloche - Type Galia {gid}"

    # ── CARTON GALIA CLOCHE Cxx
    m = re.search(r'galia.*cloche\s*([cC]\s*\d{1,3})', d_low)
    if m:
        gid = re.sub(r'\s+', '', m.group(1)).upper()
        return f"CARTON GALIA CLOCHE {gid}"

    # ── Bouclier (variantes)
    if "bouclier" in d_low:
        if "1/2" in d_low or "demi" in d_low:
            return "1/2 Carton Boucliers"
        if "kit" in d_low and "bcl" in d_low:
            return "Carton Bouclier Kit BCL"
        if ("bcl" in d_low and "av" in d_low) or "av" in d_low.split():
            # Match strict "BCL AV"
            if re.search(r'\bbcl\b.*\bav\b', d_low) or re.search(r'\bav\b', d_low):
                return "Carton Bouclier BCL AV"
        return "Carton Bouclier"

    # ── Autres types
    if "pare" in d_low and "brise" in d_low:
        return "Carton Pare-Brises"
    if "coiffe" in d_low:
        return "Coiffe carton"
    if "pochette" in d_low:
        return "Pochette carton"
    if "kit" in d_low and "palette" in d_low and "iln" in d_low:
        return "Kit Carton Palette ILN 3P"
    if "boite" in d_low or "boîte" in d_low:
        return "Boite carton"
    if "caisse" in d_low:
        return "Caisse carton"

    return ""


def get_strict_palette_type(desig: str) -> str:
    """
    Retourne le TYPE STRICT d'une palette (pour filtrage exact).
    Aligné sur la liste des types réels de la base.
    """
    if not isinstance(desig, str) or not desig.strip():
        return ""
    d_low = desig.strip().lower()

    # ── Kit Carton Palette
    if "kit" in d_low and "carton" in d_low and "palette" in d_low and "iln" in d_low:
        return "Kit Carton Palette ILN 3P"

    # ── CBOU
    if "cbou" in d_low:
        if "2840" in d_low or "1680" in d_low:
            return "Palette pour CBOU--2840 et CBOU--1680"
        if "3240" in d_low:
            return "Palette pour CBOU--3240"

    # ── CPBR
    if "cpbr" in d_low and "2100" in d_low:
        return "Palette pour CPBR--2100"

    # ── CPOR
    if "cpor" in d_low:
        if "2400" in d_low:
            return "Palette pour CPOR--2400"
        if "2945" in d_low:
            return "Palette pour CPOR--2945"
        if "3630" in d_low:
            return "Palette pour CPOR--3630"

    # ── KIT ILN xx (2E, 3C, 3E, 3G, 3P, 3R, 4C, 4E, 4G, 4H, 4L, 6K, 8K, 8L, 9K)
    m = re.search(r'kit\s*iln\s*(\d+\s*[a-zA-Z])', d_low)
    if m:
        kid = re.sub(r'\s+', '', m.group(1)).upper()
        return f"Palette pour KIT ILN {kid}"

    # ── NIMP15 (avec dimensions)
    if "nimp" in d_low:
        has_1200 = "1200" in d_low
        has_1000 = "1000" in d_low
        has_800  = "800"  in d_low
        has_600  = "600"  in d_low
        if has_1200 and has_1000:
            return "Palette NIMP15 bois 1200 x 1000mm"
        if has_1200 and has_800:
            return "Palette NIMP15 bois 1200 x 800mm"
        if has_800 and has_600:
            return "Palette NIMP15 bois 800 x 600mm"
        return "Palette Bois NIMP15"

    # ── Plastique
    if "plastique" in d_low:
        return "Palette Plastique"

    # ── Dimensions standard
    has_1200 = "1200" in d_low
    has_1000 = "1000" in d_low
    has_800  = "800"  in d_low
    has_600  = "600"  in d_low
    if has_1200 and has_1000:
        return "Palette 1200 x 1000mm"
    if has_1200 and has_800:
        return "Palette 1200 x 800mm"
    if has_800 and has_600:
        return "Palette 800 x 600mm"

    return "Palette"


def get_strict_type(desig: str, cat: str) -> str:
    """
    Type strict (selon catégorie) pour filtrage exact tableau + courbe.
    Pour sac et cale, on retombe sur le clean type (suffisamment précis).
    Pour 'autre' (composants non classifiés comme film étirable, cerclage,
    ruban, mousse, etc.), on extrait le type par mot-clé.
    """
    if cat == "carton":
        return get_strict_carton_type(desig)
    if cat == "palette":
        return get_strict_palette_type(desig)
    if cat == "film":
        # Sous-types film : bulles / étirable / rétractable / housse
        return get_clean_type_label(desig, "film")
    if cat == "autre":
        return get_strict_autre_type(desig)
    return get_clean_type_label(desig, cat)


def get_strict_autre_type(desig: str) -> str:
    """
    Type strict pour les composants 'autre' — détecte le type par mot-clé.
    Permet de comparer Film Étirable avec Film Étirable, Cerclage avec Cerclage, etc.
    """
    if not isinstance(desig, str):
        return ""
    d = desig.lower().strip()

    # Mots-clés courants pour les composants packaging "autres"
    # On retourne un libellé canonique court qui sert de clé de comparaison
    type_keywords = [
        ("film étirable",   ["film étirable", "film etirable", "film stretch", "stretch film"]),
        ("film polyéthylène",["film polyéthylène", "film polyethylene", "film pe"]),
        ("film bulle",      ["film bulle", "film à bulles", "film a bulles", "bubble wrap"]),
        ("film",            ["film"]),  # générique en dernier
        ("cerclage",        ["cerclage", "feuillard"]),
        ("ruban adhésif",   ["ruban adhésif", "ruban adhesif", "scotch", "adhesive tape"]),
        ("ruban",           ["ruban"]),
        ("mousse",          ["mousse", "foam"]),
        ("intercalaire",    ["intercalaire", "séparateur", "separateur"]),
        ("housse",          ["housse"]),
        ("coiffe",          ["coiffe"]),
        ("couvercle",       ["couvercle", "lid"]),
        ("plateau",         ["plateau", "tray"]),
        ("étiquette",       ["étiquette", "etiquette", "label", "autocollant", "sticker"]),
        ("bande",           ["bande"]),
        ("croisillon",      ["croisillon"]),
        ("anti-corrosion",  ["anti-corrosion", "anticorrosion", "vci"]),
        ("dessiccant",      ["dessiccant", "silica gel"]),
    ]
    for canonical, kws in type_keywords:
        for kw in kws:
            if kw in d:
                return canonical
    # Si rien ne matche → on prend les 2 premiers mots de la désignation
    words = desig.strip().split()
    if len(words) >= 2:
        return f"{words[0]} {words[1]}".lower()
    if words:
        return words[0].lower()
    return ""


def is_label_designation(desig: str) -> bool:
    """Détecte si la désignation correspond à une étiquette / label."""
    if not isinstance(desig, str):
        return False
    d_low = desig.lower()
    return any(k in d_low for k in NON_CARTON_KEYWORDS)


# ─────────────────────────────────────────────────────────────────────────────
# SVG illustrations pour les cartes de type (page Browse-by-type)
# Style flat coloré, professionnel et visuel sans dépendre d'images externes.
# ─────────────────────────────────────────────────────────────────────────────

def _svg_carton(variant: str = "default") -> str:
    """SVG d'un carton (variantes : galia, bouclier, rabats, cloche, default)."""
    if variant == "galia":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="g1" x1="0" y1="0" x2="1" y2="1">
            <stop offset="0%" stop-color="#fbbf24"/><stop offset="100%" stop-color="#d97706"/>
            </linearGradient></defs>
            <path d="M14 24 L40 16 L66 24 L66 60 L40 68 L14 60 Z" fill="url(#g1)" stroke="#92400e" stroke-width="1.5"/>
            <path d="M14 24 L40 32 L66 24" fill="none" stroke="#92400e" stroke-width="1.5"/>
            <path d="M40 32 L40 68" stroke="#92400e" stroke-width="1.5"/>
            <text x="40" y="50" text-anchor="middle" fill="#451a03" font-size="9" font-weight="bold" font-family="DM Sans">GALIA</text>
        </svg>"""
    if variant == "bouclier":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="g2" x1="0" y1="0" x2="0" y2="1">
            <stop offset="0%" stop-color="#fcd34d"/><stop offset="100%" stop-color="#b45309"/>
            </linearGradient></defs>
            <path d="M40 12 L62 22 L62 48 Q62 62 40 70 Q18 62 18 48 L18 22 Z" fill="url(#g2)" stroke="#78350f" stroke-width="1.8"/>
            <path d="M30 38 L37 45 L52 30" fill="none" stroke="#78350f" stroke-width="2.5" stroke-linecap="round"/>
        </svg>"""
    if variant == "cloche":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="g3" x1="0" y1="0" x2="0" y2="1">
            <stop offset="0%" stop-color="#fb923c"/><stop offset="100%" stop-color="#9a3412"/>
            </linearGradient></defs>
            <path d="M22 28 Q22 14 40 14 Q58 14 58 28 L58 60 L22 60 Z" fill="url(#g3)" stroke="#7c2d12" stroke-width="1.5"/>
            <line x1="22" y1="30" x2="58" y2="30" stroke="#7c2d12" stroke-width="1.2"/>
            <circle cx="40" cy="22" r="3" fill="#fde68a"/>
        </svg>"""
    if variant == "rabats":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="g4" x1="0" y1="0" x2="1" y2="1">
            <stop offset="0%" stop-color="#fde047"/><stop offset="100%" stop-color="#ca8a04"/>
            </linearGradient></defs>
            <path d="M14 28 L40 18 L66 28 L66 60 L14 60 Z" fill="url(#g4)" stroke="#713f12" stroke-width="1.5"/>
            <path d="M14 28 L40 38 L66 28" fill="none" stroke="#713f12" stroke-width="1.2"/>
            <path d="M40 18 L40 38" stroke="#713f12" stroke-width="1.2" stroke-dasharray="2,2"/>
            <path d="M22 22 L40 28 L58 22" fill="none" stroke="#713f12" stroke-width="1"/>
        </svg>"""
    # default
    return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
        <defs><linearGradient id="g5" x1="0" y1="0" x2="1" y2="1">
        <stop offset="0%" stop-color="#fbbf24"/><stop offset="100%" stop-color="#b45309"/>
        </linearGradient></defs>
        <rect x="18" y="22" width="44" height="40" rx="2" fill="url(#g5)" stroke="#78350f" stroke-width="1.5"/>
        <line x1="40" y1="22" x2="40" y2="62" stroke="#78350f" stroke-width="1" stroke-dasharray="3,2"/>
        <line x1="18" y1="35" x2="62" y2="35" stroke="#78350f" stroke-width="1"/>
    </svg>"""


def _svg_palette(variant: str = "default") -> str:
    """SVG d'une palette (bois/plastique/NIMP)."""
    if variant == "plastique":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="p1" x1="0" y1="0" x2="0" y2="1">
            <stop offset="0%" stop-color="#60a5fa"/><stop offset="100%" stop-color="#1e40af"/>
            </linearGradient></defs>
            <rect x="10" y="42" width="60" height="14" rx="2" fill="url(#p1)" stroke="#1e3a8a" stroke-width="1.2"/>
            <rect x="10" y="58" width="10" height="10" fill="#1e40af"/>
            <rect x="35" y="58" width="10" height="10" fill="#1e40af"/>
            <rect x="60" y="58" width="10" height="10" fill="#1e40af"/>
            <rect x="14" y="46" width="52" height="2" fill="#3b82f6" opacity="0.5"/>
        </svg>"""
    if variant == "nimp":
        return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
            <defs><linearGradient id="p2" x1="0" y1="0" x2="0" y2="1">
            <stop offset="0%" stop-color="#a3a3a3"/><stop offset="100%" stop-color="#525252"/>
            </linearGradient></defs>
            <rect x="10" y="38" width="60" height="6" fill="url(#p2)" stroke="#262626" stroke-width="1"/>
            <rect x="10" y="46" width="60" height="6" fill="url(#p2)" stroke="#262626" stroke-width="1"/>
            <rect x="10" y="54" width="60" height="14" fill="url(#p2)" stroke="#262626" stroke-width="1"/>
            <text x="40" y="65" text-anchor="middle" fill="#fef3c7" font-size="8" font-weight="bold" font-family="DM Sans">NIMP15</text>
        </svg>"""
    # default bois
    return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
        <defs><linearGradient id="p3" x1="0" y1="0" x2="0" y2="1">
        <stop offset="0%" stop-color="#d97706"/><stop offset="100%" stop-color="#7c2d12"/>
        </linearGradient></defs>
        <rect x="8" y="36" width="64" height="6" fill="url(#p3)" stroke="#451a03" stroke-width="1"/>
        <rect x="8" y="44" width="64" height="6" fill="url(#p3)" stroke="#451a03" stroke-width="1"/>
        <rect x="8" y="54" width="64" height="14" fill="url(#p3)" stroke="#451a03" stroke-width="1"/>
        <line x1="20" y1="58" x2="20" y2="64" stroke="#451a03" stroke-width="1"/>
        <line x1="40" y1="58" x2="40" y2="64" stroke="#451a03" stroke-width="1"/>
        <line x1="60" y1="58" x2="60" y2="64" stroke="#451a03" stroke-width="1"/>
    </svg>"""


def _svg_sac() -> str:
    return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
        <defs><linearGradient id="s1" x1="0" y1="0" x2="0" y2="1">
        <stop offset="0%" stop-color="#a78bfa"/><stop offset="100%" stop-color="#6d28d9"/>
        </linearGradient></defs>
        <path d="M22 24 Q22 16 30 16 L50 16 Q58 16 58 24 L62 68 Q62 70 60 70 L20 70 Q18 70 18 68 Z"
              fill="url(#s1)" stroke="#4c1d95" stroke-width="1.5"/>
        <path d="M22 24 L58 24" stroke="#4c1d95" stroke-width="1.2"/>
        <path d="M30 16 Q30 10 35 10 L45 10 Q50 10 50 16" fill="none" stroke="#4c1d95" stroke-width="1.5"/>
    </svg>"""


def _svg_cale() -> str:
    return """<svg class="t-svg" viewBox="0 0 80 80" xmlns="http://www.w3.org/2000/svg">
        <defs><linearGradient id="c1" x1="0" y1="0" x2="1" y2="1">
        <stop offset="0%" stop-color="#fb7185"/><stop offset="100%" stop-color="#9f1239"/>
        </linearGradient></defs>
        <polygon points="14,60 40,18 66,60" fill="url(#c1)" stroke="#7f1d1d" stroke-width="1.5"/>
        <line x1="14" y1="60" x2="66" y2="60" stroke="#7f1d1d" stroke-width="2"/>
        <line x1="22" y1="50" x2="58" y2="50" stroke="#fee2e2" stroke-width="1" opacity="0.6"/>
        <line x1="30" y1="40" x2="50" y2="40" stroke="#fee2e2" stroke-width="1" opacity="0.6"/>
    </svg>"""


def get_type_svg(type_label: str, cat: str) -> str:
    """Retourne le SVG approprié selon la catégorie + le type strict."""
    t_low = (type_label or "").lower()
    if cat == "carton":
        if "galia" in t_low and ("rabats" in t_low or "g0" in t_low or "g1" in t_low):
            return _svg_carton("rabats")
        if "galia" in t_low and "cloche" in t_low:
            return _svg_carton("cloche")
        if "cloche" in t_low and ("c09" in t_low or "c10" in t_low):
            return _svg_carton("cloche")
        if "bouclier" in t_low:
            return _svg_carton("bouclier")
        if "galia" in t_low:
            return _svg_carton("galia")
        return _svg_carton("default")
    if cat == "palette":
        if "plastique" in t_low:
            return _svg_palette("plastique")
        if "nimp" in t_low:
            return _svg_palette("nimp")
        return _svg_palette("default")
    if cat == "sac":
        return _svg_sac()
    if cat == "cale":
        return _svg_cale()
    return _svg_carton("default")


# ═════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_data(file_source=None, file_name: str = "") -> pd.DataFrame:
    # file_source : fichier uploadé (BytesIO) OU chemin. Si None → DATA_PATH local.
    # file_name : nom d'origine (sert à choisir le bon moteur de lecture Excel).
    src = file_source if file_source is not None else DATA_PATH

    name_for_ext = (file_name or (src if isinstance(src, str) else "")).lower()
    is_old_xls = name_for_ext.endswith(".xls")  # ancien format binaire → xlrd
    is_csv = name_for_ext.endswith((".csv", ".txt"))

    def _rewind():
        try:
            if hasattr(src, "seek"):
                src.seek(0)
        except Exception:
            pass

    # ── Détection du VRAI format par la signature binaire (magic bytes),
    # car l'extension ment parfois (un .xls renommé .xlsx, un CSV renommé, …).
    real_kind = None  # "xlsx" | "xls" | "csv" | None
    try:
        _rewind()
        head = src.read(8) if hasattr(src, "read") else open(src, "rb").read(8)
        _rewind()
        if head[:2] == b"PK":                       # zip → vrai .xlsx/.xlsm
            real_kind = "xlsx"
        elif head[:4] == b"\xD0\xCF\x11\xE0":        # OLE2 → vrai .xls binaire
            real_kind = "xls"
        else:
            real_kind = "csv"                        # probablement texte/CSV
    except Exception:
        real_kind = None

    df = None
    errors = []

    def _try(engine=None, as_csv=False):
        nonlocal df
        try:
            _rewind()
            if as_csv:
                df = pd.read_csv(src, sep=None, engine="python")
            else:
                df = pd.read_excel(src, engine=engine)
            return True
        except Exception as e:
            errors.append(f"{engine or 'csv'}: {e}")
            return False

    # Ordre d'essai guidé par la signature détectée, puis repli sur les autres
    order = []
    if real_kind == "xlsx":   order = [("openpyxl", False)]
    elif real_kind == "xls":  order = [("xlrd", False)]
    elif real_kind == "csv":  order = [(None, True)]
    # Replis génériques (au cas où la signature soit trompeuse)
    for cand in [("openpyxl", False), ("xlrd", False), (None, True)]:
        if cand not in order:
            order.append(cand)

    for engine, as_csv in order:
        if _try(engine, as_csv):
            break

    if df is None:
        raise ValueError(
            "Impossible de lire ce fichier. Vérifiez qu'il s'agit bien d'un "
            "fichier Excel .xlsx valide (et non d'un .xls renommé ou d'un fichier "
            "corrompu). Conseil : ouvrez-le dans Excel puis « Enregistrer sous » "
            "→ format « Classeur Excel (*.xlsx) », puis ré-importez-le.\n"
            f"Détails techniques : {' | '.join(errors[:3])}"
        )

    df.columns = df.columns.str.strip()

    for alt in ["Suppliername","SupplierName","supplier_name"]:
        if alt in df.columns:
            df.rename(columns={alt: COL_SUPPLIER}, inplace=True)
    # Détection des variantes de supplier_account
    for alt in ["SupplierAccount", "Supplier_account", "supplieraccount", "Supplier Account"]:
        if alt in df.columns:
            df.rename(columns={alt: COL_SUPPLIER_ACCOUNT}, inplace=True)

    if COL_SUPPLIER not in df.columns:
        df[COL_SUPPLIER] = "Inconnu"
    df[COL_SUPPLIER] = clean_supplier(df[COL_SUPPLIER])

    # ── Fallback : si suppliername est "Inconnu" / vide, on utilise supplier_account
    # à la place. Permet de toujours identifier le fournisseur même quand le nom manque.
    if COL_SUPPLIER_ACCOUNT in df.columns:
        acct_clean = df[COL_SUPPLIER_ACCOUNT].fillna("").astype(str).str.strip()
        # Là où suppliername est "Inconnu" ou vide, on remplace par supplier_account
        mask_unknown = (
            df[COL_SUPPLIER].isin(["Inconnu", "", "nan", "None", "NaN"])
            | df[COL_SUPPLIER].isna()
        )
        # Mais on ne remplace que si supplier_account contient une valeur réelle
        mask_acct_valid = (acct_clean != "") & (~acct_clean.str.lower().isin(["nan", "none", "inconnu"]))
        df.loc[mask_unknown & mask_acct_valid, COL_SUPPLIER] = (
            "Compte : " + acct_clean[mask_unknown & mask_acct_valid]
        )

    if COL_REF not in df.columns:
        df[COL_REF] = "Sans Réf"
    df[COL_REF] = df[COL_REF].fillna("Sans Réf").astype(str).str.strip()

    if COL_DESIGNATION not in df.columns:
        df[COL_DESIGNATION] = ""
    df[COL_DESIGNATION] = df[COL_DESIGNATION].fillna("").astype(str).str.strip()

    # ── Détection de la colonne FRAIS DE TRANSPORT (shipping_charge)
    for alt in ["shipping_charge","ShippingCharge","shipping charge","frais_transport",
                "shipping_cost","transport_charge","Shipping_charge","SHIPPING_CHARGE",
                "shippingcharge","frais_expedition","transport_cost"]:
        if alt in df.columns and COL_SHIPPING not in df.columns:
            df.rename(columns={alt: COL_SHIPPING}, inplace=True)
    if COL_SHIPPING not in df.columns:
        for col in df.columns:
            cl = str(col).lower().replace(" ", "").replace("_", "")
            if ("shipping" in cl) or ("transport" in cl and ("charge" in cl or "cost" in cl or "frais" in cl)):
                df.rename(columns={col: COL_SHIPPING}, inplace=True)
                break

    # ── Détection des colonnes SITE (manufacturing / dispatch)
    for alt in ["manufacturing_site","ManufacturingSite","manufacturing site",
                "site_fabrication","mfg_site","production_site","site_production"]:
        if alt in df.columns and COL_MFG_SITE not in df.columns:
            df.rename(columns={alt: COL_MFG_SITE}, inplace=True)
    for alt in ["dispatch_site","DispatchSite","dispatch site","site_expedition",
                "site_enlevement","pickup_site","shipping_site"]:
        if alt in df.columns and COL_DISPATCH not in df.columns:
            df.rename(columns={alt: COL_DISPATCH}, inplace=True)

    # ── PAYS du fournisseur — DÉRIVÉ des libellés de site.
    # Les libellés se terminent par le code pays (2 lettres), ex :
    #   "00013498 - 28 - GEODIS ... - LIMAY - FR" → FR
    # On extrait donc le pays depuis manufacturing_site (fabrication) et
    # dispatch_site (expédition) plutôt que d'utiliser des colonnes dédiées.
    if COL_MFG_SITE in df.columns:
        df[COL_COUNTRY_FAB] = df[COL_MFG_SITE].apply(country_from_site)
    if COL_DISPATCH in df.columns:
        df[COL_COUNTRY_DISPATCH] = df[COL_DISPATCH].apply(country_from_site)

    # ── Détection de la colonne DATE DE VALIDATION ÉCONOMIQUE
    for alt in ["Economic_valid_date","EconomicValidDate","economic valid date",
                "economic_validity_date","valid_date","validity_date","date_validite",
                "Economic_Valid_Date","ECONOMIC_VALID_DATE","economic_valid_dt",
                "eco_valid_date","economicvaliddate","economic_validation_date",
                "validation_date","date_validation","Economic validation date"]:
        if alt in df.columns and COL_ECON_DATE not in df.columns:
            df.rename(columns={alt: COL_ECON_DATE}, inplace=True)
    # Recherche élargie : colonne contenant "valid" + "date" (ou "econ"+"date")
    if COL_ECON_DATE not in df.columns:
        for col in df.columns:
            cl = str(col).lower().replace(" ", "").replace("_", "")
            if ("valid" in cl and "date" in cl) or ("econ" in cl and "date" in cl) \
                    or ("validation" in cl):
                df.rename(columns={col: COL_ECON_DATE}, inplace=True)
                break
    # On NE parse PAS ici (le parsing se fait au moment de l'affichage via
    # parse_econ_date, qui gère aussi les numéros de série Excel).

    # ── Détection optionnelle des sites (manufacturing_site / dispatch_site)
    for alt in ["Manufacturing_site","ManufacturingSite","manufacturing site",
                "site_manufacturing","site_fabrication","fabrication_site",
                "manufactring_site","Manufacturing Site","MANUFACTURING_SITE"]:
        if alt in df.columns and COL_MFG_SITE not in df.columns:
            df.rename(columns={alt: COL_MFG_SITE}, inplace=True)
    if COL_MFG_SITE in df.columns:
        df[COL_MFG_SITE] = df[COL_MFG_SITE].fillna("").astype(str).str.strip()

    for alt in ["Dispatch_site","DispatchSite","dispatch site","site_dispatch",
                "site_expedition","expedition_site","Dispatch Site","DISPATCH_SITE",
                "site_enlevement","pickup_site","Pickup Site"]:
        if alt in df.columns and COL_DISPATCH not in df.columns:
            df.rename(columns={alt: COL_DISPATCH}, inplace=True)
    if COL_DISPATCH in df.columns:
        df[COL_DISPATCH] = df[COL_DISPATCH].fillna("").astype(str).str.strip()

    # ── Détection du SITE CLIENT (destination) — pour la comparaison du transport
    for alt in ["Client_site","ClientSite","client site","site_client",
                "Customer_site","CustomerSite","customer_site","customer site",
                "destination_site","Destination_site","delivery_site","Site Client",
                "CLIENT_SITE","client_plant","customer_plant"]:
        if alt in df.columns and COL_CLIENT_SITE not in df.columns:
            df.rename(columns={alt: COL_CLIENT_SITE}, inplace=True)
    # Recherche élargie : toute colonne contenant "client" ou "customer"
    if COL_CLIENT_SITE not in df.columns:
        for col in df.columns:
            cl = str(col).lower()
            if ("client" in cl or "customer" in cl) and ("site" in cl or "plant" in cl):
                df.rename(columns={col: COL_CLIENT_SITE}, inplace=True)
                break
    if COL_CLIENT_SITE in df.columns:
        df[COL_CLIENT_SITE] = df[COL_CLIENT_SITE].fillna("").astype(str).str.strip()

    # Détection devise
    currency_col = None
    candidates = ["currency","Currency","CURRENCY","devise","Devise","DEVISE","cur","CUR",
                  "currency_code","CurrencyCode","CUR_CODE","devise_code"]
    for cand in candidates:
        if cand in df.columns:
            currency_col = cand
            break

    if currency_col is None:
        currency_codes = set(FX_TO_EUR.keys())
        for col in df.columns:
            try:
                vals = df[col].dropna().astype(str).str.strip().str.upper()
            except Exception:
                continue
            if vals.empty:
                continue
            uniq = vals.unique()[:200]
            matches = sum(1 for v in uniq if v in currency_codes)
            if matches >= max(1, int(len(uniq) * 0.25)):
                currency_col = col
                break
            if vals.str.contains(r'[\€\$\£\¥\₹\₩]').any():
                currency_col = col
                break

    if currency_col is None:
        price_like_cols = [c for c in df.columns if 'price' in str(c).lower() or 'prix' in str(c).lower()]
        if price_like_cols:
            code_pattern = r'\b(' + '|'.join(re.escape(c) for c in FX_TO_EUR.keys()) + r')\b'
            sym_pattern = r'([€$£¥₹₩])'
            for pc in price_like_cols:
                s = df[pc].dropna().astype(str)
                if s.str.contains(code_pattern, case=False, regex=True).any():
                    extracted_cur = s.str.extract(code_pattern, expand=False).str.upper()
                    num_str = s.str.replace(code_pattern, '', case=False, regex=True)
                    num_str = num_str.str.replace(r'[^\d\.,-]', '', regex=True).str.replace(',', '.', regex=False)
                    num_vals = pd.to_numeric(num_str, errors='coerce')
                    df[pc] = num_vals
                    df["_currency_from_price"] = extracted_cur.fillna("")
                    currency_col = "_currency_from_price"
                    break
                if s.str.contains(sym_pattern, regex=True).any():
                    extracted_sym = s.str.extract(sym_pattern, expand=False)
                    def _sym_to_code(sym):
                        return {"€":"EUR","$":"USD","£":"GBP","¥":"JPY","₹":"INR","₩":"KRW"}.get(sym, "")
                    extracted_cur = extracted_sym.map(lambda v: _sym_to_code(v))
                    num_str = s.str.replace(sym_pattern, '', regex=True)
                    num_str = num_str.str.replace(r'[^\d\.,-]', '', regex=True).str.replace(',', '.', regex=False)
                    num_vals = pd.to_numeric(num_str, errors='coerce')
                    df[pc] = num_vals
                    df["_currency_from_price"] = extracted_cur.fillna("")
                    currency_col = "_currency_from_price"
                    break

    # ── Détection des lignes ayant une currency EXPLICITE dans la source
    # (= non vide, non "nan"/"None") — utilisé pour l'inférence par prix.
    explicit_curr_mask = pd.Series(False, index=df.index)
    if currency_col:
        raw_curr_clean = df[currency_col].astype(str).str.strip().str.upper()
        explicit_curr_mask = (
            df[currency_col].notna()
            & (raw_curr_clean != "")
            & (~raw_curr_clean.isin(["NAN", "NONE", "NULL"]))
        )

    if currency_col:
        # Première passe : on prend la currency explicite quand elle existe,
        # sinon on laisse vide pour pouvoir tenter l'inférence par prix après.
        df["_currency"] = df[currency_col].astype(str).apply(normalize_currency)
        # Effacer les valeurs non explicites pour faire la place à l'inférence
        df.loc[~explicit_curr_mask, "_currency"] = ""
    else:
        df["_currency"] = ""

    # ── INFÉRENCE INTELLIGENTE : si une ligne n'a pas de currency,
    # on cherche une AUTRE ligne avec exactement le même prix dans la base.
    # Si la currency de cette ligne est connue, on l'attribue à notre ligne.
    # Si ambigu (plusieurs currencies pour le même prix) ou aucun match → UNKNOWN
    # (= "Devise non identifiée"). On ne suppose JAMAIS EUR par défaut.
    if explicit_curr_mask.any() and COL_PRICE in df.columns:
        df_explicit = df[explicit_curr_mask & df[COL_PRICE].notna()].copy()
        # Normaliser la currency
        df_explicit["_norm_cur"] = df_explicit[currency_col].astype(str).apply(normalize_currency)
        price_currency_map = {}
        for price_val, grp in df_explicit.groupby(COL_PRICE):
            try:
                pv = float(price_val)
                if pv <= 0:
                    continue
                vc = grp["_norm_cur"].value_counts()
                if vc.empty:
                    continue
                top_count = int(vc.iloc[0])
                # Si plusieurs currencies à égalité → ambigu, on ne fait rien
                if len(vc) >= 2 and int(vc.iloc[1]) == top_count:
                    continue
                price_currency_map[round(pv, 4)] = vc.index[0]
            except Exception:
                continue
    else:
        price_currency_map = {}

    # Application de l'inférence sur les lignes sans currency explicite
    def _infer_currency_for_row(row):
        idx = row.name
        if explicit_curr_mask.get(idx, False):
            return row["_currency"]
        # Tenter l'inférence par valeur de prix
        try:
            price_val = float(row.get(COL_PRICE))
            if price_val > 0:
                inferred = price_currency_map.get(round(price_val, 4))
                if inferred:
                    return inferred
        except Exception:
            pass
        # Aucune correspondance → devise non identifiée
        return "UNKNOWN"

    df["_currency"] = df.apply(_infer_currency_for_row, axis=1)

    # ── Lecture du fichier d'appoint (Copie de EPDS) — sert à combler :
    #    • la currency manquante / fausse (parfois vide dans EPDS_emballage)
    #    • le total_cost manquant (parfois vide dans EPDS_emballage)
    # Si le fichier n'est pas accessible, on continue sans bloquer.
    cur_map_aux = {}        # ref → currency (depuis le 2e fichier)
    total_cost_map_aux = {} # ref → total_cost (depuis le 2e fichier)
    try:
        import os
        if os.path.exists(DATA_PATH_CURRENCY):
            df_cur = pd.read_excel(DATA_PATH_CURRENCY, engine="openpyxl")
            df_cur.columns = df_cur.columns.str.strip()

            # Détection souple de la colonne référence
            ref_col_alt = None
            for cand in [COL_REF, "Reference_code", "reference_code", "Reference", "reference",
                         "ref", "Ref", "REF"]:
                if cand in df_cur.columns:
                    ref_col_alt = cand; break

            # Détection souple de la colonne currency
            cur_col_alt = None
            for cand in [COL_CURRENCY, "currency", "Currency", "CURRENCY", "devise", "Devise"]:
                if cand in df_cur.columns:
                    cur_col_alt = cand; break

            # Détection souple de la colonne total_cost
            tc_col_alt = None
            for cand in [COL_TOTAL_COST, "total_cost", "TotalCost", "TOTAL_COST",
                         "Total Cost", "totalcost", "cout_total", "Coût Total"]:
                if cand in df_cur.columns:
                    tc_col_alt = cand; break

            if ref_col_alt:
                df_cur[ref_col_alt] = df_cur[ref_col_alt].astype(str).str.strip()

                # Map currency
                if cur_col_alt:
                    df_cur[cur_col_alt] = df_cur[cur_col_alt].fillna("").astype(str).apply(normalize_currency)
                    cur_map_aux = (df_cur[df_cur[cur_col_alt].astype(str).str.strip() != ""]
                                   .drop_duplicates(subset=[ref_col_alt], keep="first")
                                   .set_index(ref_col_alt)[cur_col_alt]
                                   .to_dict())

                # Map total_cost
                if tc_col_alt:
                    df_cur[tc_col_alt] = pd.to_numeric(df_cur[tc_col_alt], errors="coerce")
                    df_tc_valid = df_cur[df_cur[tc_col_alt].notna() & (df_cur[tc_col_alt] > 0)]
                    total_cost_map_aux = (df_tc_valid
                                          .drop_duplicates(subset=[ref_col_alt], keep="first")
                                          .set_index(ref_col_alt)[tc_col_alt]
                                          .to_dict())
    except Exception:
        # Lecture du fichier d'appoint optionnelle — on n'interrompt pas l'app si erreur
        pass

    # Application : currency
    if cur_map_aux:
        def _override_currency(row):
            src = str(row["_currency"] or "").strip().upper()
            looked = cur_map_aux.get(str(row[COL_REF]).strip(), "")
            # On utilise la valeur du 2e fichier UNIQUEMENT si :
            #   • elle est explicite (non vide ET pas UNKNOWN)
            #   • la place actuelle est UNKNOWN ou différente
            if looked and looked != "" and looked != "UNKNOWN":
                if src in ("UNKNOWN", "", "EUR") or src != looked:
                    return looked
            return row["_currency"] if src else "UNKNOWN"
        df["_currency"] = df.apply(_override_currency, axis=1)

    # Application : total_cost (uniquement pour les lignes où le total_cost est manquant ou ≤ 0)
    if total_cost_map_aux:
        if COL_TOTAL_COST not in df.columns:
            df[COL_TOTAL_COST] = np.nan
        df[COL_TOTAL_COST] = pd.to_numeric(df[COL_TOTAL_COST], errors="coerce")
        # On stocke le map dans un attribut accessible par les fonctions d'affichage
        df.attrs["_total_cost_aux_map"] = total_cost_map_aux

        # Remplissage des total_cost manquants depuis la map auxiliaire
        def _fill_total_cost(row):
            cur_val = row.get(COL_TOTAL_COST)
            if pd.notna(cur_val) and cur_val > 0:
                return cur_val
            looked = total_cost_map_aux.get(str(row[COL_REF]).strip())
            if looked is not None and looked > 0:
                return looked
            return cur_val
        df[COL_TOTAL_COST] = df.apply(_fill_total_cost, axis=1)

    num_cols = [COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_PRICE, COL_WEIGHT, COL_THICKNESS,
                COL_PARTS_QTY, COL_PACKAGING_QTY,
                COL_LABOUR, COL_LABOUR_COST, COL_SHIPPING, COL_TOTAL_COST,
                COL_LABOUR_SECONDS, COL_PACKAGING_LEVEL]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        else:
            df[c] = np.nan

    money_cols = [COL_PRICE, COL_LABOUR, COL_LABOUR_COST, COL_SHIPPING, COL_TOTAL_COST]
    for col in money_cols:
        if col not in df.columns:
            continue
        converted_vals = []
        for val, cur in zip(df[col], df["_currency"]):
            if pd.notna(val):
                eur_val, _ = to_eur(float(val), str(cur))
                converted_vals.append(eur_val)
            else:
                converted_vals.append(np.nan)
        df[col] = converted_vals

    df = df[df[COL_DESIGNATION].apply(is_valid_desig)].copy()
    df.reset_index(drop=True, inplace=True)
    return df


@st.cache_data
def get_all_refs(df: pd.DataFrame):
    return sorted(df[COL_REF].dropna().unique().tolist())


def get_ref_components(df: pd.DataFrame, ref: str) -> pd.DataFrame:
    # Matching robuste : strip + uppercase pour absorber les variations de saisie
    # (espaces, casse, tabulations cachées dans un copier-coller depuis Excel)
    ref_clean = str(ref).strip().upper() if ref else ""
    if not ref_clean or COL_REF not in df.columns:
        return df.iloc[0:0].copy()
    # Comparaison normalisée des deux côtés
    db_refs_norm = df[COL_REF].astype(str).str.strip().str.upper()
    rows = df[db_refs_norm == ref_clean].copy()
    if COL_DESIGNATION in rows.columns:
        rows = rows[rows[COL_DESIGNATION].apply(is_valid_desig)].copy()
        rows["_category"] = rows[COL_DESIGNATION].apply(classify_component)
        rows["_clean_type"] = rows.apply(
            lambda r: get_clean_type_label(r[COL_DESIGNATION], r["_category"]), axis=1
        )
    rows.reset_index(drop=True, inplace=True)
    return rows


def get_unit_price(row: pd.Series) -> float:
    """
    Calcule le PRIX UNITAIRE = prix / partsqty packaging.

    Le prix dans la base est parfois un prix total pour N pièces ;
    le prix unitaire (pour 1 pièce) s'obtient en divisant par partsqty packaging.

    Si partsqty est manquant ou ≤ 0, on retourne le prix brut.
    La somme des prix unitaires des composants ≈ total_cost de la référence.
    """
    p = _safe_float(row.get(COL_PRICE))
    if p <= 0:
        return 0.0
    qty = _safe_float(row.get(COL_PARTS_QTY))
    if qty > 0:
        return p / qty
    return p


def _get_display_price(row: pd.Series, cat: str):
    if cat == "mod":
        v = row.get(COL_LABOUR_COST)
        if pd.notna(v) and _safe_float(v) > 0:
            return _safe_float(v)
        v = row.get(COL_LABOUR)
        return _safe_float(v) if (pd.notna(v) and _safe_float(v) > 0) else None
    if cat == "transport":
        v = row.get(COL_SHIPPING)
        return _safe_float(v) if (pd.notna(v) and _safe_float(v) > 0) else None
    # Pour les autres composants : COL_PRICE est déjà le prix UNITAIRE
    # (= prix / partsqty packaging — calculé après chargement de la base)
    v = row.get(COL_PRICE)
    return _safe_float(v) if (pd.notna(v) and _safe_float(v) > 0) else None


def compute_ref_summary(df_ref: pd.DataFrame, df_raw_full: pd.DataFrame, ref: str) -> dict:
    """Synthèse par catégorie pour la référence courante."""
    total_cost_ref = None
    if COL_TOTAL_COST in df_raw_full.columns:
        vals = df_raw_full[df_raw_full[COL_REF] == ref][COL_TOTAL_COST].dropna()
        vals = vals[vals > 0]
        if not vals.empty:
            total_cost_ref = round(float(vals.iloc[0]), 2)

    currencies_used = set()
    if "_currency" in df_raw_full.columns:
        currencies_used = set(df_raw_full[df_raw_full[COL_REF] == ref]["_currency"].dropna().unique())
    had_conversion = bool(currencies_used - {"EUR", ""})

    summary: dict = {
        "_total_cost_ref": total_cost_ref,
        "_had_conversion": had_conversion,
        "_currencies":     currencies_used,
    }

    for cat in ["carton","sac","palette","cale","mod","transport","autre"]:
        if cat == "mod":
            mod_mask = (
                (df_ref[COL_LABOUR_COST].notna() & (df_ref[COL_LABOUR_COST] > 0)) |
                (df_ref[COL_LABOUR].notna() & (df_ref[COL_LABOUR] > 0))
            )
            rows = df_ref[mod_mask].copy()
        elif cat == "transport":
            transport_mask = df_ref[COL_SHIPPING].notna() & (df_ref[COL_SHIPPING] > 0)
            rows = df_ref[transport_mask].copy()
        else:
            rows = df_ref[df_ref["_category"] == cat].copy()

        if rows.empty:
            summary[cat] = {"total": None, "type_label": None, "rows": rows, "first_row": None}
            continue

        # Type label : libellé propre extrait via get_clean_type_label
        type_label = None
        if "_clean_type" in rows.columns:
            valid_types = rows["_clean_type"].dropna()
            valid_types = valid_types[valid_types.astype(str).str.strip() != ""]
            if not valid_types.empty:
                type_label = valid_types.mode().iloc[0]

        # Total des prix
        prices = []
        for _, row in rows.iterrows():
            p = _get_display_price(row, cat)
            if p is not None:
                prices.append(p)

        total = round(sum(prices), 2) if prices else None

        # first_row = ligne correspondant au type principal (cohérence accueil ↔ analyse)
        first_row = None
        if type_label and "_clean_type" in rows.columns:
            rows_typed = rows[rows["_clean_type"] == type_label]
            if not rows_typed.empty:
                first_row = rows_typed.iloc[0]
        if first_row is None and not rows.empty:
            first_row = rows.iloc[0]

        summary[cat] = {
            "total":      total,
            "type_label": type_label,
            "rows":       rows,
            "first_row":  first_row,
        }
    return summary


def _build_component_dict(row: pd.Series, cat: str) -> dict:
    # COL_PRICE est désormais le prix UNITAIRE (déjà divisé par partsqty au chargement).
    # On garde _price_raw pour info / debug si besoin.
    return {
        "designation": str(row.get(COL_DESIGNATION, "—")),
        "category":    cat,
        "price":       _safe_float(row.get(COL_PRICE)),
        "price_raw":   _safe_float(row.get("_price_raw", row.get(COL_PRICE))),
        "parts_qty":   _safe_float(row.get(COL_PARTS_QTY, 1)) if COL_PARTS_QTY in row.index else 1.0,
        "L": _safe_float(row.get(COL_LENGTH)),
        "W": _safe_float(row.get(COL_WIDTH)),
        "H": _safe_float(row.get(COL_HEIGHT)),
        "P": _safe_float(row.get(COL_WEIGHT)),
        "thickness":   _safe_float(row.get(COL_THICKNESS)) if COL_THICKNESS in row.index else 0.0,
        "ref":         str(row.get(COL_REF, "—")),
        "supplier":    str(row.get(COL_SUPPLIER, "—")),
        "labour_rate": _safe_float(row.get(COL_LABOUR, 0)),
        "labour_cost": _safe_float(row.get(COL_LABOUR_COST, 0)),
        "shipping":    _safe_float(row.get(COL_SHIPPING, 0)),
        "currency":    str(row.get("_currency", "EUR")),
        "mfg_site":    str(row.get(COL_MFG_SITE, "")) if COL_MFG_SITE in row.index else "",
        "dispatch":    str(row.get(COL_DISPATCH, "")) if COL_DISPATCH in row.index else "",
        "client_site": str(row.get(COL_CLIENT_SITE, "")) if COL_CLIENT_SITE in row.index else "",
        "country_fab":      str(row.get(COL_COUNTRY_FAB, "")) if COL_COUNTRY_FAB in row.index else "",
        "country_dispatch": str(row.get(COL_COUNTRY_DISPATCH, "")) if COL_COUNTRY_DISPATCH in row.index else "",
    }


# ── Loaders catégories ────────────────────────────────────────────────────────

def _is_carton(desig_series: pd.Series) -> pd.Series:
    """Masque carton STRICT : exclut les étiquettes."""
    d = desig_series.fillna("").astype(str).str.lower()
    is_label = d.apply(lambda x: any(k in x for k in NON_CARTON_KEYWORDS))
    is_carton_kw = d.apply(lambda x: any(k in x for k in CATEGORY_KEYWORDS["carton"]))
    return is_carton_kw & (~is_label)


@st.cache_data
def get_carton_df(df: pd.DataFrame) -> pd.DataFrame:
    mask = _is_carton(df[COL_DESIGNATION])
    d = df[mask].dropna(subset=[COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_PRICE,COL_WEIGHT])
    d = d[(d[COL_PRICE]>0)&(d[COL_PRICE]<500)]
    return add_derived(d)

@st.cache_data
def get_sac_df(df: pd.DataFrame) -> pd.DataFrame:
    mask = df[COL_DESIGNATION].str.lower().apply(
        lambda x: any(k in x for k in CATEGORY_KEYWORDS["sac"]))
    d = df[mask].dropna(subset=[COL_PRICE]).copy()
    d = d[(d[COL_PRICE]>0)&(d[COL_PRICE]<500)]
    for c in [COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_WEIGHT]:
        d[c] = pd.to_numeric(d.get(c, pd.Series(0.,index=d.index)),errors="coerce").fillna(0.)
    return add_derived(d)

@st.cache_data
def get_palette_df(df: pd.DataFrame) -> pd.DataFrame:
    mask = df[COL_DESIGNATION].str.lower().str.contains("palette", na=False)
    d = df[mask].dropna(subset=[COL_PRICE]).copy()
    d = d[(d[COL_PRICE]>0)&(d[COL_PRICE]<5000)]
    for c in [COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_WEIGHT]:
        d[c] = pd.to_numeric(d.get(c, pd.Series(0.,index=d.index)),errors="coerce").fillna(0.)
    return add_derived(d)

@st.cache_data
def get_cale_df(df: pd.DataFrame) -> pd.DataFrame:
    mask = df[COL_DESIGNATION].str.lower().apply(
        lambda x: any(k in x for k in CATEGORY_KEYWORDS["cale"]))
    d = df[mask].dropna(subset=[COL_PRICE]).copy()
    d = d[(d[COL_PRICE]>0)&(d[COL_PRICE]<500)]
    for c in [COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_WEIGHT]:
        d[c] = pd.to_numeric(d.get(c, pd.Series(0.,index=d.index)),errors="coerce").fillna(0.)
    return add_derived(d)

@st.cache_data
def get_film_df(df: pd.DataFrame) -> pd.DataFrame:
    """Pool des films : film bulles, film étirable, film rétractable…"""
    mask = df[COL_DESIGNATION].str.lower().apply(
        lambda x: any(k in x for k in CATEGORY_KEYWORDS["film"]))
    d = df[mask].dropna(subset=[COL_PRICE]).copy()
    d = d[(d[COL_PRICE]>0)&(d[COL_PRICE]<500)]
    for c in [COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_WEIGHT]:
        d[c] = pd.to_numeric(d.get(c, pd.Series(0.,index=d.index)),errors="coerce").fillna(0.)
    return add_derived(d)

@st.cache_data
def get_mod_df(df: pd.DataFrame) -> pd.DataFrame:
    mask = (
        (df[COL_LABOUR_COST].notna() & (df[COL_LABOUR_COST] > 0)) |
        (df[COL_LABOUR].notna()      & (df[COL_LABOUR]      > 0))
    )
    return df[mask].copy()

@st.cache_data
def get_transport_df(df: pd.DataFrame) -> pd.DataFrame:
    if COL_SHIPPING not in df.columns:
        return pd.DataFrame()
    ship = pd.to_numeric(df[COL_SHIPPING], errors="coerce")
    mask = ship.notna() & (ship > 0)
    return df[mask].copy()

@st.cache_data
def get_label_df(df: pd.DataFrame) -> pd.DataFrame:
    """Pool des étiquettes/labels pour comparaison de prix."""
    mask = df[COL_DESIGNATION].fillna("").astype(str).str.lower().apply(
        lambda x: any(k in x for k in NON_CARTON_KEYWORDS))
    d = df[mask].dropna(subset=[COL_PRICE]).copy()
    d = d[(d[COL_PRICE] > 0) & (d[COL_PRICE] < 500)]
    for c in [COL_LENGTH,COL_WIDTH,COL_HEIGHT,COL_WEIGHT]:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0.)
    return d


# ═════════════════════════════════════════════════════════════════════════════
# FILTRAGE STRICT — TYPE EXACT + DIMENSIONS (TOUS FOURNISSEURS)
# ═════════════════════════════════════════════════════════════════════════════

def filter_same_component(pool: pd.DataFrame,
                          desig: str,
                          L: float, W: float, H: float, P: float,
                          cat: str,
                          tol_dim: float = 0.05,
                          tol_weight: float = 0.15,
                          use_weight: bool = False,
                          min_suppliers_for_weight: int = 3) -> pd.DataFrame:
    """
    Filtre principal — TYPE STRICT + DIMENSIONS ±5%.

    Le TYPE STRICT (ex : "Carton à rabats - Type Galia G05", "Palette 1200 x 1000mm",
    "Sac plastique", "Cale carton") est calculé via get_strict_type().

    Étapes :
      1) Filtre type STRICT (si identifié)
      2) Filtre dimensions ±5% (Length, Width, et Height si dispo)
         → si le filtre dim vide le pool, on garde le pool type-only
      3) Filtre poids optionnel si use_weight=True ET ≥ min_suppliers_for_weight
    """
    if pool is None or pool.empty:
        return pool.copy() if pool is not None else pd.DataFrame()

    df = pool.copy()

    # ── 1) Filtre TYPE STRICT (priorité absolue)
    if isinstance(desig, str) and desig.strip() and COL_DESIGNATION in df.columns:
        strict_target = get_strict_type(desig, cat)

        if strict_target:
            mask_strict = df[COL_DESIGNATION].apply(
                lambda x: get_strict_type(str(x), cat) == strict_target
            )
            df_strict = df[mask_strict]

            if not df_strict.empty:
                df = df_strict
            else:
                # Fallback : match sur clean type
                clean_target = get_clean_type_label(desig, cat).lower()
                df_clean = df[df[COL_DESIGNATION].apply(
                    lambda x: get_clean_type_label(str(x), cat).lower() == clean_target
                )]
                if not df_clean.empty:
                    df = df_clean

    # ── 2) Filtre DIMENSIONS ±5% — TOUS fournisseurs préservés
    # On filtre maintenant systématiquement pour que le graphique affiche
    # uniquement les composants de même type ET mêmes dimensions.
    df_dim = df.copy()
    if L > 0 and W > 0 and COL_LENGTH in df_dim.columns and COL_WIDTH in df_dim.columns:
        df_dim = df_dim[
            df_dim[COL_LENGTH].between(L*(1-tol_dim), L*(1+tol_dim)) &
            df_dim[COL_WIDTH].between(W*(1-tol_dim), W*(1+tol_dim))
        ]
        if H > 0 and COL_HEIGHT in df_dim.columns:
            df_dim_H = df_dim[df_dim[COL_HEIGHT].between(H*(1-tol_dim), H*(1+tol_dim))]
            if not df_dim_H.empty:
                df_dim = df_dim_H

    if not df_dim.empty:
        df = df_dim
    # sinon (filtre dim vide) : on garde le pool type-only en sécurité


    # ── 3) Filtre poids optionnel et SOUPLE (jamais activé pour les cas type-stricts
    # car use_weight=False par défaut dans tous les appels carton/sac/cale/palette)
    if use_weight and P > 0 and COL_WEIGHT in df.columns:
        df_w = df[df[COL_WEIGHT].between(P*(1-tol_weight), P*(1+tol_weight))]
        if not df_w.empty and COL_SUPPLIER in df_w.columns:
            n_sup_after = df_w[COL_SUPPLIER].nunique()
            if n_sup_after >= min_suppliers_for_weight:
                df = df_w
        elif not df_w.empty:
            df = df_w

    return df


def filter_exact_dims(pool: pd.DataFrame,
                      desig: str,
                      L: float, W: float, H: float, P: float,
                      cat: str,
                      thickness: float = 0.0,
                      tol: float = 0.001,
                      country_fab: str = "",
                      country_dispatch: str = "") -> pd.DataFrame:
    """
    Filtre EXACT (tolérance 0.1%) sur dimensions + poids + épaisseur — utilisé
    UNIQUEMENT pour les cartes statistiques (prix min / moyen / max), pour
    s'assurer qu'elles comparent des composants strictement identiques.

    Si country_fab / country_dispatch sont fournis, on restreint la comparaison
    aux composants du MÊME PAYS de fabrication et d'expédition (meilleure
    comparabilité — coûts locaux homogènes). Filtre appliqué seulement si la
    colonne existe et ne vide pas le pool.

    NB : on n'utilise PAS cette fonction pour le graphique ni le tableau,
    qui restent sur le filtrage classique (type strict + dims ±5%).
    """
    if pool is None or pool.empty:
        return pool.copy() if pool is not None else pd.DataFrame()

    df = pool.copy()

    # ── Type strict
    if isinstance(desig, str) and desig.strip() and COL_DESIGNATION in df.columns:
        strict_target = get_strict_type(desig, cat)
        if strict_target:
            mask = df[COL_DESIGNATION].apply(
                lambda x: get_strict_type(str(x), cat) == strict_target
            )
            df_strict = df[mask]
            if not df_strict.empty:
                df = df_strict

    # ── Dimensions exactes (tolérance ±0.1%)
    if L > 0 and COL_LENGTH in df.columns:
        df = df[df[COL_LENGTH].between(L*(1-tol), L*(1+tol))]
    if W > 0 and COL_WIDTH in df.columns:
        df = df[df[COL_WIDTH].between(W*(1-tol), W*(1+tol))]
    if H > 0 and COL_HEIGHT in df.columns:
        df_h = df[df[COL_HEIGHT].between(H*(1-tol), H*(1+tol))]
        if not df_h.empty:
            df = df_h

    # ── Poids exact (tolérance 0.5g pour éviter les artefacts d'arrondi)
    if P > 0 and COL_WEIGHT in df.columns:
        weight_tol = max(0.5, P * tol)
        df_w = df[df[COL_WEIGHT].between(P - weight_tol, P + weight_tol)]
        if not df_w.empty:
            df = df_w

    # ── Épaisseur exacte si présente
    if thickness > 0 and COL_THICKNESS in df.columns:
        thk_tol = max(0.01, thickness * tol)
        df_t = df[df[COL_THICKNESS].between(thickness - thk_tol, thickness + thk_tol)]
        if not df_t.empty:
            df = df_t

    # ── Filtre MÊME PAYS (fabrication + expédition) si demandé et disponible
    # On n'applique le filtre que s'il ne vide pas le pool (sinon on garde le
    # pool dimensionnel — mieux vaut un comparatif que rien).
    if country_fab and COL_COUNTRY_FAB in df.columns:
        cf = str(country_fab).strip().upper()
        if cf:
            df_cf = df[df[COL_COUNTRY_FAB].astype(str).str.strip().str.upper() == cf]
            if not df_cf.empty:
                df = df_cf
    if country_dispatch and COL_COUNTRY_DISPATCH in df.columns:
        cd = str(country_dispatch).strip().upper()
        if cd:
            df_cd = df[df[COL_COUNTRY_DISPATCH].astype(str).str.strip().str.upper() == cd]
            if not df_cd.empty:
                df = df_cd

    return df


# ═════════════════════════════════════════════════════════════════════════════
# ML
# ═════════════════════════════════════════════════════════════════════════════

@st.cache_resource
def train_dim_model(data: pd.DataFrame):
    avail = [c for c in FEATURE_COLS_DIM if c in data.columns]
    X = data[avail].fillna(0); y = data[COL_PRICE]
    sc = StandardScaler(); Xs = sc.fit_transform(X)
    m = GradientBoostingRegressor(n_estimators=500,learning_rate=0.05,max_depth=6,
                                  loss="huber",subsample=0.8,random_state=42)
    m.fit(Xs, y)
    return m, sc, avail


def predict_price(L,W,H,Pg,model,scaler,avail):
    surface = 2*(L*W+L*H+W*H); vol = L*W*H+1e-9
    row = {"length packaging":L,"width packaging":W,"height packaging":H,
           "weight packaging":Pg,"Surface_mm2":surface,
           "ratio_LW":L/(W+1e-9),"ratio_LH":L/(H+1e-9),"ratio_WH":W/(H+1e-9),
           "density":Pg/vol,"surface_vol_ratio":surface/vol}
    vec = [[row.get(c,0) for c in avail]]
    return float(model.predict(scaler.transform(vec))[0])


def get_ai_training_pool(designation: str, cat: str) -> pd.DataFrame:
    """
    Retourne les composants du MÊME TYPE STRICT utilisés pour entraîner
    le modèle IA — pour transparence : on affiche à l'utilisateur les
    références qui ont aidé à calculer le prix estimé.
    """
    pool_full = CAT_DFS.get(cat, df_carton)
    if pool_full is None or pool_full.empty or COL_DESIGNATION not in pool_full.columns:
        return pd.DataFrame()
    strict_target = get_strict_type(designation, cat)
    if strict_target:
        mask = pool_full[COL_DESIGNATION].apply(
            lambda x: get_strict_type(str(x), cat) == strict_target
        )
        pool = pool_full[mask].copy()
    else:
        clean_target = get_clean_type_label(designation, cat).lower()
        mask = pool_full[COL_DESIGNATION].apply(
            lambda x: get_clean_type_label(str(x), cat).lower() == clean_target
        )
        pool = pool_full[mask].copy()
    if COL_PRICE in pool.columns:
        pool = pool[pool[COL_PRICE].notna() & (pool[COL_PRICE] > 0)]
    return pool


@st.cache_data(show_spinner=False)
def estimate_price_by_type(designation: str, cat: str,
                           L: float, W: float, H: float, P: float,
                           thickness: float = 0.0) -> tuple[float, str, int]:
    """
    Estimateur ML spécialisé par TYPE de composant.

    Entraîne à la volée un GradientBoostingRegressor sur TOUS les composants
    du MÊME TYPE STRICT (toutes dimensions confondues), et l'utilise pour
    prédire le prix d'un composant avec dimensions L/W/H/P données.

    Retourne (prix_estimé, méthode_utilisée, n_échantillons).
    Méthodes possibles :
      • "ml_type"     → modèle ML entraîné sur ≥10 lignes du type (le mieux)
      • "median_type" → médiane simple du type (si 3-9 lignes : trop peu pour ML)
      • "no_data"     → vraiment rien (impossible d'estimer)
    """
    pool_full = CAT_DFS.get(cat, df_carton)
    if pool_full is None or pool_full.empty or COL_DESIGNATION not in pool_full.columns:
        return 0.0, "no_data", 0

    # On prend toutes les lignes du même TYPE STRICT (toutes dimensions)
    strict_target = get_strict_type(designation, cat)
    if strict_target:
        mask = pool_full[COL_DESIGNATION].apply(
            lambda x: get_strict_type(str(x), cat) == strict_target
        )
        pool = pool_full[mask].copy()
    else:
        clean_target = get_clean_type_label(designation, cat).lower()
        mask = pool_full[COL_DESIGNATION].apply(
            lambda x: get_clean_type_label(str(x), cat).lower() == clean_target
        )
        pool = pool_full[mask].copy()

    if pool.empty:
        return 0.0, "no_data", 0

    # Garder les lignes avec prix valide
    pool = pool[pool[COL_PRICE].notna() & (pool[COL_PRICE] > 0)]
    if pool.empty:
        return 0.0, "no_data", 0

    n = len(pool)

    # Si trop peu de données pour ML, on retourne la médiane
    if n < 10:
        return float(pool[COL_PRICE].median()), "median_type", n

    # Modèle ML : on utilise dimensions + poids + épaisseur comme features
    feat_cols = []
    for col in [COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT, COL_THICKNESS]:
        if col in pool.columns and pool[col].notna().any():
            feat_cols.append(col)

    if not feat_cols:
        return float(pool[COL_PRICE].median()), "median_type", n

    try:
        X = pool[feat_cols].fillna(pool[feat_cols].median(numeric_only=True)).fillna(0).astype(float)
        y = pool[COL_PRICE].astype(float)
        scaler = StandardScaler()
        Xs = scaler.fit_transform(X)

        # Modèle ML : XGBoost (très performant) avec fallback vers GradientBoosting si indisponible
        if _XGBOOST_AVAILABLE:
            model = XGBRegressor(
                n_estimators=400,
                learning_rate=0.05,
                max_depth=5,
                subsample=0.85,
                colsample_bytree=0.9,
                min_child_weight=2,
                reg_alpha=0.05,
                reg_lambda=0.1,
                objective="reg:squarederror",
                random_state=42,
                verbosity=0,
                n_jobs=1,
            )
        else:
            model = GradientBoostingRegressor(
                n_estimators=300, learning_rate=0.05, max_depth=4,
                loss="huber", subsample=0.8, random_state=42
            )
        model.fit(Xs, y)

        # Vecteur d'entrée
        feat_vals = {COL_LENGTH:L, COL_WIDTH:W, COL_HEIGHT:H,
                     COL_WEIGHT:P, COL_THICKNESS:thickness}
        vec = [[feat_vals.get(c, 0.0) for c in feat_cols]]
        predicted = float(model.predict(scaler.transform(vec))[0])

        # ── GARDE-FOU ANTI-EXTRAPOLATION ABERRANTE
        # Le ML peut extrapoler très loin quand les dimensions demandées sortent
        # de la plage d'entraînement, ou quand le pool contient des outliers.
        # On encadre la prédiction dans une fourchette robuste basée sur les prix
        # ── ESTIMATION PAR PLUS PROCHES VOISINS (KNN pondéré)
        # Le ML seul tend à SURESTIMER (extrapolation). On calcule donc une
        # estimation basée sur les composants RÉELLEMENT les plus proches en
        # dimensions/poids, et on s'appuie principalement dessus.
        y_sorted = y.sort_values()
        p05 = float(y_sorted.quantile(0.05))
        p95 = float(y_sorted.quantile(0.95))
        med_type = float(y_sorted.median())

        knn_estimate = None
        try:
            # Distance normalisée sur chaque feature disponible
            target_vals = {COL_LENGTH: L, COL_WIDTH: W, COL_HEIGHT: H,
                           COL_WEIGHT: P, COL_THICKNESS: thickness}
            dist = pd.Series(0.0, index=pool.index)
            n_feat_used = 0
            for col in feat_cols:
                tv = target_vals.get(col, 0.0)
                if tv and tv > 0:
                    base = pool[col].fillna(0).astype(float)
                    scale = max(float(base.median()), 1e-6)
                    dist += ((base - tv).abs() / scale) ** 2
                    n_feat_used += 1
            if n_feat_used > 0:
                dist = dist ** 0.5
                # On prend les 5 voisins les plus proches
                k = min(5, len(pool))
                nearest_idx = dist.nsmallest(k).index
                nb_prices = pool.loc[nearest_idx, COL_PRICE].astype(float)
                nb_dist = dist.loc[nearest_idx]
                # Pondération inverse de la distance (plus proche = plus de poids)
                weights = 1.0 / (nb_dist + 1e-6)
                if weights.sum() > 0:
                    knn_estimate = float((nb_prices * weights).sum() / weights.sum())
        except Exception:
            knn_estimate = None

        if knn_estimate is not None and knn_estimate > 0:
            # On combine : 70% KNN (fiable, ancré sur le réel) + 30% ML
            predicted = 0.7 * knn_estimate + 0.3 * predicted
            # Et on borne autour du KNN (±35%)
            p05 = min(p05, knn_estimate * 0.65)
            p95 = max(knn_estimate * 1.35, p05 + 1e-6)

        # Clip final dans la fourchette robuste
        lo = max(0.0, min(p05, med_type * 0.5))
        hi = max(p95, med_type * 1.5)
        predicted = max(lo, min(predicted, hi))
        predicted = max(0.0, predicted)
        return predicted, "ml_type", n
    except Exception:
        return float(pool[COL_PRICE].median()), "median_type", n


# ═════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION DU PRIX (BASÉE SUR L'ÉCART À LA MÉDIANE)
# ═════════════════════════════════════════════════════════════════════════════

def evaluate_price(price: float, data: pd.DataFrame, ref_col: str = COL_PRICE) -> tuple[str, str]:
    """
    Classification — basée sur la position du prix par rapport à la MÉDIANE du pool.

    Règles claires et symétriques :
      • bon    = prix ≤ médiane                   (au niveau du marché ou en-dessous)
      • moyen  = médiane < prix ≤ médiane × 1.10  (légèrement au-dessus, dans 10%)
      • élevé  = prix > médiane × 1.10            (significativement plus cher)
      • neutre = pas de comparatif disponible

    IMPORTANT — la médiane est calculée sur les prix UNIQUES (distincts).
    Sans cela, des doublons au même prix tirent la médiane et un prix qui est en
    réalité le MAXIMUM peut être classé "bon". Exemple : prix réels [30, 50, 50]
    → médiane brute = 50 → un prix de 50 (le max !) serait "bon". Avec prix uniques
    [30, 50] → médiane = 40 → un prix de 50 est correctement classé "élevé".

    Garde-fou supplémentaire : si le prix analysé est >= au maximum du pool ET
    qu'il existe au moins un prix strictement inférieur, le statut ne peut JAMAIS
    être "bon" (on est, au mieux, "moyen").
    """
    if data is None or data.empty or ref_col not in data.columns:
        return "neutre", "price-neutral"
    prices_all = data[ref_col].dropna()
    prices_all = prices_all[prices_all > 0]
    if len(prices_all) < 2:
        return "neutre", "price-neutral"

    # ── Médiane sur les valeurs UNIQUES (cf. docstring)
    prices_unique = pd.Series(sorted(prices_all.unique()))
    if len(prices_unique) < 2:
        # Toutes les valeurs sont identiques
        n_distinct_refs = data[COL_REF].nunique() if COL_REF in data.columns else 0
        ref_p = float(prices_unique.iloc[0])
        eps_e = max(ref_p * 0.005, 0.01)
        if abs(price - ref_p) <= eps_e:
            # Le prix correspond à l'unique valeur du pool
            return ("bon", "price-good") if n_distinct_refs >= 2 else ("neutre", "price-neutral")
        # Prix différent de l'unique valeur de référence
        return evaluate_price_simple(price, ref_p)

    median = float(prices_unique.median())
    p_min = float(prices_unique.min())
    p_max = float(prices_unique.max())
    if median <= 0:
        return "neutre", "price-neutral"

    eps = max(median * 0.005, 0.01)

    # ── GARDE-FOU : si le prix est au niveau du MAX du pool et qu'il existe
    # un prix strictement plus bas → on NE PEUT PAS être "bon".
    is_at_max = price >= p_max - eps
    has_cheaper = p_min < p_max - eps

    if price <= median + eps:
        if is_at_max and has_cheaper:
            # Le prix est le plus élevé du pool — incohérent avec "bon"
            return "moyen", "price-medium"
        return "bon", "price-good"
    if price <= median * 1.10:
        return "moyen", "price-medium"
    return "élevé", "price-high"


def evaluate_price_simple(price: float, reference: float) -> tuple[str, str]:
    """Classification simple d'un prix par rapport à une valeur de référence unique."""
    if reference <= 0:
        return "neutre", "price-neutral"
    if price <= reference * 1.005:
        return "bon", "price-good"
    if price <= reference * 1.10:
        return "moyen", "price-medium"
    return "élevé", "price-high"


def evaluate_price_force(price: float, data: pd.DataFrame, ref_col: str = COL_PRICE) -> tuple[str, str]:
    """
    Variante de evaluate_price qui ne renvoie JAMAIS 'neutre'.
    Si le pool est insuffisant, on tombe sur la moyenne directement.
    """
    status, color = evaluate_price(price, data, ref_col)
    if status != "neutre":
        return status, color

    if data is None or data.empty or ref_col not in data.columns:
        return "moyen", "price-medium"
    prices = data[ref_col].dropna()
    prices = prices[prices > 0]
    if prices.empty:
        return "moyen", "price-medium"

    median = float(prices.median())
    if median <= 0:
        return "moyen", "price-medium"

    eps = median * 0.005
    if price <= median + eps:
        return "bon", "price-good"
    if price <= median * 1.10:
        return "moyen", "price-medium"
    return "élevé", "price-high"


def kpi_class_for_status(status: str) -> str:
    """Couleur du gros bandeau KPI selon la classification du prix."""
    return {
        "bon":     "kpi-green",
        "moyen":   "kpi-orange",
        "élevé":   "kpi-red",
        "neutre":  "kpi-blue",
    }.get(status, "kpi-blue")


def evaluate_price_vs_estimate(price: float, estimate: float) -> tuple[str, str]:
    """
    Compare le prix réel à une ESTIMATION IA (XGBoost) — utilisé quand aucun
    comparatif réel n'existe dans la base.

    Logique (mêmes seuils que evaluate_price contre la médiane) :
      • bon    = prix ≤ estimation                  (au niveau ou en-dessous de la prédiction IA)
      • moyen  = estimation < prix ≤ estimation × 1.10  (légèrement au-dessus, dans 10%)
      • élevé  = prix > estimation × 1.10           (significativement plus cher que la prédiction)

    Cela permet de donner un statut bon/moyen/élevé même sans comparatif réel,
    en s'appuyant sur la prédiction du modèle IA entraîné sur le type complet.
    """
    if price is None or price <= 0 or estimate is None or estimate <= 0:
        return "neutre", "price-neutral"
    eps = estimate * 0.005  # tolérance 0.5%
    if price <= estimate + eps:
        return "bon", "price-good"
    if price <= estimate * 1.10:
        return "moyen", "price-medium"
    return "élevé", "price-high"


# ═════════════════════════════════════════════════════════════════════════════
# UNIFIED EVALUATION HELPERS — utilisés par accueil + ref_detail + analyse
# Garantissent la cohérence de l'écart affiché partout dans l'app.
# ═════════════════════════════════════════════════════════════════════════════

def _dedup_pool(pool: pd.DataFrame, value_col: str) -> pd.DataFrame:
    """Dédoublonnage standard utilisé par les helpers d'évaluation."""
    if pool is None or pool.empty:
        return pool
    cols = [c for c in [COL_REF, COL_SUPPLIER, COL_DESIGNATION,
                        COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT,
                        COL_MFG_SITE, COL_DISPATCH, value_col]
            if c in pool.columns]
    if cols:
        return pool.drop_duplicates(subset=cols, keep="first")
    return pool


def eval_status_dim(price: float, desig: str, L: float, W: float, H: float,
                    P: float, cat: str, thickness: float = 0.0,
                    country_fab: str = "", country_dispatch: str = "") -> tuple[str, str]:
    """Statut prix — pool = même type strict + MÊMES dimensions + MÊME poids
    + MÊME PAYS (fabrication / expédition) si disponible.

    Logique de retour :
      • Pool vide                            → neutre (ML fallback se chargera)
      • Pool = 1 prix ET current ≈ ce prix   → bon  (validé par 1 source identique)
      • Pool = 1 prix ET current ≠ ce prix   → neutre (1 seule réf, pas comparable)
      • Pool ≥ 2 prix                        → evaluate_price (médiane + seuil 10%)
    """
    if price is None or price <= 0:
        return "neutre", "price-neutral"
    pool_full = CAT_DFS.get(cat, df_carton)
    pool = filter_exact_dims(pool_full, desig, L, W, H, P, cat, thickness=thickness,
                             country_fab=country_fab, country_dispatch=country_dispatch)
    pool = _dedup_pool(pool, COL_PRICE)
    if pool.empty:
        return "neutre", "price-neutral"

    # Vérification : si le pool ne contient qu'UN SEUL prix après dédup ET que le
    # prix actuel correspond à ce prix unique, on évite de passer au ML qui
    # pourrait donner un statut incohérent (par ex. "élevé" alors qu'on est exactement
    # au niveau d'un comparable identique).
    if COL_PRICE in pool.columns:
        prices_unique = pool[COL_PRICE].dropna()
        prices_unique = prices_unique[prices_unique > 0]
        if len(prices_unique) == 1:
            ref_p = float(prices_unique.iloc[0])
            eps = max(ref_p * 0.005, 0.01)
            if abs(price - ref_p) <= eps:
                # Le prix correspond exactement à l'unique comparable
                return "bon", "price-good"
            # Sinon : pas assez de données pour juger
            return "neutre", "price-neutral"

    return evaluate_price(price, pool, COL_PRICE)


def eval_status_label(price: float) -> tuple[str, str]:
    """Statut prix étiquette — pool dédié + force bon/moyen/élevé (jamais neutre)."""
    if price is None or price <= 0:
        return "neutre", "price-neutral"
    pool = _dedup_pool(df_label, COL_PRICE)
    return evaluate_price_force(price, pool, COL_PRICE)


def get_mod_pool(supplier: str = "", mfg_site: str = "") -> pd.DataFrame:
    """
    Pool de comparaison MOD pour déterminer le BEST PRICE (= taux médian du marché).

    Logique en cascade pour TOUJOURS avoir un comparatif pertinent :
      1) Tous les taux MOD du MÊME SITE DE PRODUCTION (tous fournisseurs)
         → c'est la vraie comparaison : "pour ce site, quel est le meilleur taux ?"
      2) Si le site donne moins de 2 taux distincts → tous les taux du fournisseur
      3) Si encore insuffisant → tout le pool MOD (référence marché globale)

    On NE filtre PAS sur le site d'enlèvement (dispatch) : le taux MOD dépend du
    site de PRODUCTION uniquement.
    """
    base = df_mod
    if base is None or base.empty or COL_LABOUR not in base.columns:
        return pd.DataFrame()
    base = base[base[COL_LABOUR].notna() & (base[COL_LABOUR] > 0)]
    if base.empty:
        return pd.DataFrame()

    def _n_distinct(df):
        r = df[COL_LABOUR].dropna(); r = r[r > 0]
        return r.nunique()

    # 1) Même site de production (tous fournisseurs)
    if mfg_site and str(mfg_site).strip() and str(mfg_site).lower() not in ("nan", "none", "") \
            and COL_MFG_SITE in base.columns:
        site_pool = base[base[COL_MFG_SITE].astype(str).str.strip() == str(mfg_site).strip()]
        if _n_distinct(site_pool) >= 2:
            return _dedup_pool(site_pool, COL_LABOUR)

    # 2) Même fournisseur (tous sites)
    if supplier and supplier not in ("Inconnu", "") and COL_SUPPLIER in base.columns:
        sup_pool = base[base[COL_SUPPLIER] == supplier]
        if _n_distinct(sup_pool) >= 2:
            return _dedup_pool(sup_pool, COL_LABOUR)

    # 3) Tout le marché MOD
    return _dedup_pool(base, COL_LABOUR)


def mod_median_rate(supplier: str = "", mfg_site: str = "") -> float:
    """Médiane des taux MOD du pool (valeurs uniques) = BEST PRICE MOD du marché.
    Retourne 0.0 si pas de pool exploitable."""
    pool = get_mod_pool(supplier, mfg_site)
    if pool.empty or COL_LABOUR not in pool.columns:
        return 0.0
    rates = pool[COL_LABOUR].dropna()
    rates = rates[rates > 0]
    if rates.empty:
        return 0.0
    return float(pd.Series(sorted(rates.unique())).median())


def eval_status_mod(rate: float, supplier: str,
                    mfg_site: str = "", dispatch: str = "") -> tuple[str, str]:
    """Statut taux MOD — pool = même fournisseur + même SITE DE PRODUCTION.
    (Le paramètre `dispatch` est conservé pour compatibilité mais N'EST PLUS
    utilisé : le taux MOD ne dépend pas du site d'enlèvement.)"""
    if rate is None or rate <= 0:
        return "neutre", "price-neutral"
    pool = get_mod_pool(supplier, mfg_site)
    if pool.empty:
        return "neutre", "price-neutral"
    return evaluate_price(rate, pool, COL_LABOUR)


def eval_status_transport(shipping: float, supplier: str,
                          pickup_site: str = "", client_site: str = "",
                          parts_qty: float = 0.0) -> tuple[str, str]:
    """Statut frais transport — pool STRICT :
    - même fournisseur
    - même SITE D'ENLÈVEMENT (dispatch_site)
    - même SITE CLIENT (destination)
    - MÊME partsqty packaging
    Un transport ne se compare qu'à un autre transport du même trajet et même volume.
    """
    if shipping is None or shipping <= 0:
        return "neutre", "price-neutral"
    pool = df_transport.copy()
    pool = pool[pool[COL_SHIPPING].notna() & (pool[COL_SHIPPING] > 0)]

    if supplier and supplier not in ("Inconnu", "") and COL_SUPPLIER in pool.columns:
        pool_f = pool[pool[COL_SUPPLIER] == supplier]
        if pool_f.empty:
            return "neutre", "price-neutral"
        pool = pool_f
    # Site d'enlèvement (dispatch_site)
    if pickup_site and str(pickup_site).lower() not in ("", "nan", "none") \
            and COL_DISPATCH in pool.columns:
        pool_f = pool[pool[COL_DISPATCH].astype(str).str.strip() == str(pickup_site).strip()]
        if pool_f.empty:
            return "neutre", "price-neutral"
        pool = pool_f
    # Site client (destination)
    if client_site and str(client_site).lower() not in ("", "nan", "none") \
            and COL_CLIENT_SITE in pool.columns:
        pool_f = pool[pool[COL_CLIENT_SITE].astype(str).str.strip() == str(client_site).strip()]
        if pool_f.empty:
            return "neutre", "price-neutral"
        pool = pool_f
    # Même partsqty packaging
    if parts_qty and parts_qty > 0 and COL_PARTS_QTY in pool.columns:
        tol = max(parts_qty * 0.001, 0.5)
        pool_f = pool[pool[COL_PARTS_QTY].between(parts_qty - tol, parts_qty + tol)]
        if pool_f.empty:
            return "neutre", "price-neutral"
        pool = pool_f

    pool = _dedup_pool(pool, COL_SHIPPING)
    return evaluate_price(shipping, pool, COL_SHIPPING)


# ═════════════════════════════════════════════════════════════════════════════
# CHARTS
# ═════════════════════════════════════════════════════════════════════════════

def build_scatter(data_plot, x_col, y_col, x_label, y_label, title, price, x_target,
                  height=480, show_target: bool = True):
    """Scatter — couleur par fournisseur (légende), aucun nom sur l'abscisse.

    show_target=True (défaut) : affiche le gros point bleu "Point analysé".
    show_target=False : pas de point analysé, juste les données du pool.
    """
    if data_plot is not None and not data_plot.empty and x_col in data_plot.columns:
        dfp = data_plot.dropna(subset=[x_col,y_col]).copy()
        if not dfp.empty:
            fig = px.scatter(
                dfp, x=x_col, y=y_col,
                color=COL_SUPPLIER if COL_SUPPLIER in dfp.columns else None,
                title=f"<b>{title}</b>",
                labels={x_col:x_label, y_col:y_label},
                height=height, opacity=0.75,
                hover_data={
                    COL_SUPPLIER: True, COL_REF: True,
                    x_col: ":.4f", y_col: ":.2f"
                } if COL_SUPPLIER in dfp.columns else None
            )
        else:
            fig = go.Figure()
            fig.update_layout(title=f"<b>{title}</b>",height=height,xaxis_title=x_label,yaxis_title=y_label)
    else:
        fig = go.Figure()
        fig.update_layout(title=f"<b>{title}</b>",height=height,xaxis_title=x_label,yaxis_title=y_label)

    # Point analysé — uniquement si demandé ET valeurs valides
    if show_target and x_target is not None and price is not None and x_target > 0 and price > 0:
        fig.add_trace(go.Scatter(
            x=[x_target], y=[price], mode="markers",
            marker=dict(size=22, color="#4573de", symbol="circle",
                        line=dict(color="white", width=3)),
            name="Point analysé", showlegend=True,
            hovertemplate=f"<b>Point analysé</b><br>{x_label}: {x_target:.4f}<br>Prix: {price:.2f} €<extra></extra>"
        ))
        fig.add_annotation(
            x=x_target, y=price,
            text=f"<b>{price:.2f} €</b>",
            showarrow=True, arrowhead=3, arrowsize=1.3, arrowwidth=2, arrowcolor="#4573de",
            ax=90, ay=-70, bgcolor="white", bordercolor="#4573de", borderwidth=2, borderpad=8,
            font=dict(color="#4573de", size=12)
        )
    fig.update_layout(
        template="plotly_white",
        plot_bgcolor="#f8fafc", paper_bgcolor="white",
        hovermode="closest",
        font=dict(family="DM Sans", size=12),
        legend=dict(title="<b>Fournisseur</b>", orientation="v", x=1.02, y=1),
    )

    # ── Padding axes : assure que les points proches de zéro restent bien visibles
    # (sinon ils se collent à l'axe Y et deviennent invisibles)
    try:
        all_x_vals = []
        all_y_vals = []
        if data_plot is not None and not data_plot.empty:
            if x_col in data_plot.columns:
                all_x_vals = data_plot[x_col].dropna().tolist()
            if y_col in data_plot.columns:
                all_y_vals = data_plot[y_col].dropna().tolist()
        # On inclut le point analysé dans le padding UNIQUEMENT s'il est affiché
        if show_target and x_target is not None and x_target > 0:
            all_x_vals.append(x_target)
        if show_target and price is not None and price > 0:
            all_y_vals.append(price)

        if all_x_vals:
            x_min = min(all_x_vals); x_max = max(all_x_vals)
            x_range = max(x_max - x_min, x_max * 0.1, 1e-6)
            x_pad = x_range * 0.15
            fig.update_xaxes(range=[x_min - x_pad, x_max + x_pad])

        if all_y_vals:
            y_min = min(all_y_vals); y_max = max(all_y_vals)
            y_range = max(y_max - y_min, y_max * 0.1, 1e-6)
            y_pad = y_range * 0.15
            fig.update_yaxes(range=[max(0, y_min - y_pad), y_max + y_pad])
    except Exception:
        pass

    return fig


def _pool_has_dim_variation(df_pool: pd.DataFrame, tol: float = 0.02) -> tuple[bool, bool]:
    """
    Vérifie si le pool a de la variation en VOLUME et en POIDS.
    Tolérance : ±tol (par défaut 2%) pour considérer les valeurs comme identiques.

    Retourne (vol_varies, weight_varies).
    Si les deux sont False → on devrait afficher UN SEUL graphique
    (toutes les lignes ont les mêmes dimensions et le même poids).
    """
    vol_varies = False
    weight_varies = False
    if df_pool is None or df_pool.empty:
        return False, False

    # Volume = L × W × H — on regarde la variation
    if all(c in df_pool.columns for c in [COL_LENGTH, COL_WIDTH, COL_HEIGHT]):
        vols = (df_pool[COL_LENGTH].fillna(0)
                * df_pool[COL_WIDTH].fillna(0)
                * df_pool[COL_HEIGHT].fillna(0))
        vols = vols[vols > 0]
        if not vols.empty:
            v_min = float(vols.min()); v_max = float(vols.max())
            if v_min > 0 and v_max > v_min * (1 + tol):
                vol_varies = True

    if COL_WEIGHT in df_pool.columns:
        wts = df_pool[COL_WEIGHT].dropna()
        wts = wts[wts > 0]
        if not wts.empty:
            w_min = float(wts.min()); w_max = float(wts.max())
            if w_min > 0 and w_max > w_min * (1 + tol):
                weight_varies = True

    return vol_varies, weight_varies


def build_scatter_by_reference(data_plot, y_col, y_label, title,
                                ref_price: float = None, best_price: float = None,
                                target_supplier: str = None,
                                height: int = 480):
    """
    Scatter par fournisseur pour la page d'analyse :
    - Points colorés par fournisseur
    - Légende fournisseur à droite
    - AXE X SANS étiquettes (les noms fournisseurs sont déjà dans la légende)
    - Axe Y = prix
    - Hover : référence + prix
    - FLÈCHE + annotation pour le composant analysé (sa valeur est mise en évidence)
    """
    fig = go.Figure()
    if data_plot is None or data_plot.empty or y_col not in data_plot.columns:
        fig.update_layout(title=f"<b>{title}</b>", height=height,
                          xaxis_title="", yaxis_title=y_label)
        return fig

    dfp = data_plot.dropna(subset=[y_col]).copy()
    if dfp.empty:
        fig.update_layout(title=f"<b>{title}</b>", height=height,
                          xaxis_title="", yaxis_title=y_label)
        return fig

    sup_col = COL_SUPPLIER if COL_SUPPLIER in dfp.columns else None
    ref_col = COL_REF if COL_REF in dfp.columns else None
    target_sup_clean = str(target_supplier).strip() if target_supplier else None
    suppliers_added = []  # pour vérifier que target_supplier existe parmi les traces

    if sup_col:
        for sup_name, grp in dfp.groupby(sup_col):
            n = len(grp)
            x_pos = [str(sup_name)] * n
            suppliers_added.append(str(sup_name))
            if ref_col:
                hover_text = grp[ref_col].astype(str)
            else:
                hover_text = [""] * n
            fig.add_trace(go.Scatter(
                x=x_pos, y=grp[y_col],
                mode="markers",
                name=str(sup_name),
                marker=dict(size=12, opacity=0.8, line=dict(color="white", width=1.5)),
                customdata=hover_text,
                hovertemplate=(f"<b>{sup_name}</b><br>"
                               f"Référence : %{{customdata}}<br>"
                               f"Prix : %{{y:.2f}} €<extra></extra>"),
            ))
    else:
        fig.add_trace(go.Scatter(
            x=[""] * len(dfp), y=dfp[y_col], mode="markers",
            marker=dict(size=12, opacity=0.8, line=dict(color="white", width=1.5)),
            hovertemplate=f"Prix : %{{y:.2f}} €<extra></extra>",
        ))

    # ── INDICATEUR du composant analysé : SIMPLE FLÈCHE pointant vers le point.
    # On ne dessine PAS de marqueur supplémentaire — la flèche annotée suffit
    # à localiser visuellement la valeur sans masquer les points voisins.
    if ref_price is not None and ref_price > 0:
        if target_sup_clean and target_sup_clean in suppliers_added:
            x_target = target_sup_clean
        elif suppliers_added:
            x_target = suppliers_added[0]
        else:
            x_target = ""

        # Flèche fine + petite étiquette déportée à droite
        fig.add_annotation(
            x=x_target, y=ref_price,
            text=f"<b>← {ref_price:.2f} €</b><br><span style='font-size:9px;'>composant analysé</span>",
            showarrow=True,
            arrowhead=2, arrowsize=1.1, arrowwidth=2, arrowcolor="#dc2626",
            ax=70, ay=-10,                          # étiquette légèrement haut-droite
            font=dict(size=11, color="#7f1d1d", family="DM Sans"),
            bgcolor="rgba(254,242,242,0.85)",        # semi-transparent
            bordercolor="#dc2626", borderwidth=1.2, borderpad=4,
            xanchor="left", yanchor="middle",
            opacity=0.95,
        )

    # ── Échelle Y adaptative — points près de 0 plus visibles
    # On NE force PAS l'axe à toucher 0. On étire la plage pour mettre
    # en valeur les petites variations entre points proches de 0.
    try:
        y_vals = dfp[y_col].dropna()
        y_vals = y_vals[y_vals > 0]
        if not y_vals.empty:
            y_min = float(y_vals.min())
            y_max = float(y_vals.max())
            if ref_price is not None and ref_price > 0:
                y_min = min(y_min, ref_price)
                y_max = max(y_max, ref_price)
            spread = y_max - y_min
            if spread < 1e-9:
                # Toutes les valeurs identiques → on simule une plage de ±15 %
                pad = max(y_max * 0.15, 0.01)
                y_axis_min = max(0.001, y_min - pad)
                y_axis_max = y_max + pad
            else:
                # Padding = max(30% du spread, 30% de y_min)
                # → garantit qu'un y_min très bas obtient quand même un padding visible
                pad_down = max(spread * 0.30, y_min * 0.30)
                pad_up = max(spread * 0.20, y_max * 0.10)
                y_axis_min = max(0.001, y_min - pad_down)
                # Si y_min est très bas, on force l'axe à ne pas inclure 0
                if y_min > 0:
                    y_axis_min = max(y_axis_min, y_min * 0.5)
                y_axis_max = y_max + pad_up
            fig.update_yaxes(range=[y_axis_min, y_axis_max])
    except Exception:
        pass

    fig.update_layout(
        title=f"<b>{title}</b>",
        xaxis_title="<b>Fournisseur</b>",  # Titre clair pour éviter la confusion
        yaxis_title=y_label,
        height=height,
        template="plotly_white",
        plot_bgcolor="#f8fafc", paper_bgcolor="white",
        font=dict(family="DM Sans", size=12),
        legend=dict(title="<b>Fournisseur</b>"),
        # Étiquettes de l'axe X masquées (les fournisseurs sont dans la légende)
        # mais on garde le TITRE de l'axe pour la clarté.
        xaxis=dict(showticklabels=False, showgrid=False, zeroline=False),
    )
    return fig


# ═════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def kpi_block(price: float, label: str, is_real: bool,
              kpi_class: str = "kpi-blue", show_badge: bool = True,
              status: str = None, median_val: float = None,
              gain_text: str = None):
    """KPI : prix + statut bon/moyen/élevé selon écart à la médiane.
    median_val : si fourni et statut élevé/moyen → affiche l'écart chiffré.
    gain_text  : texte de gain potentiel déjà formaté (ex. '12.34 €')."""
    if not is_real and show_badge:
        badge_html = '<span class="badge-pos badge-orange">Estimation IA</span>'
    else:
        badge_html = ""

    status_html = ""
    if status:
        emoji_map = {"bon":"🟢","moyen":"🟡","élevé":"🔴","neutre":"⚪"}
        label_map = {"bon":"Écart faible","moyen":"Écart moyen","élevé":"Écart élevé",
                     "neutre":"Pas de comparatif"}
        status_html = (f'<div class="kpi-status">{emoji_map.get(status,"")} '
                       f'{label_map.get(status,"Prix " + str(status))}</div>')

    st.markdown(
        f'<div class="{kpi_class}">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{price:.2f} €</div>'
        f'{status_html}'
        f'{badge_html}'
        f'</div>',
        unsafe_allow_html=True
    )


def result_card(ref, supplier, price, is_real):
    ref_d = ref if (ref and is_real) else "Non trouvée"
    sup_d = supplier if is_real else "—"
    lbl   = "Correspondance Exacte" if is_real else "Estimation IA"
    return (f'<div class="result-info-card"><h4>{lbl}</h4>'
            f'<div class="result-info-row">'
            f'<div><div class="ri-label">Référence</div><div class="ri-value">{ref_d}</div></div>'
            f'<div><div class="ri-label">Fournisseur</div><div class="ri-value">{sup_d}</div></div>'
            f'<div><div class="ri-label">Prix (€)</div><div class="ri-value">{price:.2f} €</div></div>'
            f'</div></div>')


def stat_cards_from_df(df_source: pd.DataFrame, price_col: str = COL_PRICE,
                       label: str = "Pool comparé"):
    """
    Calcule min / MÉDIANE / max sur df_source — exactement les mêmes données que
    celles affichées dans le tableau et la courbe. GARANTIT la cohérence.

    Pourquoi la MÉDIANE plutôt que la moyenne :
    - Plus robuste aux valeurs aberrantes (commandes urgentes, MOQ exceptionnels,
      conversions de devise défavorables qui tirent la moyenne vers le haut)
    - Reflète mieux le « prix de marché réel » que paie la majorité
    - C'est aussi la statistique utilisée pour évaluer l'écart bon/moyen/élevé,
      donc l'affichage et le calcul du statut sont cohérents.
    """
    if df_source is None or df_source.empty or price_col not in df_source.columns:
        c1,c2,c3 = st.columns(3, gap="medium")
        for col,title in [(c1,"Prix Minimum"),(c2,"Prix Médiane"),(c3,"Prix Maximum")]:
            with col:
                st.markdown(f'<div class="stat-card"><h4>{title}</h4>'
                            f'<p class="stat-value">—</p>'
                            f'<p class="sub">{label}</p></div>', unsafe_allow_html=True)
        return

    prices = pd.to_numeric(df_source[price_col], errors="coerce").dropna()
    prices = prices[prices > 0]
    if prices.empty:
        c1,c2,c3 = st.columns(3, gap="medium")
        for col,title in [(c1,"Prix Minimum"),(c2,"Prix Médiane"),(c3,"Prix Maximum")]:
            with col:
                st.markdown(f'<div class="stat-card"><h4>{title}</h4>'
                            f'<p class="stat-value">—</p>'
                            f'<p class="sub">{label}</p></div>', unsafe_allow_html=True)
        return

    mn, med, mx = prices.min(), prices.median(), prices.max()

    # ── Date de VALIDATION économique (economic_valid_date = moment où le
    # document a été validé). On affiche, sous CHAQUE valeur, la date :
    #   - Minimum  → date de la ligne au prix minimum
    #   - Maximum  → date de la ligne au prix maximum
    #   - Médiane  → date de validation la PLUS RÉCENTE du pool
    def _date_for_price(target):
        try:
            if COL_ECON_DATE not in df_source.columns:
                return ""
            sub = df_source.copy()
            sub["_p"] = pd.to_numeric(sub[price_col], errors="coerce")
            sub = sub[sub["_p"] > 0]
            sub["_diff"] = (sub["_p"] - target).abs()
            sub = sub.sort_values("_diff")
            if sub.empty:
                return ""
            d = parse_econ_date(sub.iloc[0][COL_ECON_DATE])
            return d.strftime("%d/%m/%Y") if pd.notna(d) else ""
        except Exception:
            return ""

    # Médiane → date de validation la plus récente du pool
    date_med = ""
    try:
        if COL_ECON_DATE in df_source.columns:
            dates = parse_econ_date(df_source[COL_ECON_DATE]).dropna()
            if not dates.empty:
                date_med = dates.max().strftime("%d/%m/%Y")
    except Exception:
        date_med = ""

    date_min = _date_for_price(mn)
    date_max = _date_for_price(mx)

    def _sub(date_str):
        # Date de validation — clairement visible (📅 validé jj/mm/aaaa)
        if date_str:
            return (f'<div style="font-size:12px;color:#0c4a6e;font-weight:700;'
                    f'margin-top:4px;">📅 validé {date_str}</div>')
        return '<div style="font-size:11px;color:#94a3b8;margin-top:4px;">📅 date n/d</div>'

    # Référence associée à un prix (min/max) — pour la traçabilité
    def _ref_for_price(target):
        try:
            if COL_REF not in df_source.columns:
                return ""
            sub = df_source.copy()
            sub["_p"] = pd.to_numeric(sub[price_col], errors="coerce")
            sub = sub[sub["_p"] > 0]
            sub["_diff"] = (sub["_p"] - target).abs()
            sub = sub.sort_values("_diff")
            if sub.empty:
                return ""
            return str(sub.iloc[0][COL_REF])
        except Exception:
            return ""

    def _ref_badge(ref_str):
        if not ref_str:
            return ""
        return (f'<div style="margin-top:3px;font-size:11px;color:#475569;'
                f'background:#f1f5f9;border-radius:5px;padding:2px 6px;'
                f'display:inline-block;font-weight:600;">🔖 {ref_str}</div>')

    ref_min = _ref_for_price(mn)
    ref_max = _ref_for_price(mx)
    # Pour la médiane : référence dont le prix est le plus proche de la médiane
    ref_med = _ref_for_price(med)

    c1,c2,c3 = st.columns(3, gap="medium")
    cards = [
        (c1, mn,  "Prix Minimum", _sub(date_min), _ref_badge(ref_min)),
        (c2, med, "Prix Médiane", _sub(date_med), _ref_badge(ref_med)),
        (c3, mx,  "Prix Maximum", _sub(date_max), _ref_badge(ref_max)),
    ]
    for col,val,title,sub,refb in cards:
        with col:
            st.markdown(f'<div class="stat-card"><h4>{title}</h4>'
                        f'<p class="stat-value">{val:.2f} €</p>'
                        f'{sub}'
                        f'{refb}</div>', unsafe_allow_html=True)


def show_table(df, title, as_df=False, table_type=None):
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="table-section"><h3>{title}</h3>', unsafe_allow_html=True)
    if isinstance(df, pd.DataFrame) and not df.empty:
        if table_type == "mod":
            cols = [c for c in [COL_SUPPLIER, COL_REF, COL_LABOUR] if c in df.columns]
        elif table_type == "transport":
            cols = [c for c in [COL_SUPPLIER, COL_REF, COL_SHIPPING] if c in df.columns]
        else:
            # Tableau composant : on ajoute site de production, site d'enlèvement,
            # pays, code matière et date de validation pour la TRAÇABILITÉ.
            cols = [c for c in [COL_REF, COL_SUPPLIER, COL_DESIGNATION,
                                 COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT,
                                 COL_THICKNESS, COL_MATERIAL, COL_PRICE,
                                 COL_MFG_SITE, COL_DISPATCH,
                                 COL_COUNTRY_FAB, COL_COUNTRY_DISPATCH,
                                 COL_ECON_DATE]
                    if c in df.columns]

        # Sélection des colonnes
        if not as_df and cols:
            df_display = df[cols].copy()
        else:
            df_display = df.copy()

        # ── Forçage du typage : essentiel pour que le tri (clic sur l'en-tête)
        # fonctionne correctement. Sans ça, Streamlit interprète parfois les
        # colonnes mixtes comme du texte → tri alphabétique au lieu de numérique
        # (ex: "100.50" vient avant "99.00").
        numeric_cols = [COL_PRICE, COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT,
                        COL_THICKNESS, COL_LABOUR, COL_LABOUR_COST, COL_SHIPPING,
                        COL_TOTAL_COST]
        text_cols = [COL_REF, COL_SUPPLIER, COL_DESIGNATION, COL_MFG_SITE, COL_DISPATCH,
                     COL_COUNTRY_FAB, COL_COUNTRY_DISPATCH, COL_ECON_DATE, COL_MATERIAL]

        col_config = {}
        for c in df_display.columns:
            if c in numeric_cols:
                # Conversion stricte : tout ce qui n'est pas un nombre devient NaN
                df_display[c] = pd.to_numeric(df_display[c], errors="coerce")
                # Configuration explicite : Streamlit traite la colonne comme un nombre.
                # Format SANS unité dans la cellule — l'unité est déjà dans l'en-tête.
                if c == COL_PRICE:
                    col_config[c] = st.column_config.NumberColumn(
                        "Prix unitaire (€)", format="%.2f")
                elif c == COL_LABOUR:
                    col_config[c] = st.column_config.NumberColumn(
                        "Taux horaire (€/h)", format="%.2f")
                elif c == COL_SHIPPING:
                    col_config[c] = st.column_config.NumberColumn(
                        "Frais (€)", format="%.2f")
                elif c == COL_LENGTH:
                    col_config[c] = st.column_config.NumberColumn(
                        "Longueur (mm)", format="%g")
                elif c == COL_WIDTH:
                    col_config[c] = st.column_config.NumberColumn(
                        "Largeur (mm)", format="%g")
                elif c == COL_HEIGHT:
                    col_config[c] = st.column_config.NumberColumn(
                        "Hauteur (mm)", format="%g")
                elif c == COL_WEIGHT:
                    col_config[c] = st.column_config.NumberColumn(
                        "Poids (g)", format="%g")
                elif c == COL_THICKNESS:
                    col_config[c] = st.column_config.NumberColumn(
                        "Épaisseur (mm)", format="%g")
                elif c == COL_LABOUR_COST:
                    col_config[c] = st.column_config.NumberColumn(
                        "Coût MOD (€)", format="%.2f")
                elif c == COL_TOTAL_COST:
                    col_config[c] = st.column_config.NumberColumn(
                        "Coût total (€)", format="%.2f")
                else:
                    col_config[c] = st.column_config.NumberColumn(c, format="%.2f")
            elif c in text_cols:
                if c == COL_COUNTRY_FAB or c == COL_COUNTRY_DISPATCH:
                    # Pays : code ISO → nom complet (ES → Espagne)
                    df_display[c] = df_display[c].fillna("").astype(str).apply(country_full_name)
                    df_display[c] = df_display[c].replace("", "—")
                    label_pays = "Pays de fabrication" if c == COL_COUNTRY_FAB else "Pays d'expédition"
                    col_config[c] = st.column_config.TextColumn(label_pays)
                    continue
                if c == COL_ECON_DATE:
                    # Date de validation → format MM/AAAA
                    df_display[c] = parse_econ_date(df_display[c]).dt.strftime("%d/%m/%Y")
                    df_display[c] = df_display[c].fillna("—")
                    col_config[c] = st.column_config.TextColumn("Date validation")
                    continue
                df_display[c] = df_display[c].fillna("—").astype(str)
                if c == COL_REF:
                    col_config[c] = st.column_config.TextColumn("Référence", pinned=True)
                elif c == COL_SUPPLIER:
                    col_config[c] = st.column_config.TextColumn("Fournisseur")
                elif c == COL_DESIGNATION:
                    col_config[c] = st.column_config.TextColumn("Désignation")
                elif c == COL_MFG_SITE:
                    col_config[c] = st.column_config.TextColumn("Site de production")
                elif c == COL_DISPATCH:
                    col_config[c] = st.column_config.TextColumn("Site d'enlèvement")
                elif c == COL_MATERIAL:
                    col_config[c] = st.column_config.TextColumn("Matière")

        st.dataframe(df_display, use_container_width=True, hide_index=True,
                     column_config=col_config)
    else:
        st.markdown('<div class="info-box">Aucune donnée à afficher.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


def render_info_panel(items_text: list, items_num: list, panel_title: str = "Informations du Composant"):
    """
    Affiche un panneau d'informations PROPREMENT ALIGNÉ et stylé.

    items_text : liste de tuples (label, value_str)  → champs textes courts (Type, Fournisseur, Sites)
    items_num  : liste de tuples (label, value_str, unit)  → champs numériques avec unité

    Les items "principaux" (Prix unitaire, Coût MOD, Frais transport) reçoivent
    automatiquement la classe .main (accent vert).
    """
    parts = [f'<div class="comp-info-panel"><h3>{panel_title}</h3><div class="info-grid">']

    # Champs textes (Type, Fournisseur, Sites, etc.) — classe .txt (accent cyan)
    for label, value in items_text:
        safe_val = str(value) if value not in (None, "") else "—"
        parts.append(
            f'<div class="info-item txt">'
            f'<div class="il">{label}</div>'
            f'<div class="iv-text">{safe_val}</div>'
            f'</div>'
        )

    # Champs numériques avec unité — items "principaux" reçoivent classe .main
    # On détecte le label "Prix pour N pièces" via startswith car N varie.
    main_labels = {"Coût MOD", "Frais transport", "Taux horaire", "Prix"}
    for label, value, unit in items_num:
        safe_val = str(value) if value not in (None, "") else "—"
        unit_html = f'<span class="iu">{unit}</span>' if unit else ""
        is_main = label in main_labels or label.startswith("Prix pour ")
        item_cls = "info-item main" if is_main else "info-item"
        parts.append(
            f'<div class="{item_cls}">'
            f'<div class="il">{label}</div>'
            f'<div class="iv">{safe_val}{unit_html}</div>'
            f'</div>'
        )

    parts.append("</div></div>")
    st.markdown("".join(parts), unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# CHARGEMENT GLOBAL — avec import du fichier par l'utilisateur
# ═════════════════════════════════════════════════════════════════════════════

# L'app peut fonctionner de deux façons :
#  1) Fichier importé par l'utilisateur (bouton ci-dessous) — RECOMMANDÉ pour
#     le partage en équipe : personne n'a besoin de modifier le code.
#  2) Fichier local par défaut (DATA_PATH) — repli si aucun import.
import os as _os

@st.cache_data(show_spinner="Chargement du fichier…")
def _load_from_bytes(file_bytes: bytes, file_name: str = ""):
    import io as _io
    return load_data(_io.BytesIO(file_bytes), file_name=file_name)

def _show_upload_gate():
    """Écran d'accueil d'import du fichier EPDS — épuré et élégant."""

    st.markdown("""
    <style>
      /* Fond doux plein écran */
      .stApp { background: radial-gradient(1200px 600px at 50% -10%, #eef2ff 0%, #f8fafc 45%, #ffffff 100%); }
      .upl-wrap { max-width: 540px; margin: 8vh auto 0 auto; text-align: center; }
      .upl-mark {
        width: 84px; height: 84px; margin: 0 auto 22px auto; border-radius: 24px;
        background: linear-gradient(145deg,#1e293b 0%,#334155 100%);
        display:flex; align-items:center; justify-content:center; font-size:42px;
        box-shadow: 0 18px 40px rgba(15,23,42,0.22);
      }
      .upl-title { font-size: 38px; font-weight: 900; letter-spacing:-1.2px;
                   color:#0f172a; margin:0 0 10px 0; }
      .upl-sub { font-size: 15px; color:#64748b; margin:0 0 4px 0; font-weight:500; }
      .upl-cta { font-size: 13px; color:#94a3b8; margin: 26px 0 10px 0; font-weight:600;
                 letter-spacing:.3px; }
      /* Zone d'upload épurée */
      div[data-testid="stFileUploader"] section {
        border: 2px dashed #cbd5e1 !important; border-radius: 18px !important;
        background: #ffffff !important; padding: 30px !important;
        transition: all .2s ease;
      }
      div[data-testid="stFileUploader"] section:hover {
        border-color: #6366f1 !important; background:#f5f3ff !important;
        transform: translateY(-1px);
      }
      .upl-foot { font-size:11px; color:#cbd5e1; margin-top:26px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        '<div class="upl-wrap">'
        '<div class="upl-mark">📦</div>'
        '<h1 class="upl-title">PricePackaging</h1>'
        '<p class="upl-sub">Analyse des prix d\'emballage — Renault Group</p>'
        '<div class="upl-cta">Importez votre fichier EPDS pour commencer</div>'
        '</div>',
        unsafe_allow_html=True
    )

    _, cmid, _ = st.columns([1, 2, 1])
    with cmid:
        up = st.file_uploader("Fichier EPDS", type=["xlsx", "xls"],
                              key="epds_uploader", label_visibility="collapsed",
                              help="Fichiers volumineux acceptés.")
        if up is not None:
            st.session_state["_epds_bytes"] = up.getvalue()
            st.session_state["_epds_name"] = up.name
            st.rerun()
        if _os.path.exists(DATA_PATH):
            st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
            if st.button("Utiliser le fichier local par défaut",
                         use_container_width=True):
                st.session_state["_epds_bytes"] = "LOCAL"
                st.rerun()

    st.markdown(
        '<div class="upl-wrap" style="margin-top:0;">'
        '<div class="upl-foot">Vos données restent sur votre poste · Outil interne Renault Group</div>'
        '</div>',
        unsafe_allow_html=True
    )

# Détermine la source du fichier
_src = st.session_state.get("_epds_bytes", None)
if _src is None:
    # Aucun fichier encore importé → on affiche l'écran d'import et on arrête.
    _show_upload_gate()
    st.stop()

@st.cache_data(show_spinner="Chargement du fichier…")
def _load_local():
    return load_data()

# Chargement effectif
if _src == "LOCAL":
    df_raw = _load_local()
else:
    df_raw = _load_from_bytes(_src, st.session_state.get("_epds_name", ""))

# Bandeau discret avec le nom du fichier chargé + bouton pour changer
_fname = st.session_state.get("_epds_name", "fichier local")
_bc1, _bc2 = st.columns([5, 1])
with _bc1:
    st.caption(f"📄 Fichier chargé : **{_fname}**")
with _bc2:
    if st.button("Changer", key="change_file", use_container_width=True):
        for _k in ("_epds_bytes", "_epds_name"):
            st.session_state.pop(_k, None)
        st.rerun()

# ── Prix utilisés dans l'app
# • COL_PRICE reste le PRIX BRUT du fichier Excel (= unitprice packaging)
#   → utilisé pour l'analyse, le best price, les stats, les comparaisons.
# • _unit_price = (unitprice packaging × packagingqty packaging) / partsqty packaging
#   → utilisé UNIQUEMENT pour l'affichage des cartes sur la page d'accueil
#   (somme des _unit_price ≈ total_cost de la référence).
df_raw["_price_raw"] = df_raw[COL_PRICE].copy()
price_num = pd.to_numeric(df_raw[COL_PRICE], errors="coerce")
if COL_PARTS_QTY in df_raw.columns:
    parts_q = pd.to_numeric(df_raw[COL_PARTS_QTY], errors="coerce")
else:
    parts_q = pd.Series(1.0, index=df_raw.index)
if COL_PACKAGING_QTY in df_raw.columns:
    pack_q = pd.to_numeric(df_raw[COL_PACKAGING_QTY], errors="coerce").fillna(1.0)
else:
    pack_q = pd.Series(1.0, index=df_raw.index)
df_raw["_unit_price"] = (
    (price_num * pack_q) / parts_q.where(parts_q.notna() & (parts_q > 0), 1.0)
)

df_carton    = get_carton_df(df_raw)
df_sac       = get_sac_df(df_raw)
df_palette   = get_palette_df(df_raw)
df_cale      = get_cale_df(df_raw)
df_film      = get_film_df(df_raw)
df_mod       = get_mod_df(df_raw)
df_transport = get_transport_df(df_raw)
df_label     = get_label_df(df_raw)
all_refs     = get_all_refs(df_raw)

CAT_DFS = {
    "carton":df_carton,"sac":df_sac,"palette":df_palette,
    "cale":df_cale,"film":df_film,"mod":df_mod,"transport":df_transport,
    # Pool "autre" : on utilise df_raw filtré sur les composants packaging valides
    # (pas seulement les étiquettes). Permet de comparer Film Étirable avec Film Étirable,
    # Cerclage avec Cerclage, etc. — le filtrage par TYPE STRICT (get_strict_autre_type)
    # se charge ensuite de ne retenir que les composants de même type.
    "autre": df_raw[df_raw[COL_DESIGNATION].notna() &
                    (pd.to_numeric(df_raw[COL_PRICE], errors="coerce") > 0)].copy(),
}


@st.cache_data
def get_types_for_category(cat: str) -> list:
    """Retourne la liste triée des types stricts présents pour une catégorie."""
    pool = CAT_DFS.get(cat)
    if pool is None or pool.empty or COL_DESIGNATION not in pool.columns:
        return []
    types_count = {}
    for desig in pool[COL_DESIGNATION].dropna().astype(str):
        t = get_strict_type(desig, cat)
        if not t:
            t = get_clean_type_label(desig, cat)
        if t and t.strip() and t.lower() not in ("autre", "—", ""):
            types_count[t] = types_count.get(t, 0) + 1
    # Filtrage : on retire les types avec moins de 2 occurrences (bruit)
    items = [(t, n) for t, n in types_count.items() if n >= 2]
    # Tri : par nombre d'occurrences décroissant puis nom
    items.sort(key=lambda x: (-x[1], x[0]))
    return items  # liste de tuples (type, count)


def get_pool_for_type(cat: str, type_label: str) -> pd.DataFrame:
    """Retourne toutes les lignes de la catégorie ayant ce type strict."""
    pool = CAT_DFS.get(cat)
    if pool is None or pool.empty or COL_DESIGNATION not in pool.columns:
        return pd.DataFrame()
    mask = pool[COL_DESIGNATION].apply(
        lambda x: get_strict_type(str(x), cat) == type_label or
                  get_clean_type_label(str(x), cat) == type_label
    )
    df_t = pool[mask].copy()
    # Dédoublonnage
    dedup_cols = [c for c in [COL_REF, COL_SUPPLIER, COL_DESIGNATION,
                              COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT, COL_PRICE]
                  if c in df_t.columns]
    if dedup_cols:
        df_t = df_t.drop_duplicates(subset=dedup_cols, keep="first").reset_index(drop=True)
    return df_t


def type_has_fixed_dimensions(df_pool: pd.DataFrame) -> tuple[bool, dict]:
    """
    Détermine quelles dimensions L/W/H sont IDENTIQUES sur tout le pool
    (tolérance ±2%). Retourne (is_fixed, fixed_dims_dict) où is_fixed est True
    dès qu'AU MOINS L et W sont fixes ; fixed_dims_dict ne contient que les
    dimensions effectivement fixes (ex. {"L":..,"W":..} si H varie).
    """
    if df_pool is None or df_pool.empty:
        return False, {}

    fixed = {}
    for col, key in [(COL_LENGTH, "L"), (COL_WIDTH, "W"), (COL_HEIGHT, "H")]:
        if col not in df_pool.columns:
            continue
        vals = df_pool[col].dropna()
        vals = vals[vals > 0]
        if vals.empty:
            continue
        v_min = float(vals.min())
        v_max = float(vals.max())
        if v_min <= 0:
            continue
        # Tolérance ±2% : si max ≤ min*1.02 → cette dimension est fixe
        if v_max <= v_min * 1.02:
            fixed[key] = float(vals.median())

    # Le type est "dim-fixe" dès que L et W sont fixes (H peut varier)
    if "L" in fixed and "W" in fixed:
        return True, fixed
    return False, {}


# ═════════════════════════════════════════════════════════════════════════════
# PAGE — ACCUEIL
# ═════════════════════════════════════════════════════════════════════════════

def render_home():
    scroll_top()
    st.markdown("""<div class="home-hero">
        <div style="font-size:44px;margin-bottom:12px;">📦</div>
        <h1>PricePackaging</h1>
        <p>Entrez une référence pour visualiser et analyser tous ses composants d'emballage — prix en €</p>
    </div>""", unsafe_allow_html=True)

    # ── Barre de recherche
    st.markdown('<div class="ref-search-box"><h2>🔍 Recherche par Référence</h2>'
                '<p>Collez ou tapez une référence produit pour afficher ses composants packaging</p>',
                unsafe_allow_html=True)
    ci,cb = st.columns([4,1], gap="small")
    with ci:
        ref_input = st.text_input("ref", value=st.session_state.ref_search_input,
                                  placeholder="Ex : REF-001234", label_visibility="collapsed",
                                  key="ref_txt")
    with cb:
        searched = st.button("Analyser →", use_container_width=True, type="primary")
    st.markdown("</div>", unsafe_allow_html=True)

    ref_trimmed = ref_input.strip() if ref_input else ""

    if searched and ref_trimmed:
        st.session_state.ref_search_input = ref_trimmed
        if ref_trimmed in all_refs:
            df_ref  = get_ref_components(df_raw, ref_trimmed)
            summary = compute_ref_summary(df_ref, df_raw, ref_trimmed)
            st.session_state.ref_components_summary = summary
            st.session_state.selected_ref = ref_trimmed
            st.session_state.ref_not_found = False
        else:
            st.session_state.ref_components_summary = None
            st.session_state.selected_ref = None
            st.session_state.ref_not_found = True
        st.rerun()

    active_ref = st.session_state.selected_ref
    summary    = st.session_state.ref_components_summary
    ref_not_found = st.session_state.ref_not_found

    if ref_not_found and st.session_state.ref_search_input:
        st.markdown(
            f'<div class="ref-not-found-banner">❌ Référence '
            f'<b>"{st.session_state.ref_search_input}"</b> introuvable. '
            f'Vérifiez la saisie.</div>',
            unsafe_allow_html=True,
        )
        # Si ref non trouvée → on ne rend PAS la suite (cartes, voir tous, etc.)
        return

    # ── Cartes composants
    st.markdown("<br>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────────
    # MODE A — Une référence est active : on affiche UNE CARTE PAR COMPOSANT
    # (au lieu d'une carte par catégorie). On regroupe :
    #   • Tous les composants packaging (carton, sac, palette, cale, étiquette…)
    #   • Le ou les coûts MOD associés (labour_cost > 0)
    #   • Le ou les coûts Transport associés (shipping > 0)
    # → Total cost ≈ somme de tous ces prix unitaires affichés.
    # ─────────────────────────────────────────────────────────────────────────
    if active_ref and summary is not None:
        df_ref_full = get_ref_components(df_raw, active_ref)
        if COL_DESIGNATION in df_ref_full.columns and not df_ref_full.empty:
            df_ref_full = df_ref_full.drop_duplicates(
                subset=[COL_DESIGNATION], keep="first"
            ).reset_index(drop=True)

        # ── Construction de la liste de cartes à afficher
        # Chaque entrée = (cat, row_data, label_pour_affichage, price_value, source)
        cards = []  # liste de dicts

        # 1) Composants packaging classiques (carton/sac/palette/cale + étiquettes)
        for _, row in df_ref_full.iterrows():
            cat_raw = row.get("_category", "autre")
            desig = str(row.get(COL_DESIGNATION, "—"))

            # Affichage : pour les étiquettes/labels on dit "Étiquette"
            # Pour les composants "autre" non identifiés, on garde la catégorie packaging générique
            if cat_raw == "autre":
                if is_label_designation(desig):
                    cat_display = "etiquette"
                    cat_for_analysis = "autre"
                    cat_label = "🏷️ Étiquette"
                else:
                    # Catégorie inconnue : on l'affiche comme "Packaging" sans dire "autre"
                    cat_display = "autre"
                    cat_for_analysis = "autre"
                    cat_label = "📦 Packaging"
            else:
                cat_display = cat_raw
                cat_for_analysis = cat_raw
                cat_label = CAT_LABELS.get(cat_raw, "📦 Packaging")

            # Prix unitaire (= prix / partsqty) — pour que la somme ≈ total_cost
            unit_p = _safe_float(row.get("_unit_price"))
            if unit_p <= 0:
                unit_p = _safe_float(row.get(COL_PRICE))
            if unit_p <= 0:
                continue  # pas de prix valide → on n'affiche pas la carte

            cards.append({
                "cat_display": cat_display,
                "cat_for_analysis": cat_for_analysis,
                "cat_label": cat_label,
                "icon": CATEGORY_ICONS.get(cat_display, "📦"),
                "name": desig,
                "supplier": str(row.get(COL_SUPPLIER, "—")),
                "price": unit_p,           # prix UNITAIRE (affichage homepage)
                "price_for_analysis": _safe_float(row.get(COL_PRICE)),  # prix BRUT (pour analyse)
                "unit": "€",
                "row": row,
                "badge_cls": "cat-price-badge",
            })

        # 2) MOD : on prend la PREMIÈRE ligne ayant labour_cost > 0 (ou labour_rate > 0)
        # Une référence n'a typiquement qu'un seul fournisseur et donc qu'un seul taux MOD.
        if not df_ref_full.empty:
            mod_mask = pd.Series(False, index=df_ref_full.index)
            if COL_LABOUR_COST in df_ref_full.columns:
                mod_mask |= (df_ref_full[COL_LABOUR_COST].notna() & (df_ref_full[COL_LABOUR_COST] > 0))
            if COL_LABOUR in df_ref_full.columns:
                mod_mask |= (df_ref_full[COL_LABOUR].notna() & (df_ref_full[COL_LABOUR] > 0))
            mod_rows = df_ref_full[mod_mask]
            if not mod_rows.empty:
                mod_row = mod_rows.iloc[0]
                lc = _safe_float(mod_row.get(COL_LABOUR_COST, 0))
                lr = _safe_float(mod_row.get(COL_LABOUR, 0))
                # Coût MOD pour l'affichage homepage (pour la somme = total_cost)
                # et taux horaire pour l'indicateur de comparaison
                mod_price_display = lc if lc > 0 else lr
                cards.append({
                    "cat_display": "mod",
                    "cat_for_analysis": "mod",
                    "cat_label": "🏭 Taux MOD",
                    "icon": "🏭",
                    "name": "Main d'œuvre directe",
                    "supplier": str(mod_row.get(COL_SUPPLIER, "—")),
                    "price": mod_price_display,
                    "price_for_analysis": lr if lr > 0 else lc,
                    "unit": "€" if lc > 0 else "€/h",
                    "row": mod_row,
                    "badge_cls": "cat-price-badge mod-color",
                })

        # 3) Transport
        if not df_ref_full.empty and COL_SHIPPING in df_ref_full.columns:
            trans_mask = df_ref_full[COL_SHIPPING].notna() & (df_ref_full[COL_SHIPPING] > 0)
            trans_rows = df_ref_full[trans_mask]
            if not trans_rows.empty:
                trans_row = trans_rows.iloc[0]
                sh = _safe_float(trans_row.get(COL_SHIPPING, 0))
                cards.append({
                    "cat_display": "transport",
                    "cat_for_analysis": "transport",
                    "cat_label": "🚚 Transport",
                    "icon": "🚚",
                    "name": "Frais d'expédition",
                    "supplier": str(trans_row.get(COL_SUPPLIER, "—")),
                    "price": sh,
                    "price_for_analysis": sh,
                    "unit": "€",
                    "row": trans_row,
                    "badge_cls": "cat-price-badge transport-color",
                })

        # ── Affichage en grille (4 par ligne)
        if not cards:
            st.markdown(
                '<div class="info-box">Aucun composant valide trouvé pour cette référence.</div>',
                unsafe_allow_html=True
            )
        else:
            # ── En-tête complet de la référence
            n_cards = len(cards)
            sup_set = {c["supplier"] for c in cards if c["supplier"] and c["supplier"] != "—"}
            supplier_label = next(iter(sup_set)) if len(sup_set) == 1 else f"{len(sup_set)} fournisseur(s)"
            currencies_used = (summary.get("_currencies", set()) - {"EUR", ""}) if summary else set()
            conv_note = ""
            if currencies_used:
                conv_note = (f" <span style='font-size:11px;color:#94a3b8;'>"
                             f"· converti depuis {', '.join(currencies_used)}</span>")

            # ── SOURCE UNIQUE DE VÉRITÉ : _compute_ref_savings
            # Fournit total_cost (calculé si la colonne est vide), Σ économies,
            # nombre de composants en écart élevé, et coût optimisé.
            total_cost_ref = None
            total_cost_estime = False
            total_savings = 0.0
            n_high_ecart = 0
            cout_opt = None
            try:
                _opt_res = _compute_ref_savings(active_ref)
                total_cost_ref = _opt_res.get("cout_total")
                total_cost_estime = _opt_res.get("cout_total_estime", False)
                total_savings = _opt_res.get("total_savings", 0.0) or 0.0
                n_high_ecart = _opt_res.get("n_high_ecart", 0)
                cout_opt = _opt_res.get("cout_optimise")
            except Exception:
                pass

            # Coût total : valeur réelle, ou calculée (Σ coûts/pièce) — jamais "introuvable"
            # On N'AFFICHE PAS la mention "estimé" (demande utilisateur).
            if total_cost_ref is not None:
                total_cost_display = f"{total_cost_ref:.2f} €"
                cost_color = "#38bdf8"
            else:
                total_cost_display = "—"
                cost_color = "#94a3b8"

            # ── Affichage dans l'ORDRE : Coût total → N écarts → Économie → Coût optimisé
            # Si économie potentielle = 0 → on n'affiche QUE le coût total
            # (pas de ligne redondante "économie 0,00 / coût optimisé = coût total").
            opt_html = ""
            try:
                # Si l'économie est négligeable (< 0,01 €), on considère que tout va
                # bien : on n'affiche QUE le coût total (pas de coût optimisé ≈ total).
                if n_high_ecart > 0 and total_savings >= 0.01:
                    # Nombre de composants en écart — en JAUNE
                    opt_html += (
                        f'<div class="ref-meta" style="margin-top:6px;">'
                        f'<strong style="color:#fbbf24;font-size:14px;">'
                        f'⚠ {n_high_ecart} composant(s) en écart</strong></div>'
                    )
                    # Économie potentielle (= somme des écarts)
                    opt_html += (
                        f'<div class="ref-meta" style="margin-top:4px;">'
                        f'Économie potentielle : '
                        f'<strong style="color:#10b981;font-size:16px;">'
                        f'{total_savings:.2f} €</strong></div>'
                    )
                    # Coût optimisé — uniquement s'il est strictement positif
                    if cout_opt is not None and cout_opt > 0.005:
                        opt_html += (
                            f'<div class="ref-meta" style="margin-top:4px;">'
                            f'Coût optimisé : '
                            f'<strong style="color:#10b981;font-size:16px;">'
                            f'{cout_opt:.2f} €</strong></div>'
                        )
            except Exception:
                opt_html = ""

            banner_html = (
                f'<div class="ref-summary">'
                f'<div><div class="ref-code">{active_ref}</div></div>'
                f'<div style="flex:1">'
                f'<strong class="ref-meta">{n_cards} composants packaging</strong>'
                f'<div class="ref-meta">Fournisseur : '
                f'<strong style="color:white;">{supplier_label}</strong></div>'
                f'<div class="ref-meta" style="margin-top:4px;">Coût total : '
                f'<strong style="color:{cost_color};font-size:16px;">{total_cost_display}</strong>'
                f'{conv_note}</div>'
                f'{opt_html}'
                f'</div>'
                f'</div>'
            )
            st.markdown(banner_html, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Cartes des composants
            n_per_row = 4
            for row_start in range(0, n_cards, n_per_row):
                row_cards = st.columns(n_per_row, gap="small")
                for j, card in enumerate(cards[row_start:row_start + n_per_row]):
                    card_idx = row_start + j
                    with row_cards[j]:
                        row = card["row"]
                        cat = card["cat_display"]
                        cat_for_analysis = card["cat_for_analysis"]
                        desig = card["name"]
                        supplier = card["supplier"]
                        pv = card["price"]
                        pv_analysis = card["price_for_analysis"]
                        unit = card["unit"]

                        Lv = _safe_float(row.get(COL_LENGTH, 0))
                        Wv = _safe_float(row.get(COL_WIDTH, 0))
                        Hv = _safe_float(row.get(COL_HEIGHT, 0))
                        Pv = _safe_float(row.get(COL_WEIGHT, 0))

                        # Indicateur d'écart : utilise le prix BRUT pour la comparaison
                        # (cohérent avec les pools de comparaison qui contiennent les prix bruts)
                        indicator_html = ""
                        try:
                            if cat == "mod":
                                # Pour MOD on compare le taux horaire EN FILTRANT
                                # par fournisseur ET par sites manufacturing+dispatch
                                rate_for_eval = _safe_float(row.get(COL_LABOUR, 0))
                                if rate_for_eval > 0:
                                    mfg_m = str(row.get(COL_MFG_SITE, "")) if COL_MFG_SITE in row.index else ""
                                    disp_m = str(row.get(COL_DISPATCH, "")) if COL_DISPATCH in row.index else ""
                                    status, color_cls = eval_status_mod(rate_for_eval, supplier,
                                                                         mfg_site=mfg_m, dispatch=disp_m)
                                else:
                                    status, color_cls = "neutre", "price-neutral"
                            elif cat == "transport":
                                disp = str(row.get(COL_DISPATCH, "")) if COL_DISPATCH in row.index else ""
                                client = str(row.get(COL_CLIENT_SITE, "")) if COL_CLIENT_SITE in row.index else ""
                                pq_tr = _safe_float(row.get(COL_PARTS_QTY, 0))
                                status, color_cls = eval_status_transport(
                                    pv_analysis, supplier, pickup_site=disp,
                                    client_site=client, parts_qty=pq_tr)
                            elif cat in ("carton", "sac", "palette", "cale", "autre", "etiquette"):
                                # UNIFIÉ : toutes les catégories packaging (y compris étiquettes)
                                # passent par eval_status_dim → même logique de filtrage et même
                                # résultat sur la page d'accueil que sur la page d'analyse.
                                cat_eval = "autre" if cat == "etiquette" else cat_for_analysis
                                status, color_cls = eval_status_dim(
                                    pv_analysis, desig, Lv, Wv, Hv, Pv, cat_eval,
                                    country_fab=str(row.get(COL_COUNTRY_FAB, "")) if COL_COUNTRY_FAB in row.index else "",
                                    country_dispatch=str(row.get(COL_COUNTRY_DISPATCH, "")) if COL_COUNTRY_DISPATCH in row.index else ""
                                )
                            else:
                                status, color_cls = "neutre", "price-neutral"

                            # ── Fallback IA : si pas de comparatif réel, on utilise XGBoost
                            # pour donner quand même un statut bon/moyen/élevé.
                            # On MÉMORISE l'estimation IA pour afficher l'écart correspondant.
                            ai_estimate_card = None
                            if status == "neutre" and cat in ("carton", "sac", "palette", "cale", "autre", "etiquette"):
                                try:
                                    cat_ml = "autre" if cat == "etiquette" else cat_for_analysis
                                    est_p, _, _ = estimate_price_by_type(
                                        desig, cat_ml, Lv, Wv, Hv, Pv
                                    )
                                    if est_p > 0 and pv_analysis > 0:
                                        status, color_cls = evaluate_price_vs_estimate(pv_analysis, est_p)
                                        ai_estimate_card = est_p
                                except Exception:
                                    pass

                            # ── Calcul de l'ÉCART à afficher sur la carte
                            # IMPORTANT : on divise par partsqty pour avoir un écart PAR PIÈCE,
                            # cohérent avec le prix par pièce affiché sur la carte (qui est
                            # _unit_price = (price × packagingqty) / partsqty).
                            # Règles d'affichage :
                            #   • bon   → mot "Bon" (couleur verte)
                            #   • moyen → écart par pièce en € (couleur orange)
                            #   • tout écart (moyen OU élevé) → écart affiché en ROUGE
                            #   • bon → mot "Bon" en vert
                            ecart_label = ""
                            ecart_is_zero = False   # écart effectivement nul → traiter comme "Bon"
                            parts_q_card = _safe_float(row.get(COL_PARTS_QTY, 0))

                            def _fmt_ecart(val, unit):
                                # Écart affiché UNIQUEMENT si le prix dépasse la médiane.
                                # Toujours 2 décimales avec arrondi. Si l'écart est nul ou
                                # négligeable → ("", True) → "Bon".
                                if val is None:
                                    return "", False
                                if unit == "€/pièce":
                                    if round(val, 2) <= 0:        # négligeable / sous médiane
                                        return "", True
                                    return f"+{val:.2f} {unit}", False
                                else:
                                    if round(val, 2) <= 0:
                                        return "", True
                                    return f"+{val:.2f} {unit}", False

                            try:
                                if status == "bon":
                                    ecart_label = "Bon"
                                elif status in ("élevé", "moyen") and cat in (
                                    "carton", "sac", "palette", "cale", "autre", "etiquette"
                                ):
                                    cat_eval_ec = "autre" if cat == "etiquette" else cat_for_analysis
                                    pool_full_ec = CAT_DFS.get(cat_eval_ec, df_carton)
                                    pool_ec = filter_exact_dims(pool_full_ec, desig, Lv, Wv, Hv, Pv, cat_eval_ec)
                                    pool_ec = _dedup_pool(pool_ec, COL_PRICE)
                                    med_ec = None
                                    if not pool_ec.empty and COL_PRICE in pool_ec.columns:
                                        vp_ec = pool_ec[COL_PRICE].dropna()
                                        vp_ec = vp_ec[vp_ec > 0]
                                        if not vp_ec.empty:
                                            med_ec = float(vp_ec.median())
                                    if med_ec is None and ai_estimate_card is not None:
                                        med_ec = ai_estimate_card
                                    if med_ec is not None and med_ec > 0:
                                        diff_raw = pv_analysis - med_ec
                                        if parts_q_card > 0:
                                            ecart_label, ecart_is_zero = _fmt_ecart(diff_raw / parts_q_card, "€/pièce")
                                        else:
                                            ecart_label, ecart_is_zero = _fmt_ecart(diff_raw, "€")
                                elif status in ("élevé", "moyen") and cat == "mod":
                                    # On affiche l'ÉCART RÉEL calculé (économie MOD en €/pièce),
                                    # pas la simple différence de taux en €/h.
                                    mfg_m = str(row.get(COL_MFG_SITE, "")) if COL_MFG_SITE in row.index else ""
                                    med_r = mod_median_rate(supplier, mfg_m)
                                    if med_r > 0 and pv_analysis > med_r:
                                        # temps_total + partsqty(niv.4) de la réf active
                                        secs_c = 0.0; pq_c = 0.0
                                        _ref_c = st.session_state.get("selected_ref", "") or ""
                                        _ar = df_raw[df_raw[COL_REF].astype(str).str.strip().str.upper()
                                                     == str(_ref_c).strip().upper()] \
                                              if _ref_c else pd.DataFrame()
                                        if not _ar.empty:
                                            if COL_PACKAGING_LEVEL in _ar.columns and COL_PARTS_QTY in _ar.columns:
                                                _lv = pd.to_numeric(_ar[COL_PACKAGING_LEVEL], errors="coerce")
                                                _m4 = _ar[_lv == 4]
                                                if not _m4.empty:
                                                    _p = _m4[COL_PARTS_QTY].dropna(); _p = _p[_p > 0]
                                                    if not _p.empty: pq_c = float(_p.iloc[0])
                                                    if COL_LABOUR_SECONDS in _m4.columns:
                                                        _s = _m4[COL_LABOUR_SECONDS].dropna(); _s = _s[_s > 0]
                                                        if not _s.empty: secs_c = float(_s.iloc[0])
                                            if secs_c <= 0 and COL_LABOUR_SECONDS in _ar.columns:
                                                _s = _ar[COL_LABOUR_SECONDS].dropna(); _s = _s[_s > 0]
                                                if not _s.empty: secs_c = float(_s.iloc[0])
                                            if pq_c <= 0 and COL_PARTS_QTY in _ar.columns:
                                                _p = _ar[COL_PARTS_QTY].dropna(); _p = _p[_p > 0]
                                                if not _p.empty: pq_c = float(_p.iloc[0])
                                        if secs_c > 0 and pq_c > 0:
                                            eco_pp = (pv_analysis - med_r) * (secs_c / 3600.0 / pq_c)
                                            ecart_label, ecart_is_zero = _fmt_ecart(eco_pp, "€/pièce")
                                elif status in ("élevé", "moyen") and cat == "transport":
                                    pool_t_ec = _dedup_pool(df_transport, COL_SHIPPING)
                                    if COL_SHIPPING in pool_t_ec.columns:
                                        frais_ec = pool_t_ec[COL_SHIPPING].dropna()
                                        frais_ec = frais_ec[frais_ec > 0]
                                        if len(frais_ec) >= 2:
                                            med_f = float(frais_ec.median())
                                            diff_f = pv_analysis - med_f
                                            if parts_q_card > 0:
                                                ecart_label, ecart_is_zero = _fmt_ecart(diff_f / parts_q_card, "€/pièce")
                                            else:
                                                ecart_label, ecart_is_zero = _fmt_ecart(diff_f, "€")
                            except Exception:
                                ecart_label = ""

                            # ── Si l'écart est effectivement NUL/négligeable → le
                            # composant est au niveau de la médiane : "Bon" (vert).
                            if ecart_is_zero:
                                status = "bon"
                                ecart_label = "Bon"

                            # ── Code couleur : tout écart (moyen OU élevé) → ROUGE.
                            if status in ("élevé", "moyen"):
                                color_cls = "price-high"
                            elif status == "bon":
                                color_cls = "price-good"

                            if status == "neutre":
                                # "neutre" = AUCUN comparatif disponible (pool < 2 points)
                                indicator_html = '<div class="price-indicator price-neutral">Pas de comparatif</div>'
                            elif status in ("élevé", "moyen"):
                                # Il Y A un comparatif. Si on a su chiffrer l'écart → on
                                # l'affiche (rouge, 2 décimales). Sinon, l'écart est
                                # négligeable → "Écart négligeable" (PAS "pas de comparatif").
                                if ecart_label:
                                    indicator_html = f'<div class="price-indicator {color_cls}">{ecart_label}</div>'
                                else:
                                    indicator_html = '<div class="price-indicator price-good">Écart négligeable</div>'
                            elif status == "bon":
                                indicator_html = f'<div class="price-indicator price-good">{ecart_label or "Bon"}</div>'
                            else:
                                indicator_html = ""
                        except Exception:
                            indicator_html = ""

                        # Nom tronqué
                        disp_name = desig if len(desig) <= 32 else desig[:29] + "…"
                        price_disp = f"{pv:.2f} {unit}"

                        # Quantités : on N'AFFICHE PAS le partsqty sur la carte
                        # accueil — il sera visible dans "Informations du Composant"
                        # quand l'utilisateur clique sur "Analyser".
                        qty_html = ""

                        # On ne montre PAS la catégorie en doublon sous le nom
                        # (le nom contient souvent déjà l'info, par ex. "Étiquette packaging XYZ")
                        st.markdown(
                            f'<div class="cat-card" title="{desig}">'
                            f'<div class="cat-icon">{card["icon"]}</div>'
                            f'<div class="cat-label">{disp_name}</div>'
                            f'<div class="{card["badge_cls"]}">{price_disp}</div>'
                            f'{qty_html}'
                            f'{indicator_html}'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                        if st.button("Analyser", key=f"hc_comp_{card_idx}",
                                     use_container_width=True):
                            st.session_state.selected_component = _build_component_dict(
                                row, cat_for_analysis
                            )
                            # On RÉINITIALISE les filtres fournisseur des analyses
                            # (sinon un filtre saisi pour un composant reste actif
                            # pour le composant suivant).
                            for _k in ["dim_sup_carton","dim_sup_sac","dim_sup_palette",
                                       "dim_sup_cale","dim_sup_film","dim_sup_autre",
                                       "sac_sup_filter","palette_sup_filter"]:
                                st.session_state.pop(_k, None)
                            st.session_state.run_analysis = True
                            go_page("component_analysis")
                            st.rerun()

    else:
        # ─────────────────────────────────────────────────────────────────────
        # MODE B — Pas de référence active : on n'affiche AUCUNE carte de catégorie.
        # L'utilisateur n'a pas encore choisi de référence, on lui propose juste
        # l'exploration par type (section ci-dessous).
        # ─────────────────────────────────────────────────────────────────────
        pass

    # ─────────────────────────────────────────────────────────────────────────
    # NOUVEAU : Bouton d'accès à la page Optimisation Multi-Références
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
        <div style="background:linear-gradient(135deg,#065f46 0%,#10b981 50%,#34d399 100%);
                    border-radius:18px;padding:24px 28px;margin:8px 0 12px 0;
                    box-shadow:0 10px 30px rgba(16,185,129,0.25);
                    position:relative;overflow:hidden;">
            <div style="position:absolute;top:0;right:0;width:200px;height:200px;
                        background:radial-gradient(circle,rgba(255,255,255,0.15) 0%,transparent 70%);
                        border-radius:50%;"></div>
            <div style="position:relative;z-index:1;display:flex;justify-content:space-between;
                        align-items:center;gap:24px;flex-wrap:wrap;">
                <div>
                    <h3 style="margin:4px 0 2px 0;color:white;font-size:22px;font-weight:900;
                               letter-spacing:-0.4px;">
                        💰 Calculer l'économie potentielle sur plusieurs références
                    </h3>
                    <p style="margin:0;color:#ecfdf5;font-size:13px;">
                        Saisissez une liste de références — l'outil identifie tous les composants
                        en écart élevé et calcule combien vous pouvez gagner au total.
                    </p>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    _, ccta, _ = st.columns([1, 2, 1])
    with ccta:
        if st.button("🔎 Ouvrir l'analyse multi-références",
                     key="open_batch_opt",
                     use_container_width=True,
                     type="primary"):
            go_page("batch_optimization")
            st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # Navigation par TYPE de composant (toujours visible, avec ou sans référence)
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="browse-section">'
        '<h3>🔎 Explorer par Type de Composant</h3>'
        '<p>Pas de référence ? Parcourez tous les types disponibles dans la base, '
        'avec leurs prix de marché.</p>'
        '</div>',
        unsafe_allow_html=True
    )

    BROWSE_CATS = [
        ("carton",  "📦", "Carton",  "Galia, Bouclier, Cloche…",  _svg_carton("default")),
        ("sac",     "🛍️", "Sac",     "Plastique, bulles, VCI…",   _svg_sac()),
        ("palette", "🪵", "Palette", "Bois, plastique, NIMP15…",  _svg_palette("default")),
        ("cale",    "🧩", "Cale",    "Cales, inserts…",            _svg_cale()),
        ("film",    "🎞️", "Film",    "Bulles, étirable, rétractable", _svg_sac()),
    ]

    n_browse = len(BROWSE_CATS)
    cols_browse = st.columns(n_browse)
    for i, (cat_b, icon_b, label_b, desc_b, _svg) in enumerate(BROWSE_CATS):
        with cols_browse[i]:
            types_count = get_types_for_category(cat_b)
            n_types = len(types_count)
            count_html = (f'<div class="b-count">{n_types} type(s)</div>'
                          if n_types > 0 else '<div class="b-count" style="background:#94a3b8;">Aucun</div>')

            st.markdown(
                f'<div class="browse-card">'
                f'<div class="b-icon">{icon_b}</div>'
                f'<div class="b-label">{label_b}</div>'
                f'<div class="b-desc">{desc_b}</div>'
                f'{count_html}'
                f'</div>',
                unsafe_allow_html=True,
            )
            if st.button(f"Explorer", key=f"browse_{cat_b}", use_container_width=True,
                         disabled=(n_types == 0)):
                st.session_state.browse_category = cat_b
                st.session_state.browse_type = None
                go_page("browse_types"); st.rerun()

    # ─────────────────────────────────────────────────────────────────────
    # EXPLORER LES TAUX MOD — par fournisseur + site de production
    # ─────────────────────────────────────────────────────────────────────
    try:
        if df_mod is not None and not df_mod.empty and COL_LABOUR in df_mod.columns:
            st.markdown(
                '<div class="browse-section" style="margin-top:18px;">'
                '<h3>⚙️ Explorer les taux MOD</h3>'
                '<p>Choisissez un fournisseur et un site de production pour comparer '
                'les taux de main d\'œuvre.</p>'
                '</div>',
                unsafe_allow_html=True
            )

            mod_pool = df_mod[df_mod[COL_LABOUR].notna() & (df_mod[COL_LABOUR] > 0)].copy()

            # Liste des fournisseurs (valeurs réelles distinctes)
            sup_list = sorted({str(s).strip() for s in mod_pool[COL_SUPPLIER].dropna().unique()
                               if str(s).strip() and str(s).strip() not in ("nan", "None")}) \
                       if COL_SUPPLIER in mod_pool.columns else []

            mc1, mc2 = st.columns(2)
            with mc1:
                sel_sup = st.selectbox("Fournisseur", ["— Tous —"] + sup_list, key="mod_explore_sup",
                                       help="Tapez quelques lettres pour filtrer la liste")

            # Matching TOLÉRANT du fournisseur (défini globalement plus bas)
            # Liste des sites de production — FILTRÉE selon le fournisseur choisi
            site_source = mod_pool
            if sel_sup != "— Tous —" and COL_SUPPLIER in site_source.columns:
                site_source = site_source[_sup_match(site_source[COL_SUPPLIER], sel_sup)]
            site_list = sorted({str(s).strip() for s in site_source[COL_MFG_SITE].dropna().unique()
                                if str(s).strip() and str(s).strip() not in ("nan", "None")}) \
                        if COL_MFG_SITE in site_source.columns else []
            with mc2:
                sel_site = st.selectbox("Site de production", ["— Tous —"] + site_list,
                                        key="mod_explore_site")

            # Filtrage (matching tolérant pour le fournisseur)
            filt = mod_pool
            if sel_sup != "— Tous —" and COL_SUPPLIER in filt.columns:
                filt = filt[_sup_match(filt[COL_SUPPLIER], sel_sup)]
            if sel_site != "— Tous —" and COL_MFG_SITE in filt.columns:
                filt = filt[filt[COL_MFG_SITE].astype(str).str.strip() == sel_site]

            if filt.empty:
                st.markdown('<div class="info-box">Aucun taux MOD pour cette combinaison.</div>',
                            unsafe_allow_html=True)
            else:
                rates_e = filt[COL_LABOUR].dropna()
                rates_e = rates_e[rates_e > 0]
                if not rates_e.empty:
                    med_e = float(rates_e.median())
                    min_e = float(rates_e.min())
                    max_e = float(rates_e.max())

                    # Dates de validation associées (min/max = ligne correspondante ;
                    # médiane = date de validation la plus récente du pool)
                    def _mod_date(target):
                        if COL_ECON_DATE not in filt.columns:
                            return ""
                        diff = (pd.to_numeric(filt[COL_LABOUR], errors="coerce") - target).abs()
                        if diff.dropna().empty:
                            return ""
                        idx = diff.idxmin()
                        d = parse_econ_date(filt.loc[idx, COL_ECON_DATE])
                        return d.strftime("%d/%m/%Y") if pd.notna(d) else ""
                    d_recent = ""
                    if COL_ECON_DATE in filt.columns:
                        _dd = parse_econ_date(filt[COL_ECON_DATE]).dropna()
                        if not _dd.empty: d_recent = _dd.max().strftime("%d/%m/%Y")
                    dmin_l, dmax_l = _mod_date(min_e), _mod_date(max_e)

                    # Ordre : minimum (gauche) · médiane (milieu) · maximum (droite)
                    k1, k2, k3 = st.columns(3)
                    k1.metric("Taux minimum", f"{min_e:.2f} €/h",
                              delta=f"📅 validé {dmin_l}" if dmin_l else None, delta_color="off")
                    k2.metric("Taux médiane", f"{med_e:.2f} €/h",
                              delta=f"📅 validé {d_recent}" if d_recent else None, delta_color="off")
                    k3.metric("Taux maximum", f"{max_e:.2f} €/h",
                              delta=f"📅 validé {dmax_l}" if dmax_l else None, delta_color="off")

                # Tableau : fournisseur, site de production, pays, taux + date validité
                cols_mod_e = [c for c in [COL_REF, COL_SUPPLIER, COL_MFG_SITE,
                              COL_COUNTRY_FAB, COL_LABOUR, COL_ECON_DATE]
                              if c in filt.columns]
                tbl = filt[cols_mod_e].copy()
                # Tri par taux croissant puis LIMITE d'affichage (perf : éviter de
                # rendre des dizaines de milliers de lignes, ce qui fige la page).
                if COL_LABOUR in tbl.columns:
                    tbl = tbl.sort_values(COL_LABOUR)
                _MAX_ROWS = 500
                _total_rows = len(tbl)
                tbl = tbl.head(_MAX_ROWS).reset_index(drop=True)
                # Pays complet — VECTORISÉ (map sur dict, pas apply ligne à ligne)
                if COL_COUNTRY_FAB in tbl.columns:
                    tbl[COL_COUNTRY_FAB] = (tbl[COL_COUNTRY_FAB].astype(str).str.strip()
                                            .str.upper().map(COUNTRY_NAMES)
                                            .fillna(tbl[COL_COUNTRY_FAB]))
                if COL_ECON_DATE in tbl.columns:
                    tbl[COL_ECON_DATE] = parse_econ_date(tbl[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")
                tbl = tbl.rename(columns={
                    COL_REF: "Référence", COL_SUPPLIER: "Fournisseur",
                    COL_MFG_SITE: "Site de production", COL_COUNTRY_FAB: "Pays de fabrication",
                    COL_LABOUR: "Taux MOD (€/h)", COL_ECON_DATE: "Validité",
                })
                if _total_rows > _MAX_ROWS:
                    st.caption(f"Affichage des {_MAX_ROWS} taux les plus bas "
                               f"sur {_total_rows} au total — filtrez par fournisseur "
                               f"ou site pour affiner.")
                st.dataframe(tbl, use_container_width=True, hide_index=True,
                             height=min(380, 60 + 35 * min(len(tbl), 9)),
                             column_config={
                                 "Taux MOD (€/h)": st.column_config.NumberColumn(format="%.2f"),
                                 "Référence": st.column_config.TextColumn("Référence", pinned=True),
                             })

    except Exception as _e_sec:
        st.error(f"⚠ Section « Explorer les taux MOD » indisponible : {_e_sec}")
    # ─────────────────────────────────────────────────────────────────────
    # EXPLORER LES FRAIS DE TRANSPORT — par site d'enlèvement + site client
    # ─────────────────────────────────────────────────────────────────────
    try:
        st.markdown(
            '<div class="browse-section" style="margin-top:18px;">'
            '<h3>🚚 Explorer les frais de transport</h3>'
            '<p>Choisissez un fournisseur, un site d\'enlèvement et un site client '
            'pour comparer les frais d\'expédition.</p>'
            '</div>',
            unsafe_allow_html=True
        )
        _trp_ok = (df_transport is not None and not df_transport.empty
                   and COL_SHIPPING in df_transport.columns)
        if not _trp_ok:
            _cands = ", ".join([c for c in df_raw.columns
                                if any(k in str(c).lower()
                                       for k in ["ship", "transport", "frais", "expe"])]) or "aucune"
            st.info(f"Aucune colonne de frais de transport détectée "
                    f"(attendu : « shipping_charge »). Colonnes ressemblantes : {_cands}.")
        if _trp_ok:

            trp_pool = df_transport[df_transport[COL_SHIPPING].notna()
                                    & (df_transport[COL_SHIPPING] > 0)].copy()

            # Liste des fournisseurs
            trp_sup_list = sorted({str(s).strip() for s in trp_pool[COL_SUPPLIER].dropna().unique()
                                   if str(s).strip() and str(s).strip() not in ("nan", "None")}) \
                           if COL_SUPPLIER in trp_pool.columns else []

            tc1, tc2, tc3 = st.columns(3)
            with tc1:
                sel_trp_sup = st.selectbox("Fournisseur", ["— Tous —"] + trp_sup_list,
                                           key="trp_explore_sup")

            # Liste des sites d'enlèvement — FILTRÉE selon le fournisseur choisi
            # (matching tolérant via _sup_match défini dans le bloc MOD ci-dessus)
            pickup_source = trp_pool
            if sel_trp_sup != "— Tous —" and COL_SUPPLIER in pickup_source.columns:
                pickup_source = pickup_source[_sup_match(pickup_source[COL_SUPPLIER], sel_trp_sup)]
            pickup_list = sorted({str(s).strip() for s in pickup_source[COL_DISPATCH].dropna().unique()
                                  if str(s).strip() and str(s).strip() not in ("nan", "None")}) \
                          if COL_DISPATCH in pickup_source.columns else []
            with tc2:
                sel_pickup = st.selectbox("Site d'enlèvement", ["— Tous —"] + pickup_list,
                                          key="trp_explore_pickup")

            # Liste des sites client — FILTRÉE selon fournisseur + site d'enlèvement
            client_source = trp_pool
            if sel_trp_sup != "— Tous —" and COL_SUPPLIER in client_source.columns:
                client_source = client_source[_sup_match(client_source[COL_SUPPLIER], sel_trp_sup)]
            if sel_pickup != "— Tous —" and COL_DISPATCH in client_source.columns:
                client_source = client_source[
                    client_source[COL_DISPATCH].astype(str).str.strip() == sel_pickup]
            client_list = sorted({str(s).strip() for s in client_source[COL_CLIENT_SITE].dropna().unique()
                                  if str(s).strip() and str(s).strip() not in ("nan", "None")}) \
                          if COL_CLIENT_SITE in client_source.columns else []
            with tc3:
                sel_client = st.selectbox("Site client", ["— Tous —"] + client_list,
                                          key="trp_explore_client")

            # Filtrage (matching tolérant pour le fournisseur)
            filt_t = trp_pool
            if sel_trp_sup != "— Tous —" and COL_SUPPLIER in filt_t.columns:
                filt_t = filt_t[_sup_match(filt_t[COL_SUPPLIER], sel_trp_sup)]
            if sel_pickup != "— Tous —" and COL_DISPATCH in filt_t.columns:
                filt_t = filt_t[filt_t[COL_DISPATCH].astype(str).str.strip() == sel_pickup]
            if sel_client != "— Tous —" and COL_CLIENT_SITE in filt_t.columns:
                filt_t = filt_t[filt_t[COL_CLIENT_SITE].astype(str).str.strip() == sel_client]

            if filt_t.empty:
                st.markdown('<div class="info-box">Aucun frais de transport pour cette combinaison.</div>',
                            unsafe_allow_html=True)
            else:
                frais_e = filt_t[COL_SHIPPING].dropna()
                frais_e = frais_e[frais_e > 0]
                if not frais_e.empty:
                    med_t = float(frais_e.median())
                    min_t = float(frais_e.min())
                    max_t = float(frais_e.max())

                    def _trp_date(target):
                        if COL_ECON_DATE not in filt_t.columns:
                            return ""
                        diff = (pd.to_numeric(filt_t[COL_SHIPPING], errors="coerce") - target).abs()
                        if diff.dropna().empty:
                            return ""
                        idx = diff.idxmin()
                        d = parse_econ_date(filt_t.loc[idx, COL_ECON_DATE])
                        return d.strftime("%d/%m/%Y") if pd.notna(d) else ""
                    d_recent_t = ""
                    if COL_ECON_DATE in filt_t.columns:
                        _dd = parse_econ_date(filt_t[COL_ECON_DATE]).dropna()
                        if not _dd.empty: d_recent_t = _dd.max().strftime("%d/%m/%Y")
                    dmin_t, dmax_t = _trp_date(min_t), _trp_date(max_t)

                    # Ordre : minimum (gauche) · médiane (milieu) · maximum (droite)
                    t1, t2, t3 = st.columns(3)
                    t1.metric("Frais minimum", f"{min_t:.2f} €",
                              delta=f"📅 validé {dmin_t}" if dmin_t else None, delta_color="off")
                    t2.metric("Frais médiane", f"{med_t:.2f} €",
                              delta=f"📅 validé {d_recent_t}" if d_recent_t else None, delta_color="off")
                    t3.metric("Frais maximum", f"{max_t:.2f} €",
                              delta=f"📅 validé {dmax_t}" if dmax_t else None, delta_color="off")

                cols_trp_e = [c for c in [COL_REF, COL_SUPPLIER, COL_DISPATCH,
                                          COL_CLIENT_SITE, COL_COUNTRY_DISPATCH,
                                          COL_SHIPPING, COL_ECON_DATE]
                              if c in filt_t.columns]
                tbl_t = filt_t[cols_trp_e].copy()
                if COL_SHIPPING in tbl_t.columns:
                    tbl_t = tbl_t.sort_values(COL_SHIPPING)
                _MAX_T = 500
                _tot_t = len(tbl_t)
                tbl_t = tbl_t.head(_MAX_T).reset_index(drop=True)
                if COL_COUNTRY_DISPATCH in tbl_t.columns:
                    tbl_t[COL_COUNTRY_DISPATCH] = (tbl_t[COL_COUNTRY_DISPATCH].astype(str)
                                                   .str.strip().str.upper().map(COUNTRY_NAMES)
                                                   .fillna(tbl_t[COL_COUNTRY_DISPATCH]))
                if COL_ECON_DATE in tbl_t.columns:
                    tbl_t[COL_ECON_DATE] = parse_econ_date(tbl_t[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")
                tbl_t = tbl_t.rename(columns={
                    COL_REF: "Référence", COL_SUPPLIER: "Fournisseur",
                    COL_DISPATCH: "Site d'enlèvement", COL_CLIENT_SITE: "Site client",
                    COL_COUNTRY_DISPATCH: "Pays d'expédition",
                    COL_SHIPPING: "Frais transport (€)", COL_ECON_DATE: "Validité",
                })
                if _tot_t > _MAX_T:
                    st.caption(f"Affichage des {_MAX_T} frais les plus bas "
                               f"sur {_tot_t} au total — filtrez pour affiner.")
                st.dataframe(tbl_t, use_container_width=True, hide_index=True,
                             height=min(380, 60 + 35 * min(len(tbl_t), 9)),
                             column_config={
                                 "Frais transport (€)": st.column_config.NumberColumn(format="%.2f"),
                                 "Référence": st.column_config.TextColumn("Référence", pinned=True),
                             })
    except Exception as _e_sec:
        st.error(f"⚠ Section « Explorer les frais de transport » indisponible : {_e_sec}")
    # ─────────────────────────────────────────────────────────────────────
    # COMPARER PAR PAYS — taux MOD ou frais transport par pays fournisseur
    # ─────────────────────────────────────────────────────────────────────
    try:
        st.markdown(
            '<div class="browse-section" style="margin-top:18px;">'
            '<h3>🌍 Comparer par pays</h3>'
            '<p>Comparez les niveaux de prix selon le pays du fournisseur '
            '(fabrication ou expédition).</p>'
            '</div>',
            unsafe_allow_html=True
        )
        # Le pays est dérivé des libellés de site. On vérifie qu'on a au moins
        # quelques codes pays exploitables (2 lettres) dans fab OU dispatch.
        def _has_codes(col):
            if col not in df_raw.columns:
                return False
            vals = df_raw[col].astype(str).str.strip()
            return (vals.str.len() == 2).any()
        _has_country = _has_codes(COL_COUNTRY_FAB) or _has_codes(COL_COUNTRY_DISPATCH)
        if not _has_country:
            # Diagnostic : on indique pourquoi (sites manquants ou sans code pays)
            site_cols = [c for c in [COL_MFG_SITE, COL_DISPATCH] if c in df_raw.columns]
            if site_cols:
                ex = ""
                try:
                    s0 = df_raw[site_cols[0]].dropna().astype(str)
                    if not s0.empty:
                        ex = f" Exemple de valeur : « {s0.iloc[0][:60]} »."
                except Exception:
                    ex = ""
                hint = (f" Les colonnes de site ({', '.join(site_cols)}) existent mais "
                        f"le code pays (2 lettres en fin de libellé) n'a pas pu être extrait.{ex}")
            else:
                hint = (" Aucune colonne <b>manufacturing_site</b> / <b>dispatch_site</b> "
                        "n'a été trouvée pour en déduire le pays.")
            st.info(f"Comparaison par pays indisponible.{re.sub('<[^>]+>', '', hint)}")
        if _has_country:
            pc1, pc2 = st.columns(2)
            with pc1:
                pays_metric = st.selectbox(
                    "Quoi comparer ?",
                    ["Taux MOD (€/h)", "Frais transport (€)"],
                    key="pays_metric")
            with pc2:
                # MOD → toujours pays de fabrication. Transport → enlèvement ou client.
                if pays_metric.startswith("Taux MOD"):
                    pays_dim = "Pays de fabrication"
                    st.selectbox("Pays selon", ["Pays de fabrication"],
                                 key="pays_dim_mod", disabled=True)
                else:
                    pays_dim = st.selectbox(
                        "Pays selon",
                        ["Pays du site d'enlèvement", "Pays du site client"],
                        key="pays_dim_trp")

            # Source de données + colonne valeur + colonne pays
            if pays_metric.startswith("Taux MOD"):
                src = df_mod[df_mod[COL_LABOUR].notna() & (df_mod[COL_LABOUR] > 0)].copy()
                val_col = COL_LABOUR; unit = "€/h"
                # MOD : pays = pays de fabrication (site de production)
                country_col = COL_COUNTRY_FAB
            else:
                src = df_transport[df_transport[COL_SHIPPING].notna()
                                   & (df_transport[COL_SHIPPING] > 0)].copy()
                val_col = COL_SHIPPING; unit = "€"
                # Transport : pays du site d'enlèvement (dispatch) OU du site client
                if pays_dim == "Pays du site client":
                    # Pays dérivé du libellé client_site
                    if COL_CLIENT_SITE in src.columns:
                        src["_country_client"] = src[COL_CLIENT_SITE].apply(country_from_site)
                        country_col = "_country_client"
                    else:
                        country_col = COL_COUNTRY_DISPATCH
                else:
                    country_col = COL_COUNTRY_DISPATCH

            if country_col not in src.columns or src.empty:
                st.markdown('<div class="info-box">Donnée pays indisponible pour ce choix.</div>',
                            unsafe_allow_html=True)
            else:
                # Agrégation par pays
                grp = src[src[country_col].astype(str).str.strip() != ""].copy()
                grp["_code"] = grp[country_col].astype(str).str.strip().str.upper()
                grp = grp[grp["_code"].str.len() == 2]
                grp["_pays"] = grp["_code"].apply(country_full_name)
                agg = grp.groupby(["_code", "_pays"])[val_col].agg(
                    Médiane="median", Minimum="min", Maximum="max", Nombre="count"
                ).reset_index().sort_values("Médiane")

                if agg.empty:
                    st.markdown('<div class="info-box">Aucune donnée pays à comparer.</div>',
                                unsafe_allow_html=True)
                else:
                    agg = agg.reset_index(drop=True)
                    cheapest = agg.iloc[0]
                    priciest = agg.iloc[-1]
                    rng = (priciest["Médiane"] - cheapest["Médiane"]) or 1

                    st.markdown(
                        f"<h4 style='color:#0f172a;font-weight:800;margin:6px 0 12px 0;'>"
                        f"🌍 {pays_metric} par pays "
                        f"<span style='font-size:12px;color:#64748b;font-weight:500;'>"
                        f"· {pays_dim.lower()} · du moins cher au plus cher</span></h4>",
                        unsafe_allow_html=True
                    )

                    # ── CLASSEMENT STYLÉ (une ligne par pays)
                    rank_html = ['<div style="display:flex;flex-direction:column;gap:8px;">']
                    for rk, (_, prow) in enumerate(agg.iterrows(), start=1):
                        code = prow["_code"]; pays = prow["_pays"]
                        med = prow["Médiane"]; mn = prow["Minimum"]
                        mx = prow["Maximum"]; nb = int(prow["Nombre"])
                        fl = country_flag(code)
                        t = (med - cheapest["Médiane"]) / rng
                        if t <= 0.33:   accent = "#10b981"
                        elif t <= 0.66: accent = "#f59e0b"
                        else:           accent = "#ef4444"
                        wpct = max(6, min(100, int(100 * med / (priciest["Médiane"] or 1))))
                        rank_badge = (f'<div style="min-width:30px;height:30px;border-radius:50%;'
                                      f'background:#0f172a;color:white;display:flex;align-items:center;'
                                      f'justify-content:center;font-weight:800;font-size:13px;">{rk}</div>')
                        rank_html.append(
                            f'<div style="display:flex;align-items:center;gap:14px;'
                            f'background:white;border:1px solid #e2e8f0;border-radius:14px;'
                            f'padding:12px 16px;box-shadow:0 2px 8px rgba(0,0,0,0.03);">'
                            f'{rank_badge}'
                            f'<div style="font-size:26px;">{fl}</div>'
                            f'<div style="flex:1;min-width:0;">'
                            f'<div style="font-size:14px;font-weight:800;color:#0f172a;">{pays} '
                            f'<span style="font-size:11px;color:#94a3b8;font-weight:600;">({code})</span></div>'
                            f'<div style="height:8px;background:#f1f5f9;border-radius:4px;'
                            f'overflow:hidden;margin-top:5px;">'
                            f'<div style="height:100%;width:{wpct}%;background:{accent};'
                            f'border-radius:4px;"></div></div>'
                            f'<div style="font-size:11px;color:#64748b;margin-top:4px;">'
                            f'min {mn:.2f} · max {mx:.2f} · {nb} référence(s)</div>'
                            f'</div>'
                            f'<div style="text-align:right;">'
                            f'<div style="font-size:19px;font-weight:900;color:{accent};">'
                            f'{med:.2f}</div>'
                            f'<div style="font-size:10px;color:#94a3b8;font-weight:600;">{unit}</div>'
                            f'</div></div>'
                        )
                    rank_html.append("</div>")
                    st.markdown("".join(rank_html), unsafe_allow_html=True)

                    # ── DRILL-DOWN : cliquer sur un pays pour voir le détail
                    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)
                    st.markdown(
                        "<h4 style='color:#0f172a;font-weight:800;margin:6px 0 8px 0;'>"
                        "🔍 Détail d'un pays</h4>", unsafe_allow_html=True)
                    pays_options = [f"{country_flag(c)} {p} ({c})"
                                    for c, p in zip(agg["_code"], agg["_pays"])]
                    code_by_label = {f"{country_flag(c)} {p} ({c})": c
                                     for c, p in zip(agg["_code"], agg["_pays"])}
                    sel_pays_lbl = st.selectbox("Sélectionnez un pays",
                                                ["— Choisir —"] + pays_options,
                                                key="pays_drilldown")
                    if sel_pays_lbl != "— Choisir —":
                        sel_code = code_by_label[sel_pays_lbl]
                        detail = grp[grp["_code"] == sel_code].copy()
                        fl = country_flag(sel_code)
                        pays_name = country_full_name(sel_code)
                        dv = detail[val_col].dropna(); dv = dv[dv > 0]
                        if not dv.empty:
                            # Bandeau KPI du pays
                            st.markdown(
                                f'<div style="background:linear-gradient(135deg,#0f172a,#1e3a8a);'
                                f'border-radius:18px;padding:22px 24px;color:white;margin:6px 0 14px 0;'
                                f'display:flex;align-items:center;gap:20px;">'
                                f'<div style="font-size:50px;">{fl}</div>'
                                f'<div style="flex:1;">'
                                f'<div style="font-size:22px;font-weight:900;margin-bottom:8px;">{pays_name}</div>'
                                f'<div style="display:flex;gap:26px;flex-wrap:wrap;">'
                                f'<div><div style="font-size:11px;color:#bfdbfe;">MÉDIANE</div>'
                                f'<div style="font-size:20px;font-weight:800;">{dv.median():.2f} {unit}</div></div>'
                                f'<div><div style="font-size:11px;color:#bfdbfe;">MINIMUM</div>'
                                f'<div style="font-size:20px;font-weight:800;color:#86efac;">{dv.min():.2f} {unit}</div></div>'
                                f'<div><div style="font-size:11px;color:#bfdbfe;">MAXIMUM</div>'
                                f'<div style="font-size:20px;font-weight:800;color:#fca5a5;">{dv.max():.2f} {unit}</div></div>'
                                f'<div><div style="font-size:11px;color:#bfdbfe;">RÉFÉRENCES</div>'
                                f'<div style="font-size:20px;font-weight:800;">{len(dv)}</div></div>'
                                f'</div></div></div>',
                                unsafe_allow_html=True
                            )
                            # Tableau détaillé des lignes du pays
                            d_cols = [c for c in [COL_REF, COL_SUPPLIER, COL_MFG_SITE,
                                      COL_DISPATCH, val_col, COL_ECON_DATE]
                                      if c in detail.columns]
                            d_show = detail[d_cols].copy()
                            if COL_ECON_DATE in d_show.columns:
                                d_show[COL_ECON_DATE] = parse_econ_date(d_show[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")
                            d_show = d_show.sort_values(val_col).rename(columns={
                                COL_REF: "Référence", COL_SUPPLIER: "Fournisseur",
                                COL_MFG_SITE: "Site de production", COL_DISPATCH: "Site d'enlèvement",
                                val_col: pays_metric, COL_ECON_DATE: "Date validation",
                            })
                            st.dataframe(d_show, use_container_width=True, hide_index=True,
                                         height=min(420, 60 + 35 * min(len(d_show), 10)),
                                         column_config={
                                             pays_metric: st.column_config.NumberColumn(format="%.2f"),
                                             "Référence": st.column_config.TextColumn("Référence", pinned=True),
                                         })
    except Exception as _e_sec:
        st.error(f"⚠ Section « Comparer par pays » indisponible : {_e_sec}")

def render_browse_types():
    """Page intermédiaire : grille des différents TYPES disponibles dans la catégorie."""
    scroll_top()
    cat = st.session_state.browse_category
    if not cat:
        go_page("home"); st.rerun(); return

    cat_label_map = {"carton":"📦 Carton","sac":"🛍️ Sac","palette":"🪵 Palette","cale":"🧩 Cale","film":"🎞️ Film"}
    cat_label = cat_label_map.get(cat, cat.capitalize())

    # Bouton retour
    cb1, _ = st.columns([1, 5])
    with cb1:
        if st.button("← Retour à l'Accueil", key="bk_home_browse"):
            go_page("home"); st.rerun()

    # Header coloré
    st.markdown(
        f'<div class="type-grid-header">'
        f'<h2>{cat_label} &nbsp;·&nbsp; Choisir un type</h2>'
        f"<p>Cliquez sur un type pour voir tous les fournisseurs, "
        f"comparer leurs prix et accéder au détail.</p>"
        f'</div>',
        unsafe_allow_html=True
    )

    types_with_count = get_types_for_category(cat)

    if not types_with_count:
        st.markdown('<div class="warning-box">⚠ Aucun type identifié pour cette catégorie.</div>',
                    unsafe_allow_html=True)
        return

    # ── Barre de recherche pour filtrer les types par nom (insensible à la casse)
    search_query = st.text_input(
        "🔍 Rechercher un type",
        value="",
        key=f"browse_search_{cat}",
        placeholder="Tapez quelques lettres pour filtrer (ex: galia, NIMP, bouclier…)",
    )

    if search_query and search_query.strip():
        q = search_query.strip().lower()
        types_with_count = [(t, n) for t, n in types_with_count if q in t.lower()]
        if not types_with_count:
            st.markdown(
                f'<div class="info-box">Aucun type ne contient « <b>{search_query}</b> ». '
                f'Effacez le champ pour voir tous les types.</div>',
                unsafe_allow_html=True
            )
            return
        st.markdown(
            f'<div class="info-box" style="margin:6px 0 14px 0;">'
            f'<b>{len(types_with_count)}</b> type(s) correspondant à « <b>{search_query}</b> »</div>',
            unsafe_allow_html=True
        )

    # Affichage en grille — 4 colonnes desktop, responsive
    cols_per_row = 4
    pool = CAT_DFS.get(cat)

    for row_start in range(0, len(types_with_count), cols_per_row):
        cols = st.columns(cols_per_row)
        for col_idx, (type_label, count) in enumerate(types_with_count[row_start:row_start + cols_per_row]):
            with cols[col_idx]:
                # Stats : best price (min) UNIQUEMENT si type à dimensions FIXES
                # (sinon, le best price n'a pas de sens car dépend des dimensions)
                df_t = get_pool_for_type(cat, type_label)
                price_html = ""
                stats_html = f'<div class="t-stats">{count} ligne(s)</div>'
                if not df_t.empty and COL_PRICE in df_t.columns:
                    valid_prices = df_t[COL_PRICE].dropna()
                    valid_prices = valid_prices[valid_prices > 0]
                    if not valid_prices.empty:
                        n_sup = df_t[COL_SUPPLIER].nunique() if COL_SUPPLIER in df_t.columns else 0
                        stats_html = f'<div class="t-stats">{count} ligne(s) · {n_sup} fournisseur(s)</div>'
                        # Best price affiché si TOUTES les lignes ont les mêmes dimensions ET le même poids.
                        # Sinon on affiche un indicateur expliquant de quoi dépend le prix.
                        is_fixed, _ = type_has_fixed_dimensions(df_t)
                        # Vérifier aussi si les poids varient pour ce type
                        weight_varies = False
                        if COL_WEIGHT in df_t.columns:
                            valid_weights = df_t[COL_WEIGHT].dropna()
                            valid_weights = valid_weights[valid_weights > 0]
                            if not valid_weights.empty:
                                w_min = float(valid_weights.min())
                                w_max = float(valid_weights.max())
                                # Si max > min * 1.02 → poids variable (±2% de tolérance)
                                weight_varies = (w_min > 0 and w_max > w_min * 1.02)

                        if is_fixed and not weight_varies:
                            # Dimensions ET poids fixes → best price a du sens
                            best = float(valid_prices.min())
                            price_html = (
                                f'<div class="t-price" title="Meilleur prix">'
                                f'💰 {best:.2f} €</div>'
                            )
                        elif is_fixed and weight_varies:
                            # Dimensions fixes mais poids variable → prix dépend du poids
                            price_html = (
                                f'<div class="t-price" title="Poids variable — '
                                f'le prix dépend du poids" '
                                f'style="color:#94a3b8;">⚖️ Dépend du poids</div>'
                            )
                        else:
                            # Dimensions variables → prix dépend des dimensions
                            price_html = (
                                f'<div class="t-price" title="Dimensions variables — '
                                f'définissez vos dimensions à l\'analyse" '
                                f'style="color:#94a3b8;">📏 Dépend des dimensions</div>'
                            )

                svg_html = get_type_svg(type_label, cat)
                # Tronquer le nom si trop long
                disp_name = type_label if len(type_label) <= 38 else type_label[:35] + "…"

                st.markdown(
                    f'<div class="type-card" title="{type_label}">'
                    f'{svg_html}'
                    f'<div class="t-name">{disp_name}</div>'
                    f'{stats_html}'
                    f'{price_html}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                # Clé unique pour le bouton (en évitant les caractères posant souci)
                btn_key = f"type_{cat}_{row_start}_{col_idx}"
                if st.button("Voir l'analyse →", key=btn_key, use_container_width=True):
                    st.session_state.browse_type = type_label
                    go_page("type_analysis"); st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# PAGE — TYPE ANALYSIS (analyse d'un type sélectionné, courbe + min/max + tableau)
# ═════════════════════════════════════════════════════════════════════════════

def render_type_analysis():
    """Page d'analyse d'un type sélectionné : courbe fournisseurs/prix + stats + tableau."""
    scroll_top()
    cat = st.session_state.browse_category
    type_label = st.session_state.browse_type
    if not cat or not type_label:
        go_page("home"); st.rerun(); return

    cat_label_map = {"carton":"📦 Carton","sac":"🛍️ Sac","palette":"🪵 Palette","cale":"🧩 Cale","film":"🎞️ Film"}
    cat_label = cat_label_map.get(cat, cat.capitalize())

    # Boutons retour
    cb1, cb2 = st.columns([1, 1])
    with cb1:
        if st.button("← Retour aux types", key="bk_browse"):
            go_page("browse_types"); st.rerun()
    with cb2:
        if st.button("🏠 Accueil", key="bk_home_ta"):
            go_page("home"); st.rerun()

    # Header avec SVG du type
    svg = get_type_svg(type_label, cat)
    st.markdown(
        f'<div class="type-grid-header" style="display:flex;align-items:center;gap:20px;">'
        f'<div style="background:white;border-radius:14px;padding:8px;flex-shrink:0;">{svg}</div>'
        f'<div style="flex:1;">'
        f'<h2 style="margin:0;">{cat_label} &nbsp;·&nbsp; {type_label}</h2>'
        f'<p>Analyse complète tous fournisseurs · prix de marché · détail des références</p>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True
    )

    # ── Récupérer les données
    df_pool = get_pool_for_type(cat, type_label)
    if df_pool.empty:
        st.markdown('<div class="warning-box">⚠ Aucune donnée pour ce type.</div>',
                    unsafe_allow_html=True)
        return

    # Filtrer prix valides
    df_pool = df_pool[df_pool[COL_PRICE].notna() & (df_pool[COL_PRICE] > 0)].copy()
    if df_pool.empty:
        st.markdown('<div class="warning-box">⚠ Aucun prix valide pour ce type.</div>',
                    unsafe_allow_html=True)
        return

    # ── Détection : ce TYPE a-t-il des dimensions FIXES partout (mêmes L/W/H) ?
    is_fixed_dims, fixed_dims = type_has_fixed_dimensions(df_pool)

    # ── Si dimensions identiques partout : badges visuels élégants (lecture seule)
    if is_fixed_dims:
        L_info = fixed_dims.get("L", 0.)
        W_info = fixed_dims.get("W", 0.)
        H_info = fixed_dims.get("H", 0.)
        badges = []
        for label, val in [("Longueur", L_info), ("Largeur", W_info), ("Hauteur", H_info)]:
            if val > 0:
                badges.append(
                    f'<div class="dim-badge">'
                    f'<span class="dim-badge-label">{label}</span>'
                    f'<span class="dim-badge-value">{val:g}</span>'
                    f'<span class="dim-badge-unit">mm</span>'
                    f'</div>'
                )
        if badges:
            st.markdown(
                '<div class="dim-badges-container">'
                '<div class="dim-badges-title">📐 Dimensions de ce type</div>'
                '<div class="dim-badges-row">' + "".join(badges) + '</div>'
                '</div>',
                unsafe_allow_html=True
            )

    # ── Filtre fournisseur (utile pour la courbe — un fournisseur ressort en couleur)
    sup_opts = sorted(df_pool[COL_SUPPLIER].dropna().unique().tolist()) \
               if COL_SUPPLIER in df_pool.columns else []

    if sup_opts:
        sup_sel = st.selectbox("Filtrer par Fournisseur",
                               ["Tous les Fournisseurs"] + sup_opts,
                               index=0, key="type_sup_filter")
    else:
        sup_sel = "Tous les Fournisseurs"

    df_view = df_pool.copy()
    if sup_sel and sup_sel != "Tous les Fournisseurs" and COL_SUPPLIER in df_view.columns:
        df_view = df_view[df_view[COL_SUPPLIER] == sup_sel]

    # ── LECTURE PRÉCOCE des filtres utilisateur (poids / dimensions / épaisseur)
    if is_fixed_dims:
        filter_specs_early = [("Poids (g)", COL_WEIGHT, "Filtrer par poids (g)"),
                              ("Épaisseur (mm)", COL_THICKNESS, "Filtrer par épaisseur (mm)")]
    else:
        filter_specs_early = [
            ("Longueur (mm)", COL_LENGTH,    "Filtrer par longueur (mm)"),
            ("Largeur (mm)",  COL_WIDTH,     "Filtrer par largeur (mm)"),
            ("Hauteur (mm)",  COL_HEIGHT,    "Filtrer par hauteur (mm)"),
            ("Poids (g)",     COL_WEIGHT,    "Filtrer par poids (g)"),
            ("Épaisseur (mm)", COL_THICKNESS, "Filtrer par épaisseur (mm)"),
        ]
    # On ne garde que les filtres dont la colonne existe
    filter_specs_early = [(l, o, p) for (l, o, p) in filter_specs_early
                          if o in df_view.columns]

    # ── WIDGETS DE FILTRE — rendus EN HAUT (avant courbe, stats et tableau)
    if filter_specs_early:
        st.markdown(
            '<div style="margin:4px 0 8px 0;font-size:13px;font-weight:700;color:#0f172a;">'
            '🔎 Filtrer les composants</div>',
            unsafe_allow_html=True
        )
        n_fc = min(len(filter_specs_early), 4)
        fcols = st.columns(n_fc)
        for i, (col_label, orig_col, prompt) in enumerate(filter_specs_early):
            with fcols[i % n_fc]:
                st.text_input(prompt, key=f"type_filter_{col_label}",
                              placeholder="Valeur exacte (ex: 800)")

    user_filters_early = {}  # {label: (val_str, orig_col)}
    for col_label, orig_col, _ in filter_specs_early:
        key = f"type_filter_{col_label}"
        val = str(st.session_state.get(key, "")).strip()
        if val:
            user_filters_early[col_label] = (val, orig_col)

    # Application sur df_view (graphique + stats + tableau = même pool filtré)
    for col_label, (val, orig_col) in user_filters_early.items():
        v = val.replace(",", ".").strip()
        try:
            target = float(v)
            tol = max(abs(target) * 0.001, 0.01)
            if orig_col in df_view.columns:
                df_view = df_view[df_view[orig_col].between(target - tol, target + tol)]
        except ValueError:
            pass

    # ── Médiane utilisée pour la courbe (référence visuelle), sans bandeau récap
    median_price = float(df_view[COL_PRICE].median()) if not df_view.empty else 0.

    # ── Si les filtres ont VIDÉ le pool : la courbe affiche les références
    # LES PLUS PROCHES des filtres (les mêmes que le tableau d'estimation),
    # SANS ligne médiane (médiane=0 n'a pas de sens).
    df_chart_src = df_view
    chart_is_nearest = False
    if df_view.empty:
        try:
            tp_chart = get_ai_training_pool(type_label, cat)
            if not tp_chart.empty:
                tpc = tp_chart.copy()
                tpc["_score"] = 0.0
                # Cibles depuis les filtres utilisateur
                Lc = Wc = Hc = Pc = Tc = 0.0
                if is_fixed_dims and fixed_dims:
                    Lc = fixed_dims.get("L", 0.0); Wc = fixed_dims.get("W", 0.0)
                    Hc = fixed_dims.get("H", 0.0)
                for col_label, (val, orig_col) in user_filters.items():
                    try:
                        v = float(str(val).replace(",", "."))
                    except ValueError:
                        continue
                    if orig_col == COL_LENGTH: Lc = v
                    elif orig_col == COL_WIDTH: Wc = v
                    elif orig_col == COL_HEIGHT: Hc = v
                    elif orig_col == COL_WEIGHT: Pc = v
                    elif orig_col == COL_THICKNESS: Tc = v
                for colname, target in [(COL_LENGTH, Lc), (COL_WIDTH, Wc),
                                        (COL_HEIGHT, Hc), (COL_WEIGHT, Pc),
                                        (COL_THICKNESS, Tc)]:
                    if target and target > 0 and colname in tpc.columns:
                        base = tpc[colname].fillna(0).astype(float)
                        tpc["_score"] += ((base - target).abs() / max(target, 1e-6))
                df_chart_src = tpc.sort_values("_score").head(8)
                chart_is_nearest = True
        except Exception:
            df_chart_src = df_view

    # ── Courbe : un point par ligne (fournisseur, prix), couleur par fournisseur
    chart_title_note = (" — références les plus proches de vos filtres"
                        if chart_is_nearest else "")
    st.markdown(
        f"<h3 style='color:#0f172a;font-weight:800;margin:18px 0 12px 0;'>"
        f"Distribution des Prix par Fournisseur{chart_title_note}</h3>",
        unsafe_allow_html=True
    )

    df_chart = df_chart_src.copy()
    # On utilise le poids comme axe X par défaut s'il y a de la variation, sinon le volume
    has_weight = COL_WEIGHT in df_chart.columns and df_chart[COL_WEIGHT].notna().any()
    has_dims = (COL_LENGTH in df_chart.columns and COL_WIDTH in df_chart.columns
                and df_chart[COL_LENGTH].notna().any() and df_chart[COL_WIDTH].notna().any())

    if has_weight and has_dims and not chart_is_nearest:
        df_chart["weight_kg"] = df_chart[COL_WEIGHT] / 1000.0
        df_chart["Volume_m3"] = (
            df_chart.get(COL_LENGTH, 0).fillna(0)
            * df_chart.get(COL_WIDTH, 0).fillna(0)
            * df_chart.get(COL_HEIGHT, 0).fillna(0)
        ) / 1e9
        t1, t2 = st.tabs(["Volume vs Prix", "Poids vs Prix"])
        for tab, xc, xl in [(t1, "Volume_m3", "Volume (m³)"),
                            (t2, "weight_kg", "Poids (kg)")]:
            with tab:
                fig = build_scatter(
                    df_chart, xc, COL_PRICE, xl, "Prix (€)",
                    f"{xl.split('(')[0].strip()} vs Prix",
                    0, 0,
                    show_target=False  # JAMAIS de point analysé sur cette page
                )
                st.plotly_chart(fig, use_container_width=True)
    else:
        # Scatter par fournisseur — points du pool (ou des références proches)
        fig = go.Figure()
        if COL_SUPPLIER in df_chart.columns and not df_chart.empty:
            for sup_name, grp in df_chart.groupby(COL_SUPPLIER):
                fig.add_trace(go.Scatter(
                    x=[str(sup_name)] * len(grp),
                    y=grp[COL_PRICE],
                    mode="markers",
                    name=str(sup_name),
                    marker=dict(size=12, opacity=0.85, line=dict(color="white", width=1.5)),
                    hovertemplate=f"<b>{sup_name}</b><br>Prix : %{{y:.2f}} €<extra></extra>",
                ))
            # Ligne médiane UNIQUEMENT si le pool réel n'est pas vide
            # (jamais de "médiane 0,00" avec trait pointillé)
            if not chart_is_nearest and median_price > 0:
                fig.add_hline(y=median_price, line_dash="dash", line_color="#64748b",
                              annotation_text=f"Médiane : {median_price:.2f} €",
                              annotation_position="top right",
                              annotation_font=dict(color="#0f172a", size=11))
        fig.update_layout(
            title=f"<b>Prix par fournisseur — {type_label}</b>",
            xaxis_title="Fournisseur",
            yaxis_title="Prix (€)",
            height=480,
            template="plotly_white",
            plot_bgcolor="#f8fafc", paper_bgcolor="white",
            font=dict(family="DM Sans", size=12),
            legend=dict(title="<b>Fournisseur</b>"),
        )
        st.plotly_chart(fig, use_container_width=True)

    # ── Construction du tableau (valeurs NUMÉRIQUES pour bon tri/filtrage)
    cols_to_show = []
    rename_map = {}
    if COL_REF in df_view.columns:
        cols_to_show.append(COL_REF); rename_map[COL_REF] = "Référence"
    if COL_SUPPLIER in df_view.columns:
        cols_to_show.append(COL_SUPPLIER); rename_map[COL_SUPPLIER] = "Fournisseur"

    df_table = df_view[cols_to_show].copy() if cols_to_show else df_view.copy()
    df_table.insert(min(2, len(df_table.columns)), "Type", type_label)

    if COL_LENGTH in df_view.columns:
        df_table["Longueur (mm)"] = pd.to_numeric(df_view[COL_LENGTH], errors="coerce")
    if COL_WIDTH in df_view.columns:
        df_table["Largeur (mm)"] = pd.to_numeric(df_view[COL_WIDTH], errors="coerce")
    if COL_HEIGHT in df_view.columns:
        df_table["Hauteur (mm)"] = pd.to_numeric(df_view[COL_HEIGHT], errors="coerce")
    if COL_WEIGHT in df_view.columns:
        df_table["Poids (g)"] = pd.to_numeric(df_view[COL_WEIGHT], errors="coerce")
    if COL_THICKNESS in df_view.columns and df_view[COL_THICKNESS].notna().any():
        df_table["Épaisseur (mm)"] = pd.to_numeric(df_view[COL_THICKNESS], errors="coerce")
    if COL_PRICE in df_view.columns:
        df_table["Prix unitaire (€)"] = pd.to_numeric(df_view[COL_PRICE], errors="coerce")

    df_table = df_table.rename(columns=rename_map)

    # ── Spécification des filtres (selon le type)
    # • Type à dimensions FIXES → uniquement Poids (les L/W/H sont déjà identiques)
    # • Type à dimensions VARIABLES → L, W, H ET Poids
    if is_fixed_dims:
        filter_specs = [("Poids (g)", COL_WEIGHT, "Filtrer par poids (g)"),
                        ("Épaisseur (mm)", COL_THICKNESS, "Filtrer par épaisseur (mm)")]
    else:
        filter_specs = [
            ("Longueur (mm)", COL_LENGTH,    "Filtrer par longueur (mm)"),
            ("Largeur (mm)",  COL_WIDTH,     "Filtrer par largeur (mm)"),
            ("Hauteur (mm)",  COL_HEIGHT,    "Filtrer par hauteur (mm)"),
            ("Poids (g)",     COL_WEIGHT,    "Filtrer par poids (g)"),
            ("Épaisseur (mm)", COL_THICKNESS, "Filtrer par épaisseur (mm)"),
        ]
    filter_specs = [(label, orig, prompt) for (label, orig, prompt) in filter_specs
                    if label in df_table.columns]

    # ── Lecture des valeurs de filtre depuis session_state (sans rendre les widgets)
    # On les rendra plus tard, juste avant le tableau.
    user_filters = {}  # {col_label: (value_str, orig_col)}
    for col_label, orig_col, _ in filter_specs:
        key = f"type_filter_{col_label}"
        val = str(st.session_state.get(key, "")).strip()
        if val:
            user_filters[col_label] = (val, orig_col)

    # ── Application des filtres (sur tableau ET sur stats pour cohérence)
    df_filtered = df_table.copy()
    df_view_for_stats = df_view.copy()
    invalid_filters = []  # collecte des saisies invalides à signaler plus tard

    for col_label, (val, orig_col) in user_filters.items():
        v = val.replace(",", ".").strip()
        try:
            target = float(v)
            tol = max(abs(target) * 0.001, 0.01)  # 0.1% de tolérance flottante
            df_filtered = df_filtered[
                df_filtered[col_label].between(target - tol, target + tol)
            ]
            if orig_col in df_view_for_stats.columns:
                df_view_for_stats = df_view_for_stats[
                    df_view_for_stats[orig_col].between(target - tol, target + tol)
                ]
        except ValueError:
            invalid_filters.append((col_label, val))

    # ── 1) Cartes statistiques (reflètent les filtres saisis en haut)
    st.markdown("<br>", unsafe_allow_html=True)
    if df_view_for_stats.empty:
        # Aucun composant exact → on propose une ESTIMATION IA si l'utilisateur
        # a saisi des dimensions (le modèle XGBoost est entraîné sur tout le type).
        L_est = W_est = H_est = P_est = thick_est = 0.0
        if is_fixed_dims and fixed_dims:
            L_est = fixed_dims.get("L", 0.0)
            W_est = fixed_dims.get("W", 0.0)
            H_est = fixed_dims.get("H", 0.0)
        for col_label, (val, orig_col) in user_filters.items():
            try:
                v = float(str(val).replace(",", "."))
            except ValueError:
                continue
            if orig_col == COL_LENGTH: L_est = v
            elif orig_col == COL_WIDTH: W_est = v
            elif orig_col == COL_HEIGHT: H_est = v
            elif orig_col == COL_WEIGHT: P_est = v
            elif orig_col == COL_THICKNESS: thick_est = v

        est_price = 0.0
        try:
            est_price, method_lbl, n_train = estimate_price_by_type(
                type_label, cat, L_est, W_est, H_est, P_est, thickness=thick_est
            )
        except Exception:
            est_price = 0.0

        if est_price > 0:
            # Message clair : pas de comparatif EXACT pour ces filtres + estimation visible
            st.markdown(
                f'<div class="info-box" style="background:#fffbeb;'
                f'border-left-color:#f59e0b;color:#92400e;font-size:14px;">'
                f'⚠️ <b>Pas de comparatif exact pour ces filtres.</b><br>'
                f'<span style="font-size:13px;">Aucun composant ne correspond '
                f'exactement aux dimensions / poids saisis.</span>'
                f'</div>',
                unsafe_allow_html=True
            )
            st.markdown(
                f'<div class="info-box" style="background:#eff6ff;'
                f'border-left-color:#0ea5e9;color:#075985;text-align:center;'
                f'padding:18px;">'
                f'🤖 <b style="font-size:15px;">Best Price estimé par le modèle IA</b><br>'
                f'<b style="font-size:30px;color:#0c4a6e;">{est_price:.2f} €</b>'
                f'</div>',
                unsafe_allow_html=True
            )
            # Tableau des références LES PLUS PROCHES qui ont aidé à l'estimation
            try:
                train_pool = get_ai_training_pool(type_label, cat)
                if not train_pool.empty:
                    tp = train_pool.copy()
                    # Score de proximité par rapport aux filtres saisis (dims + poids)
                    tp["_score"] = 0.0
                    def _add_dist(colname, target):
                        if target and target > 0 and colname in tp.columns:
                            base = tp[colname].fillna(0).astype(float)
                            tp["_score"] += ((base - target).abs() / max(target, 1e-6))
                    _add_dist(COL_LENGTH, L_est)
                    _add_dist(COL_WIDTH, W_est)
                    _add_dist(COL_HEIGHT, H_est)
                    _add_dist(COL_WEIGHT, P_est)
                    _add_dist(COL_THICKNESS, thick_est)
                    # On garde les 8 références les PLUS PROCHES
                    tp = tp.sort_values("_score").head(8)

                    # Structure du tableau : PRIX EN DERNIÈRE COLONNE + date validation
                    tcols = [c for c in [COL_REF, COL_DESIGNATION, COL_SUPPLIER,
                             COL_LENGTH, COL_WIDTH, COL_HEIGHT,
                             COL_WEIGHT, COL_THICKNESS, COL_ECON_DATE, COL_PRICE]
                             if c in tp.columns]
                    tshow = tp[tcols].copy()
                    if COL_ECON_DATE in tshow.columns:
                        tshow[COL_ECON_DATE] = parse_econ_date(tshow[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")
                    tshow = tshow.rename(columns={
                        COL_REF: "Référence", COL_DESIGNATION: "Composant",
                        COL_SUPPLIER: "Fournisseur", COL_PRICE: "Prix (€)",
                        COL_LENGTH: "L (mm)", COL_WIDTH: "l (mm)",
                        COL_HEIGHT: "H (mm)", COL_WEIGHT: "Poids (g)",
                        COL_THICKNESS: "Épaisseur (mm)", COL_ECON_DATE: "Date validation",
                    })
                    st.markdown(
                        f"<h4 style='color:#0f172a;font-weight:800;margin:14px 0 6px 0;"
                        f"font-size:15px;'>📚 Références qui ont aidé à l'estimation "
                        f"<span style='color:#64748b;font-weight:500;font-size:12px;'>"
                        f"— {len(tshow)} références les plus proches de vos filtres</span></h4>",
                        unsafe_allow_html=True
                    )
                    st.dataframe(tshow, use_container_width=True, hide_index=True,
                                 height=min(380, 38 + 35 * min(len(tshow), 9)),
                                 column_config={
                                     "Prix (€)": st.column_config.NumberColumn(format="%.2f"),
                                     "L (mm)": st.column_config.NumberColumn(format="%.0f"),
                                     "l (mm)": st.column_config.NumberColumn(format="%.0f"),
                                     "H (mm)": st.column_config.NumberColumn(format="%.0f"),
                                     "Poids (g)": st.column_config.NumberColumn(format="%.0f"),
                                     "Épaisseur (mm)": st.column_config.NumberColumn(format="%g"),
                                 })
            except Exception:
                pass
        else:
            st.markdown(
                '<div class="info-box" style="background:#f1f5f9;border-left-color:#94a3b8;">'
                '⚪ <b>Pas de comparatif pour ces filtres.</b> '
                'Effacez un ou plusieurs champs ci-dessous pour voir plus de résultats.'
                '</div>',
                unsafe_allow_html=True
            )
    else:
        if user_filters:
            filter_summary = " · ".join(f"{lbl.split(' (')[0]} = {val[0]}"
                                        for lbl, val in user_filters.items())
            label_pool = f"{type_label} · {filter_summary}"
        else:
            label_pool = f"{type_label} · {sup_sel}"
        stat_cards_from_df(df_view_for_stats, COL_PRICE, label_pool)

    # Signalement des saisies invalides
    for col_label, val in invalid_filters:
        st.warning(
            f"« {val} » n'est pas une valeur numérique valide pour {col_label}. "
            f"Tapez une valeur numérique (ex: 800).",
            icon="⚠️"
        )

    # ── 3) Tableau — uniquement s'il y a des composants à montrer
    if df_filtered is None or df_filtered.empty:
        # Pas de tableau ni de titre "0 ligne(s)" : on s'arrête proprement ici.
        return

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        f'<div class="table-section"><h3>Détail des Composants — {type_label} '
        f'<span style="font-size:13px;color:#64748b;font-weight:600;">'
        f'({len(df_filtered)} ligne(s))</span></h3>',
        unsafe_allow_html=True
    )

    # Tri par prix croissant initial (modifiable par clic sur l'en-tête)
    if "Prix unitaire (€)" in df_filtered.columns:
        df_filtered = df_filtered.sort_values("Prix unitaire (€)", na_position="last").reset_index(drop=True)

    # Configuration des colonnes : formatage SANS unité (déjà dans l'en-tête)
    col_config = {}
    if "Longueur (mm)"   in df_filtered.columns: col_config["Longueur (mm)"]   = st.column_config.NumberColumn(format="%g")
    if "Largeur (mm)"    in df_filtered.columns: col_config["Largeur (mm)"]    = st.column_config.NumberColumn(format="%g")
    if "Hauteur (mm)"    in df_filtered.columns: col_config["Hauteur (mm)"]    = st.column_config.NumberColumn(format="%g")
    if "Poids (g)"       in df_filtered.columns: col_config["Poids (g)"]       = st.column_config.NumberColumn(format="%g")
    if "Épaisseur (mm)"  in df_filtered.columns: col_config["Épaisseur (mm)"]  = st.column_config.NumberColumn(format="%g")
    if "Prix unitaire (€)" in df_filtered.columns: col_config["Prix unitaire (€)"] = st.column_config.NumberColumn(format="%.2f")

    if df_filtered.empty:
        st.markdown('<div class="info-box">Aucune ligne à afficher avec ces filtres.</div>',
                    unsafe_allow_html=True)
    else:
        st.dataframe(df_filtered, use_container_width=True, hide_index=True,
                     column_config=col_config)
    st.markdown('</div>', unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# PAGE — DÉTAIL RÉFÉRENCE
# ═════════════════════════════════════════════════════════════════════════════

def render_ref_detail():
    scroll_top()
    ref = st.session_state.selected_ref

    if st.button("← Retour à l'Accueil", key="bk_home_rd"):
        go_page("home"); st.rerun()

    df_ref = get_ref_components(df_raw, ref)

    if df_ref.empty:
        st.error(f"Aucun composant trouvé pour la référence **{ref}**."); return

    # ── Dédoublonnage agressif : pour une référence donnée, deux lignes avec
    # la même désignation sont considérées comme un seul composant (on garde la
    # première occurrence). Cela évite que la liste affiche plusieurs fois le
    # même composant même si certaines valeurs annexes diffèrent légèrement.
    if COL_DESIGNATION in df_ref.columns:
        df_ref = df_ref.drop_duplicates(
            subset=[COL_DESIGNATION], keep="first"
        ).reset_index(drop=True)

    n_comp    = len(df_ref)

    # Fournisseur unique (la plupart des refs n'ont qu'un seul fournisseur)
    sup_uniq = sorted(df_ref[COL_SUPPLIER].dropna().unique().tolist()) if not df_ref.empty else []
    supplier_label = sup_uniq[0] if len(sup_uniq) == 1 else (
        " · ".join(sup_uniq) if sup_uniq else "—"
    )

    # Coût total — priorité au total_cost de la base (déjà complété par fichier d'appoint)
    total_cost_ref = None
    if COL_TOTAL_COST in df_raw.columns:
        vals = df_raw[df_raw[COL_REF] == ref][COL_TOTAL_COST].dropna()
        vals = vals[vals > 0]
        if not vals.empty:
            total_cost_ref = round(float(vals.iloc[0]), 2)

    # Si le total_cost reste introuvable → afficher explicitement "Donnée introuvable"
    # plutôt qu'une somme calculée qui peut être fausse (devises mélangées, lignes manquantes).
    total_cost_display = (f"{total_cost_ref:.2f} €" if total_cost_ref is not None
                          else "Donnée introuvable")

    currencies_used: set = set()
    if "_currency" in df_ref.columns:
        currencies_used = set(df_ref["_currency"].dropna().unique()) - {"EUR",""}
    conv_note = ""
    if currencies_used:
        conv_note = (f" <span style='font-size:11px;color:#94a3b8;'>"
                     f"· converti depuis {', '.join(currencies_used)}</span>")

    cost_color = "#38bdf8" if total_cost_ref is not None else "#fbbf24"

    st.markdown(f"""<div class="ref-summary">
        <div><div class="ref-code">{ref}</div></div>
        <div style="flex:1">
            <strong class="ref-meta">{n_comp} composants packaging</strong>
            <div class="ref-meta">Fournisseur :
                <strong style="color:white;">{supplier_label}</strong>
            </div>
            <div class="ref-meta" style="margin-top:4px;">Coût total :
                <strong style="color:{cost_color};font-size:16px;">{total_cost_display}</strong>{conv_note}
            </div>
        </div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="comp-table-wrap"><h2>📋 Composants du Packaging</h2>'
                "<p class='sub'>Revenez à l'accueil et cliquez sur <b>Analyser</b> "
                "dans la catégorie souhaitée pour une analyse détaillée.</p>",
                unsafe_allow_html=True)

    has_components = False
    for _, row in df_ref.iterrows():
        desig    = str(row.get(COL_DESIGNATION,"—"))
        cat      = row.get("_category","autre")
        clean_t  = row.get("_clean_type", get_clean_type_label(desig, cat))
        icon     = CATEGORY_ICONS.get(cat,"🔹")
        supplier = str(row.get(COL_SUPPLIER,"—"))
        cur      = str(row.get("_currency","EUR"))

        Lv = row.get(COL_LENGTH); Wv = row.get(COL_WIDTH)
        Hv = row.get(COL_HEIGHT); Pv = row.get(COL_WEIGHT)

        dims = []
        if pd.notna(Lv) and Lv>0: dims.append(f"L:{Lv:.0f} mm")
        if pd.notna(Wv) and Wv>0: dims.append(f"l:{Wv:.0f} mm")
        if pd.notna(Hv) and Hv>0: dims.append(f"H:{Hv:.0f} mm")
        if pd.notna(Pv) and Pv>0: dims.append(f"{Pv:.0f} g")
        if cat == "mod":
            lc = _safe_float(row.get(COL_LABOUR_COST))
            lr = _safe_float(row.get(COL_LABOUR))
            if lc>0: dims.append(f"Coût MOD : {lc:.2f} €")
            if lr>0: dims.append(f"Taux : {lr:.2f} €/h")
        elif cat == "transport":
            sh = _safe_float(row.get(COL_SHIPPING))
            if sh>0: dims.append(f"Frais : {sh:.2f} €")

        pv = _get_display_price(row, cat)
        if not dims or pv is None:
            continue

        has_components = True
        dims_str = " &nbsp;·&nbsp; ".join(dims) if dims else ""
        price_disp = f"{pv:.2f} €" if pv is not None else "—"
        orig_note  = (f'<div class="c-price-orig">converti depuis {cur}</div>'
                      if cur not in ("EUR","") else "")

        # ── Indicateur de prix — HELPERS UNIFIÉS (cohérent avec accueil + analyse)
        indicator_html = ""
        best_price_html = ""
        try:
            if pv is not None:
                pool_for_best = None
                price_col = COL_PRICE
                status = "neutre"
                color_cls = "price-neutral"
                # Pool for best price : on calcule TOUJOURS, peu importe le statut
                # (la demande utilisateur : afficher le best price même si le prix est "bon")
                if cat == "mod":
                    mfg_m = str(row.get(COL_MFG_SITE, "")) if COL_MFG_SITE in row.index else ""
                    disp_m = str(row.get(COL_DISPATCH, "")) if COL_DISPATCH in row.index else ""
                    status, color_cls = eval_status_mod(pv, supplier,
                                                         mfg_site=mfg_m, dispatch=disp_m)
                    if supplier and COL_SUPPLIER in df_mod.columns:
                        pool_for_best = df_mod[df_mod[COL_SUPPLIER] == supplier]
                        # Filtre par sites pour cohérence avec eval_status_mod
                        if mfg_m and mfg_m.strip() and mfg_m.lower() not in ("nan", "none", "") \
                                and COL_MFG_SITE in pool_for_best.columns:
                            p2 = pool_for_best[pool_for_best[COL_MFG_SITE].astype(str).str.strip() == mfg_m.strip()]
                            if not p2.empty: pool_for_best = p2
                        if disp_m and disp_m.strip() and disp_m.lower() not in ("nan", "none", "") \
                                and COL_DISPATCH in pool_for_best.columns:
                            p2 = pool_for_best[pool_for_best[COL_DISPATCH].astype(str).str.strip() == disp_m.strip()]
                            if not p2.empty: pool_for_best = p2
                        price_col = COL_LABOUR
                elif cat == "transport":
                    disp = str(row.get(COL_DISPATCH, "")) if COL_DISPATCH in row.index else ""
                    client = str(row.get(COL_CLIENT_SITE, "")) if COL_CLIENT_SITE in row.index else ""
                    pq_tr = _safe_float(row.get(COL_PARTS_QTY, 0)) if hasattr(row, 'get') else 0
                    status, color_cls = eval_status_transport(
                        pv, supplier, pickup_site=disp, client_site=client, parts_qty=pq_tr)
                    pool_for_best = df_transport.copy()
                    if supplier and COL_SUPPLIER in pool_for_best.columns:
                        pool_for_best = pool_for_best[pool_for_best[COL_SUPPLIER] == supplier]
                    if disp and COL_DISPATCH in pool_for_best.columns:
                        p2 = pool_for_best[pool_for_best[COL_DISPATCH] == disp]
                        if not p2.empty: pool_for_best = p2
                    if client and COL_CLIENT_SITE in pool_for_best.columns:
                        p2 = pool_for_best[pool_for_best[COL_CLIENT_SITE] == client]
                        if not p2.empty: pool_for_best = p2
                    price_col = COL_SHIPPING
                elif cat == "autre" and is_label_designation(desig):
                    status, color_cls = eval_status_label(pv)
                    pool_for_best = df_label
                    price_col = COL_PRICE
                else:
                    status, color_cls = eval_status_dim(
                        pv, desig,
                        _safe_float(Lv), _safe_float(Wv),
                        _safe_float(Hv), _safe_float(Pv), cat
                    )
                    # IMPORTANT : on utilise filter_exact_dims pour le best price
                    # (MÊME type + MÊMES dimensions + MÊME poids) — pas filter_same_component
                    # qui était trop large (±5%) → best price faussé
                    pool_full = CAT_DFS.get(cat, df_carton)
                    pool_for_best = filter_exact_dims(
                        pool_full, desig,
                        _safe_float(Lv), _safe_float(Wv),
                        _safe_float(Hv), _safe_float(Pv), cat
                    )
                    price_col = COL_PRICE

                # Affichage statut
                if status == "neutre":
                    indicator_html = (
                        f'<span class="price-indicator price-neutral" '
                        f'style="margin-top:6px;display:inline-block;">Pas de comparatif</span>'
                    )
                else:
                    indicator_html = (
                        f'<span class="price-indicator {color_cls}" '
                        f'style="margin-top:6px;display:inline-block;">{status}</span>'
                    )

                # Affichage best price — TOUJOURS si pool non vide (même si status = bon)
                if pool_for_best is not None and not pool_for_best.empty:
                    pool_for_best = _dedup_pool(pool_for_best, price_col)
                    if price_col in pool_for_best.columns:
                        valid_prices = pool_for_best[price_col].dropna()
                        valid_prices = valid_prices[valid_prices > 0]
                        if not valid_prices.empty:
                            best_price = float(valid_prices.min())
                            unit = "€/h" if price_col == COL_LABOUR else "€"
                            if best_price < pv:
                                saving = pv - best_price
                                best_price_html = (
                                    f'<div style="margin-top:6px;font-size:11px;color:#059669;'
                                    f'font-weight:700;background:#ecfdf5;border-radius:6px;'
                                    f'padding:4px 8px;border:1px solid #bbf7d0;display:inline-block;">'
                                    f'💰 Meilleur prix : <b>{best_price:.2f} {unit}</b> '
                                    f'(économie {saving:.2f} {unit})'
                                    f'</div>'
                                )
                            elif best_price == pv:
                                best_price_html = (
                                    f'<div style="margin-top:6px;font-size:11px;color:#059669;'
                                    f'font-weight:700;background:#ecfdf5;border-radius:6px;'
                                    f'padding:4px 8px;border:1px solid #bbf7d0;display:inline-block;">'
                                    f'💰 Vous avez le meilleur prix : <b>{best_price:.2f} {unit}</b>'
                                    f'</div>'
                                )
        except Exception:
            indicator_html = ""
            best_price_html = ""

        # ── Affichage de la ligne — ligne meta retirée (évite la répétition nom+fournisseur)
        st.markdown(
            f'<div class="comp-row">'
            f'<div class="c-icon">{icon}</div>'
            f'<div class="c-body">'
            f'<div class="c-name">{desig}</div>'
            f'<div class="c-dims">{dims_str}</div>'
            f'</div>'
            f'<div style="align-self:center;text-align:right;">'
            f'<div class="c-price">{price_disp}</div>'
            f'{indicator_html}'
            f'{best_price_html}'
            f'{orig_note}'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    if not has_components:
        st.markdown("<p style='color:#94a3b8;text-align:center;padding:20px;'>Aucun composant avec données complètes à afficher.</p>",
                    unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    _,cb,_ = st.columns([2,2,2])
    with cb:
        if st.button("🏠 Retour à l'Accueil", use_container_width=True, type="primary"):
            go_page("home"); st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# PAGE — ANALYSE COMPOSANT (depuis bouton Analyser de l'accueil)
# ═════════════════════════════════════════════════════════════════════════════

def render_component_analysis():
    scroll_top()
    comp       = st.session_state.selected_component
    ref_active = st.session_state.selected_ref

    if not comp:
        st.markdown(f'<div class="ref-not-found-banner">❌ <b>Référence {ref_active}</b> : '
                    f'composant introuvable pour cette catégorie.</div>',
                    unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🏠 Retour à l'Accueil", use_container_width=True, type="primary"):
            go_page("home"); st.rerun()
        return

    cat       = comp["category"]
    desig     = comp["designation"]
    L,W,H,P   = comp["L"], comp["W"], comp["H"], comp["P"]
    thickness = comp.get("thickness", 0.)
    price     = comp["price"]
    supplier  = comp["supplier"]
    cur       = comp.get("currency","EUR")
    ref_code  = comp.get("ref","—")
    vol_m3    = (L*W*H)/1e9 if (L>0 and W>0 and H>0) else 0.
    poids_kg  = P/1000

    if st.button("← Retour à l'Accueil", key="bk_home_ca"):
        go_page("home"); st.rerun()

    # ── Header — affiche le NOM COMPLET du composant (sans troncature)
    # + référence à côté du titre
    if cat == "mod":
        title_html = "🏭 MOD &nbsp;·&nbsp; Taux horaire"
        display_type = "Taux horaire"
    elif cat == "transport":
        title_html = "🚚 Transport &nbsp;·&nbsp; Frais d'expédition"
        display_type = "Frais d'expédition"
    else:
        strict_type = get_strict_type(desig, cat) if cat in ("carton","palette") else get_clean_type_label(desig, cat)
        display_type = strict_type if strict_type else get_clean_type_label(desig, cat)
        cat_icon_label = CAT_LABELS.get(cat, cat)
        # Titre = icône catégorie + NOM COMPLET du composant
        title_html = f"{cat_icon_label} &nbsp;·&nbsp; {desig}"

    ref_badge = (
        f'<span style="display:inline-block;background:rgba(255,255,255,0.18);'
        f'border:1px solid rgba(255,255,255,0.35);border-radius:8px;'
        f'padding:6px 14px;font-size:14px;font-weight:700;color:white;'
        f'letter-spacing:0.4px;margin-left:14px;vertical-align:middle;'
        f'font-family:JetBrains Mono, monospace;">'
        f'Réf : {ref_code}</span>'
    ) if ref_code and ref_code not in ("—","") else ""

    st.markdown(
        f"<div class='section-header' style='display:flex;align-items:center;flex-wrap:wrap;'>"
        f"<h2 style='margin:0;display:inline-block;'>{title_html}</h2>{ref_badge}"
        f"</div>",
        unsafe_allow_html=True
    )

    if cur not in ("EUR",""):
        st.markdown(f'<div class="currency-note">💱 Les prix de ce composant ont été convertis '
                    f'depuis <b>{cur}</b> vers <b>EUR (€)</b> avec les taux de change embarqués.</div>',
                    unsafe_allow_html=True)

    # ── Panneau informations
    items_text = [
        ("Type",        display_type),
        ("Fournisseur", supplier),
    ]
    mfg_site = comp.get("mfg_site", "")
    disp_site = comp.get("dispatch", "")
    client_site = comp.get("client_site", "")
    if cat == "mod":
        # MOD : on affiche le SITE DE PRODUCTION (manufacturing) uniquement —
        # pas le site d'enlèvement.
        if mfg_site and str(mfg_site).lower() not in ("", "nan", "none"):
            items_text.append(("Site de production", mfg_site))
    elif cat == "transport":
        # Transport : site d'enlèvement + site client
        if disp_site and str(disp_site).lower() not in ("", "nan", "none"):
            items_text.append(("Site d'enlèvement", disp_site))
        if client_site and str(client_site).lower() not in ("", "nan", "none"):
            items_text.append(("Site client", client_site))

    items_num = []
    if cat == "mod":
        lc = comp.get("labour_cost",0.)
        lr = comp.get("labour_rate",0.)
        if lc>0: items_num.append(("Coût MOD",     f"{lc:.2f}", "€"))
        if lr>0: items_num.append(("Taux horaire",  f"{lr:.2f}", "€/h"))
    elif cat == "transport":
        sh = comp.get("shipping",0.)
        if sh>0: items_num.append(("Frais transport", f"{sh:.2f}", "€"))
    else:
        if L>0:     items_num.append(("Longueur",      f"{L:g}",  "mm"))
        if W>0:     items_num.append(("Largeur",        f"{W:g}",  "mm"))
        if H>0:     items_num.append(("Hauteur",        f"{H:g}",  "mm"))
        # Poids : valeur exacte, pas d'arrondi (1.5 ne devient pas 2)
        if P>0:     items_num.append(("Poids",          f"{P:g}",  "g"))
        # Épaisseur si disponible
        if thickness > 0:
            items_num.append(("Épaisseur",  f"{thickness:g}", "mm"))
        # Partsqty packaging — affichage dédié dans le panel
        parts_q_label = _safe_float(comp.get("parts_qty", 0))
        if parts_q_label > 0:
            items_num.append(("Pièces / packaging", f"{parts_q_label:g}", ""))
        if price>0:
            if parts_q_label > 0:
                items_num.append((f"Prix pour {parts_q_label:.0f} pièces",
                                  f"{price:.2f}", "€"))
            else:
                items_num.append(("Prix", f"{price:.2f}", "€"))
        if vol_m3>0:items_num.append(("Volume",         f"{vol_m3*1e6:.1f}", "cm³"))

    render_info_panel(items_text, items_num, "Informations du Composant")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── ANALYSE PAR TYPE — filtrage TYPE STRICT + DIMS pour tous (carton/sac/palette/cale)
    if cat in ("carton","cale"):
        pool = CAT_DFS.get(cat, df_carton)
        df_pool = filter_same_component(pool, desig, L, W, H, P, cat, use_weight=False)
        _run_dim_analysis(df_pool if not df_pool.empty else pool,
                          pool, L, W, H, P, vol_m3, poids_kg,
                          forced_price=price, forced_ref=comp["ref"],
                          forced_sup=supplier, designation=desig, cat=cat,
                          forced_thickness=thickness,
                          forced_parts_qty=comp.get("parts_qty", 0),
                          country_fab=comp.get("country_fab", ""),
                          country_dispatch=comp.get("country_dispatch", ""))

    elif cat == "sac":
        pool = df_sac
        surface = 2*(L*W+(L+W)*H) if H>0 else 2*L*W
        df_pool = filter_same_component(pool, desig, L, W, H, P, cat, use_weight=False)
        _run_sac_analysis(df_pool if not df_pool.empty else pool,
                          L, W, H, P, surface,
                          forced_price=price, forced_ref=comp["ref"],
                          forced_sup=supplier, designation=desig,
                          forced_thickness=thickness,
                          forced_parts_qty=comp.get("parts_qty", 0),
                          country_fab=comp.get("country_fab", ""),
                          country_dispatch=comp.get("country_dispatch", ""))

    elif cat == "palette":
        pool = df_palette
        df_pool = filter_same_component(pool, desig, L, W, H, P, cat, use_weight=False)
        _run_palette_analysis(df_pool if not df_pool.empty else pool, pool,
                              L, W, H, poids_kg, vol_m3,
                              forced_price=price, forced_ref=comp["ref"],
                              forced_sup=supplier, designation=desig,
                              forced_thickness=thickness,
                              forced_parts_qty=comp.get("parts_qty", 0),
                              country_fab=comp.get("country_fab", ""),
                              country_dispatch=comp.get("country_dispatch", ""))

    elif cat == "mod":
        # MOD : UNIQUEMENT pour le fournisseur de la référence
        # ET le MÊME manufacturing_site + MÊME dispatch_site (sinon comparaison faussée)
        taux = comp.get("labour_rate", 0.)
        _run_mod_for_supplier(df_mod, supplier, taux,
                              display_cost=comp.get("labour_cost", 0.),
                              mfg_site=comp.get("mfg_site", ""),
                              dispatch=comp.get("dispatch", ""))

    elif cat == "transport":
        frais = comp.get("shipping",0.)
        if frais<=0:
            frais = float(df_transport[COL_SHIPPING].dropna().median()) if (
                not df_transport.empty and COL_SHIPPING in df_transport.columns and
                df_transport[COL_SHIPPING].notna().any()) else 50.
        # Filtrage : même fournisseur + site d'enlèvement + site client + partsqty
        _run_transport_analysis(df_transport, frais,
                                forced_supplier=supplier,
                                forced_dispatch=disp_site,
                                forced_client_site=comp.get("client_site", ""),
                                forced_parts_qty=comp.get("parts_qty", 0))

    else:
        # ── Catégorie "autre" : on choisit l'analyse en fonction de la nature
        # • Si c'est une étiquette → on l'analyse via le pool des étiquettes
        # • Sinon (composant packaging non standard) → on tente l'analyse dim générique
        #   en utilisant le pool carton comme référence type (mieux que rien)
        if is_label_designation(desig):
            # Analyse étiquette : pool = df_label, comparaison par prix uniquement
            pool = df_label
            df_pool = pool.copy()
            # On utilise _run_dim_analysis avec cat="autre" pour avoir la même UX
            _run_dim_analysis(df_pool if not df_pool.empty else pool,
                              pool, L, W, H, P, vol_m3, poids_kg,
                              forced_price=price, forced_ref=comp["ref"],
                              forced_sup=supplier, designation=desig, cat="autre",
                              forced_thickness=thickness,
                              forced_parts_qty=comp.get("parts_qty", 0),
                              country_fab=comp.get("country_fab", ""),
                              country_dispatch=comp.get("country_dispatch", ""))
        else:
            # Composant packaging générique : on traite comme un carton générique
            # (utilise le pool df_raw pour rechercher les composants similaires)
            pool_generic = df_raw[df_raw[COL_DESIGNATION].notna()].copy() if not df_raw.empty else pd.DataFrame()
            _run_dim_analysis(pool_generic, pool_generic,
                              L, W, H, P, vol_m3, poids_kg,
                              forced_price=price, forced_ref=comp["ref"],
                              forced_sup=supplier, designation=desig, cat="autre",
                              forced_thickness=thickness,
                              forced_parts_qty=comp.get("parts_qty", 0),
                              country_fab=comp.get("country_fab", ""),
                              country_dispatch=comp.get("country_dispatch", ""))


# ═════════════════════════════════════════════════════════════════════════════
# PAGE — ANALYSE LIBRE
# ═════════════════════════════════════════════════════════════════════════════

def render_analysis():
    scroll_top()
    cat = st.session_state.category or "carton"

    if st.button("← Retour à l'Accueil", key="bk_home_an"):
        go_page("home"); st.rerun()

    st.markdown(f"<div class='section-header'><h2>{CAT_LABELS.get(cat,cat)}</h2>"
                "<p>Saisissez les caractéristiques pour une analyse libre · prix en €</p></div>",
                unsafe_allow_html=True)

    dispatch = {
        "carton":    lambda: _render_dim_ui(df_carton, "carton"),
        "sac":       lambda: _render_sac_ui(df_sac),
        "palette":   lambda: _render_palette_ui(df_palette),
        "cale":      lambda: _render_dim_ui(df_cale, "cale"),
        "mod":       lambda: _render_mod_ui(df_mod),
        "transport": lambda: _render_transport_ui(df_transport),
    }
    dispatch.get(cat, lambda: st.info("Catégorie non disponible."))()


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSE — DIMENSIONS  (Carton + Cale)
# ═════════════════════════════════════════════════════════════════════════════

def _render_dim_ui(df_cat, cat="carton"):
    cl,cr = st.columns([2.2,1.4], gap="large")
    designation = ""
    with cr:
        st.markdown('<div class="config-section"><h3>⚙️ Configuration</h3>', unsafe_allow_html=True)
        mode = st.radio("Mode de Saisie",["Saisie Manuelle","Par Référence"],horizontal=False, key=f"mode_{cat}")
        if mode == "Saisie Manuelle":
            L=st.number_input("Longueur (mm)",value=200.,step=10.,min_value=1., key=f"L_{cat}")
            W=st.number_input("Largeur (mm)", value=150.,step=10.,min_value=1., key=f"W_{cat}")
            H=st.number_input("Hauteur (mm)", value=100.,step=10.,min_value=1., key=f"H_{cat}")
            P=st.number_input("Poids (g)",    value=50., step=5., min_value=1., key=f"P_{cat}")
            thickness_in=st.number_input("Épaisseur (mm)", value=0., step=0.5, min_value=0., key=f"T_{cat}",
                                         help="0 = ne pas filtrer par épaisseur")
        else:
            refs = sorted(df_cat[COL_REF].unique().tolist()) if not df_cat.empty else ["—"]
            rs   = st.selectbox("Référence", refs, key=f"ref_{cat}")
            rr   = df_cat[df_cat[COL_REF]==rs]
            if not rr.empty:
                r=rr.iloc[0]
                L=float(r[COL_LENGTH]) if pd.notna(r[COL_LENGTH]) else 200.
                W=float(r[COL_WIDTH])  if pd.notna(r[COL_WIDTH])  else 150.
                H=float(r[COL_HEIGHT]) if pd.notna(r[COL_HEIGHT]) else 100.
                P=float(r[COL_WEIGHT]) if pd.notna(r[COL_WEIGHT]) else 50.
                designation = str(r.get(COL_DESIGNATION,""))
            else: L=W=H=P=0.
        vm  = (L*W*H)/1e9; pk=P/1000
        st.markdown(f'<div class="dim-display">{L:.0f} × {W:.0f} × {H:.0f} mm<br>'
                    f'<span style="font-size:12px;opacity:.75;">Poids : {P:.0f} g &nbsp;|&nbsp; '
                    f'Volume : {vm*1e6:.1f} cm³</span></div>', unsafe_allow_html=True)
        if st.button("🚀 Lancer l'Analyse",use_container_width=True,type="primary", key=f"go_{cat}"):
            st.session_state.run_analysis=True
        st.markdown('</div>', unsafe_allow_html=True)

    with cl:
        if not st.session_state.run_analysis:
            st.markdown('<div class="info-box">Saisissez les caractéristiques puis cliquez sur '
                        "<b>Lancer l'Analyse</b> pour voir tous les fournisseurs proposant des "
                        "composants similaires.</div>", unsafe_allow_html=True)
            return
        df_pool = filter_same_component(df_cat, designation, L, W, H, P, cat) if designation \
                  else _filter_dims_only(df_cat, L, W, H, P)
        if df_pool.empty:
            df_pool = _filter_dims_only(df_cat, L, W, H, P)
        _run_dim_analysis(df_pool, df_cat, L, W, H, P, vm, pk,
                          designation=designation, cat=cat,
                          forced_thickness=float(st.session_state.get(f"T_{cat}", 0.0) or 0.0))


def _filter_dims_only(pool: pd.DataFrame, L, W, H, P,
                      tol_dim: float = 0.05, tol_w: float = 0.15) -> pd.DataFrame:
    """Filtrage par dimensions seulement (mode saisie manuelle sans désignation)."""
    if pool is None or pool.empty:
        return pd.DataFrame()
    df = pool.copy()
    if L > 0 and W > 0 and COL_LENGTH in df.columns and COL_WIDTH in df.columns:
        df = df[
            df[COL_LENGTH].between(L*(1-tol_dim), L*(1+tol_dim)) &
            df[COL_WIDTH].between(W*(1-tol_dim), W*(1+tol_dim))
        ]
        if H > 0 and COL_HEIGHT in df.columns:
            dfh = df[df[COL_HEIGHT].between(H*(1-tol_dim), H*(1+tol_dim))]
            if not dfh.empty:
                df = dfh
    if P > 0 and COL_WEIGHT in df.columns:
        dfw = df[df[COL_WEIGHT].between(P*(1-tol_w), P*(1+tol_w))]
        if not dfw.empty:
            df = dfw
    return df


def _run_dim_analysis(df_pool, df_cat, L, W, H, P, vol_m3, poids_kg,
                      forced_price=None, forced_ref=None, forced_sup=None,
                      designation=None, cat="carton", forced_thickness=0.0,
                      forced_parts_qty=None,
                      country_fab="", country_dispatch=""):
    """
    Analyse d'un composant — TYPE STRICT + MÊMES dimensions + MÊME poids
    + MÊME PAYS (fabrication / expédition) quand l'info est disponible.
    Tableau, courbe, stats, statut et best price : tous calculés sur le MÊME pool exact.
    forced_parts_qty : partsqty packaging du composant analysé — affiché dans les KPI
    """
    # ── 1) Détermination du prix affiché
    is_real = False
    if forced_price is not None and forced_price > 0:
        price = forced_price; ref_r = forced_ref; sup_r = forced_sup; is_real = True
    else:
        # Mode analyse libre : on cherche d'abord une correspondance exacte
        pool_cat = CAT_DFS.get(cat, df_carton)
        df_exact_search = filter_exact_dims(pool_cat, designation or "", L, W, H, P, cat,
                                            thickness=forced_thickness)
        if not df_exact_search.empty:
            r = df_exact_search.iloc[0]
            price = float(r[COL_PRICE]); ref_r = r[COL_REF]; sup_r = r[COL_SUPPLIER]
            is_real = True
        else:
            # Pas de correspondance exacte : on tente une estimation ML par TYPE
            est_price, method, n_sample = estimate_price_by_type(
                designation or "", cat, L, W, H, P, thickness=forced_thickness
            )
            if est_price > 0:
                price = est_price; ref_r = None
                sup_r = f"Estimation IA (n={n_sample})" if method == "ml_type" else f"Médiane type (n={n_sample})"
            else:
                price = 0.; ref_r = None; sup_r = "Données insuffisantes"

    # ── 2) Pool de référence UNIQUE : MÊME TYPE + MÊMES DIMS + MÊME POIDS
    #        + MÊME PAYS (fabrication / expédition) si l'info est disponible.
    pool_cat = CAT_DFS.get(cat, df_carton)
    df_exact = filter_exact_dims(pool_cat, designation or "", L, W, H, P, cat,
                                 thickness=forced_thickness,
                                 country_fab=country_fab,
                                 country_dispatch=country_dispatch)
    df_exact = _dedup_pool(df_exact, COL_PRICE)
    has_real_comparison = (not df_exact.empty and len(df_exact) >= 2)

    # ── 3) ESTIMATION IA (toujours calculée — utilisée comme fallback de référence)
    est_price, est_method, est_n_sample = estimate_price_by_type(
        designation or "", cat, L, W, H, P, thickness=forced_thickness
    )

    # ── 4) Détermination du STATUT (bon / moyen / élevé) :
    #    • Si comparatif RÉEL ≥ 2 → comparaison contre le pool exact
    #    • Sinon (pas/peu de comparatif) → comparaison contre L'ESTIMATION IA
    if has_real_comparison:
        status, _ = eval_status_dim(price, designation or "", L, W, H, P, cat,
                                    thickness=forced_thickness)
        status_source = "real"
    elif est_price > 0 and price > 0:
        status, _ = evaluate_price_vs_estimate(price, est_price)
        status_source = "ml"
    else:
        status = "neutre"
        status_source = "none"

    # ── 5) KPI principal avec le statut déterminé ci-dessus
    # Label : on précise "Prix pour N pièces" si on connaît la partsqty (au lieu
    # de "Prix Trouvé") — pour distinguer ce prix du prix unitaire affiché sur
    # les cartes de la page d'accueil (qui est par pièce).
    if forced_parts_qty is not None and forced_parts_qty > 0:
        kpi_label = f"Prix pour {forced_parts_qty:.0f} pièces"
    else:
        kpi_label = "Prix Trouvé" if is_real else "Prix Estimé"
    # Best price = médiane du pool exact (sur valeurs uniques) si comparatif réel
    _median_kpi = None
    _gain_kpi = None
    try:
        if has_real_comparison and not df_exact.empty and COL_PRICE in df_exact.columns:
            _vp = df_exact[COL_PRICE].dropna(); _vp = _vp[_vp > 0]
            if not _vp.empty:
                _median_kpi = float(pd.Series(sorted(_vp.unique())).median())
                if price > _median_kpi:
                    # Gain potentiel = écart ramené à la pièce si partsqty connu
                    if forced_parts_qty and forced_parts_qty > 0:
                        _gain_pp = (price - _median_kpi) / forced_parts_qty
                        _gain_kpi = f"{_gain_pp:.2f} €/pièce"
                    else:
                        _gain_kpi = f"{price - _median_kpi:.2f} €"
    except Exception:
        _median_kpi = None

    kpi_block(price,
              kpi_label,
              is_real,
              kpi_class=kpi_class_for_status(status),
              status=status,
              median_val=_median_kpi,
              gain_text=_gain_kpi)

    # ── 6) MODE ESTIMATION IA — quand pas de comparatif réel
    # Le modèle IA fournit À LA FOIS :
    #   • Le statut (déjà calculé via evaluate_price_vs_estimate)
    #   • Le "best price" (= la prédiction IA elle-même)
    if not has_real_comparison:
        if est_price > 0:
            method_label = (
                f"XGBoost — {est_n_sample} composants du même type"
                if est_method == "ml_type"
                else f"Médiane du type — {est_n_sample} composants"
            )

            # Message adapté au statut détecté
            if status == "bon":
                status_msg = "✅ Votre prix est <b>bon</b> : il est au niveau ou en-dessous de la prédiction du modèle IA."
                bg_color = "linear-gradient(135deg,#ecfdf5,#d1fae5)"; border_color = "#10b981"; text_color = "#065f46"
            elif status == "moyen":
                status_msg = "🟡 Votre prix est <b>légèrement au-dessus</b> de la prédiction du modèle IA (écart moyen)."
                bg_color = "linear-gradient(135deg,#fef9c3,#fef3c7)"; border_color = "#eab308"; text_color = "#854d0e"
            elif status == "élevé":
                status_msg = "🔴 Votre prix est <b>significativement au-dessus</b> de la prédiction du modèle IA (écart élevé)."
                bg_color = "linear-gradient(135deg,#fef2f2,#fee2e2)"; border_color = "#ef4444"; text_color = "#991b1b"
            else:
                status_msg = ""
                bg_color = "linear-gradient(135deg,#eff6ff,#dbeafe)"; border_color = "#3b82f6"; text_color = "#1e3a8a"

            # Bloc estimation IA — message simplifié : juste le Best Price estimé
            st.markdown(
                f'<div class="info-box" style="background:{bg_color};'
                f'border-left-color:{border_color};color:{text_color};">'
                f'🤖 <b>Best Price estimé par le modèle IA :</b> '
                f'<b style="color:{text_color};font-size:22px;">{est_price:.2f} €</b>'
                f'</div>',
                unsafe_allow_html=True
            )

            # ── Tableau des références qui ont aidé à calculer cette estimation
            # (transparence : l'utilisateur voit la marge de prix réelle dans la base)
            try:
                train_pool = get_ai_training_pool(designation or "", cat)
                if not train_pool.empty:
                    train_show = train_pool[[c for c in [
                        COL_REF, COL_DESIGNATION, COL_SUPPLIER, COL_PRICE,
                        COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT, COL_THICKNESS
                    ] if c in train_pool.columns]].copy()
                    # Trier par prix croissant pour mieux visualiser la marge
                    if COL_PRICE in train_show.columns:
                        train_show = train_show.sort_values(COL_PRICE).reset_index(drop=True)
                    # Renommer pour la lisibilité
                    rename_map = {
                        COL_REF: "Référence",
                        COL_DESIGNATION: "Composant",
                        COL_SUPPLIER: "Fournisseur",
                        COL_PRICE: "Prix (€)",
                        COL_LENGTH: "L (mm)",
                        COL_WIDTH: "l (mm)",
                        COL_HEIGHT: "H (mm)",
                        COL_WEIGHT: "Poids (g)",
                        COL_THICKNESS: "Épaisseur (mm)",
                    }
                    train_show = train_show.rename(columns={k: v for k, v in rename_map.items()
                                                             if k in train_show.columns})
                    n_train = len(train_show)
                    st.markdown(
                        f"<h4 style='color:#0f172a;font-weight:800;margin:14px 0 6px 0;font-size:15px;'>"
                        f"📚 Références qui ont aidé à calculer cette estimation "
                        f"<span style='color:#64748b;font-weight:500;font-size:12px;'>"
                        f"— {n_train} composant(s) du même type</span></h4>",
                        unsafe_allow_html=True
                    )
                    st.dataframe(train_show, use_container_width=True, hide_index=True,
                                 height=min(380, 38 + 35 * min(n_train, 10)),
                                 column_config={
                                     "Prix (€)": st.column_config.NumberColumn(format="%.2f"),
                                     "L (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "l (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "H (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "Poids (g)": st.column_config.NumberColumn(format="%.0f"),
                                     "Épaisseur (mm)": st.column_config.NumberColumn(format="%g"),
                                 })
            except Exception:
                pass
        else:
            st.markdown(
                '<div class="info-box" style="background:#f1f5f9;border-left-color:#94a3b8;">'
                '⚪ <b>Pas de comparatif disponible</b> — données insuffisantes dans la base '
                'pour ce type de composant (impossible d\'entraîner le modèle IA).</div>',
                unsafe_allow_html=True
            )
        return

    # ── 7) MODE COMPARATIF RÉEL — comparaison contre les composants similaires en base
    # (best_price gardé en variable pour l'annotation de la courbe, mais pas affiché en bloc)
    best_price = float(df_exact[COL_PRICE].min())
    median_price = float(df_exact[COL_PRICE].median())  # Stat de cohérence (médiane = robuste aux outliers)

    # ── 6) FILTRE FOURNISSEUR optionnel — placé AVANT les stats pour que
    # le prix min/médiane/max se recalcule selon le fournisseur sélectionné.
    sup_opts = sorted(df_exact[COL_SUPPLIER].dropna().unique().tolist()) \
               if COL_SUPPLIER in df_exact.columns else []
    if sup_opts:
        sup_sel = st.selectbox("Filtrer par Fournisseur",
                               ["Tous les Fournisseurs"] + sup_opts,
                               index=0, key=f"dim_sup_{cat}")
    else:
        sup_sel = "Tous les Fournisseurs"

    df_show = df_exact.copy()
    if sup_sel and sup_sel != "Tous les Fournisseurs" and COL_SUPPLIER in df_show.columns:
        df_show = df_show[df_show[COL_SUPPLIER] == sup_sel]

    # ── 7) STATS — calculées sur df_show (= pool filtré par fournisseur)
    st.markdown("<br>", unsafe_allow_html=True)
    _stat_lbl = "Mêmes type + dimensions + poids" + (f" · {sup_sel}" if sup_sel != "Tous les Fournisseurs" else "")
    stat_cards_from_df(df_show, COL_PRICE, _stat_lbl)

    # ── 8) COURBES — Sélection intelligente :
    # • Si les composants comparés ont TOUS les mêmes dimensions ET le même poids
    #   → UN SEUL graphique (Prix par référence), car Volume vs Prix et Poids vs Prix
    #   seraient identiques (un point répété)
    # • Sinon → 2 onglets Volume vs Prix et Poids vs Prix comme avant
    if not df_show.empty and COL_PRICE in df_show.columns:
        df_chart = df_show.copy()
        if "Volume_m3" not in df_chart.columns and all(c in df_chart.columns for c in [COL_LENGTH, COL_WIDTH, COL_HEIGHT]):
            df_chart["Volume_m3"] = (
                df_chart[COL_LENGTH].fillna(0) *
                df_chart[COL_WIDTH].fillna(0) *
                df_chart[COL_HEIGHT].fillna(0)
            ) / 1e9
        if "weight_kg" not in df_chart.columns and COL_WEIGHT in df_chart.columns:
            df_chart["weight_kg"] = df_chart[COL_WEIGHT] / 1000.0

        # Toujours un graphique simple par fournisseur (cohérent sur tous les composants
        # d'une référence — étiquettes incluses).
        st.markdown("<h3 style='color:#0f172a;font-weight:800;margin-bottom:12px;'>"
                    "Graphique d'Analyse</h3>", unsafe_allow_html=True)
        fig = build_scatter_by_reference(
            df_chart, COL_PRICE, "Prix (€)",
            "Prix par fournisseur (mêmes dimensions, même poids)",
            ref_price=price, best_price=best_price,
            target_supplier=sup_r,
        )
        st.plotly_chart(fig, use_container_width=True)

    # ── 9) TABLEAU — même pool que la courbe, mêmes lignes affichées
    st.markdown("<br>", unsafe_allow_html=True)
    title_tbl = f"Composants Similaires — {sup_sel}" if sup_sel != "Tous les Fournisseurs" \
                else "Composants Similaires — Tous fournisseurs"
    show_table(df_show.sort_values(COL_PRICE) if COL_PRICE in df_show.columns else df_show,
               title_tbl)


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSE — SAC
# ═════════════════════════════════════════════════════════════════════════════

def _render_sac_ui(df_sac):
    cl,cr = st.columns([2.2,1.4],gap="large")
    designation = ""
    with cr:
        st.markdown('<div class="config-section"><h3>⚙️ Configuration</h3>',unsafe_allow_html=True)
        mode=st.radio("Mode de Saisie",["Saisie Manuelle","Par Référence"],horizontal=False, key="mode_sac")
        if mode=="Saisie Manuelle":
            L=st.number_input("Longueur (mm)",value=300.,step=10.,min_value=1., key="L_sac")
            W=st.number_input("Largeur (mm)", value=200.,step=10.,min_value=1., key="W_sac")
            H=st.number_input("Hauteur (mm)", value=0.,  step=1., min_value=0., key="H_sac")
            P=st.number_input("Poids (g)",    value=20., step=5., min_value=0., key="P_sac")
            thickness_in=st.number_input("Épaisseur (mm)", value=0., step=0.05, min_value=0., key="T_sac",
                                         help="0 = ne pas filtrer par épaisseur")
        else:
            refs=sorted(df_sac[COL_REF].unique().tolist()) if not df_sac.empty else ["—"]
            rs=st.selectbox("Référence",refs, key="ref_sac"); rr=df_sac[df_sac[COL_REF]==rs]
            if not rr.empty:
                r=rr.iloc[0]; L=_safe_float(r.get(COL_LENGTH,0))
                W=_safe_float(r.get(COL_WIDTH,0)); H=_safe_float(r.get(COL_HEIGHT,0))
                P=_safe_float(r.get(COL_WEIGHT,0))
                designation = str(r.get(COL_DESIGNATION,""))
            else: L=W=H=P=0.
        surf=2*(L*W+L*H+W*H)
        st.markdown(f'<div class="dim-display">{L:.0f} × {W:.0f} mm<br>'
                    f'<span style="font-size:12px;opacity:.75;">Poids : {P:.0f} g &nbsp;|&nbsp; '
                    f'Surface : {surf/1e4:.1f} cm²</span></div>',unsafe_allow_html=True)
        if st.button("🚀 Lancer l'Analyse",use_container_width=True,type="primary", key="go_sac"):
            st.session_state.run_analysis=True
        st.markdown('</div>', unsafe_allow_html=True)

    with cl:
        if not st.session_state.run_analysis:
            st.markdown('<div class="info-box">Saisissez les caractéristiques puis cliquez sur '
                        "<b>Lancer l'Analyse</b>.</div>", unsafe_allow_html=True)
            return
        df_pool = filter_same_component(df_sac, designation, L, W, H, P, "sac") if designation \
                  else _filter_dims_only(df_sac, L, W, H, P)
        if df_pool.empty:
            df_pool = _filter_dims_only(df_sac, L, W, H, P)
        _run_sac_analysis(df_pool if not df_pool.empty else df_sac,
                          L, W, H, P, surf, designation=designation,
                          forced_thickness=float(st.session_state.get("T_sac", 0.0) or 0.0))


def _run_sac_analysis(df_pool, L, W, H, P, surface_mm2,
                      forced_price=None, forced_ref=None, forced_sup=None, designation=None,
                      forced_thickness=0.0, forced_parts_qty=None,
                      country_fab="", country_dispatch=""):
    """Analyse SAC — pool unique : type strict + mêmes dims + même poids + même pays."""
    # 1) Prix affiché
    is_real = False
    if forced_price is not None and forced_price > 0:
        price = forced_price; ref_r = forced_ref; sup_r = forced_sup; is_real = True
    else:
        df_exact_search = filter_exact_dims(df_sac, designation or "", L, W, H, P, "sac",
                                            thickness=forced_thickness)
        if not df_exact_search.empty:
            r = df_exact_search.iloc[0]
            price = float(r[COL_PRICE]); ref_r = r[COL_REF]; sup_r = r[COL_SUPPLIER]
            is_real = True
        else:
            est_p, _, _ = estimate_price_by_type(designation or "", "sac", L, W, H, P,
                                                  thickness=forced_thickness)
            price = est_p if est_p > 0 else 0.
            ref_r = None; sup_r = "Estimation IA"

    # 2) Pool unique (+ même pays)
    df_exact = filter_exact_dims(df_sac, designation or "", L, W, H, P, "sac",
                                 thickness=forced_thickness,
                                 country_fab=country_fab, country_dispatch=country_dispatch)
    df_exact = _dedup_pool(df_exact, COL_PRICE)
    has_real_comparison = (not df_exact.empty and len(df_exact) >= 2)

    # 3) ESTIMATION IA (toujours calculée)
    est_price, est_method, est_n_sample = estimate_price_by_type(
        designation or "", "sac", L, W, H, P, thickness=forced_thickness
    )

    # 4) Statut : comparatif réel sinon ML
    if has_real_comparison:
        status, _ = eval_status_dim(price, designation or "", L, W, H, P, "sac",
                                    thickness=forced_thickness)
    elif est_price > 0 and price > 0:
        status, _ = evaluate_price_vs_estimate(price, est_price)
    else:
        status = "neutre"

    kpi_label = (f"Prix pour {forced_parts_qty:.0f} pièces"
                 if forced_parts_qty is not None and forced_parts_qty > 0
                 else ("Prix Trouvé" if is_real else "Prix Estimé"))
    _med_kpi = None; _gain_kpi = None
    try:
        if has_real_comparison and not df_exact.empty and COL_PRICE in df_exact.columns:
            _vp = df_exact[COL_PRICE].dropna(); _vp = _vp[_vp > 0]
            if not _vp.empty:
                _med_kpi = float(pd.Series(sorted(_vp.unique())).median())
                if price > _med_kpi:
                    if forced_parts_qty and forced_parts_qty > 0:
                        _gain_kpi = f"{(price - _med_kpi)/forced_parts_qty:.2f} €/pièce"
                    else:
                        _gain_kpi = f"{price - _med_kpi:.2f} €"
    except Exception:
        _med_kpi = None
    kpi_block(price, kpi_label, is_real,
              kpi_class=kpi_class_for_status(status), status=status,
              median_val=_med_kpi, gain_text=_gain_kpi)

    # 5) Mode estimation IA si pas de comparatif réel
    if not has_real_comparison:
        if est_price > 0:
            method_label = (f"XGBoost — {est_n_sample} sacs du même type"
                            if est_method == "ml_type"
                            else f"Médiane du type — {est_n_sample} sacs")
            if status == "bon":
                status_msg = "✅ Votre prix est <b>bon</b> : au niveau ou en-dessous de la prédiction IA."
                bg_color = "linear-gradient(135deg,#ecfdf5,#d1fae5)"; border_color = "#10b981"; text_color = "#065f46"
            elif status == "moyen":
                status_msg = "🟡 Votre prix est <b>légèrement au-dessus</b> de la prédiction IA (écart moyen)."
                bg_color = "linear-gradient(135deg,#fef9c3,#fef3c7)"; border_color = "#eab308"; text_color = "#854d0e"
            elif status == "élevé":
                status_msg = "🔴 Votre prix est <b>significativement au-dessus</b> de la prédiction IA (écart élevé)."
                bg_color = "linear-gradient(135deg,#fef2f2,#fee2e2)"; border_color = "#ef4444"; text_color = "#991b1b"
            else:
                status_msg = ""
                bg_color = "linear-gradient(135deg,#eff6ff,#dbeafe)"; border_color = "#3b82f6"; text_color = "#1e3a8a"

            # Best Price estimé — message simplifié + tableau des références d'apprentissage
            st.markdown(
                f'<div class="info-box" style="background:{bg_color};'
                f'border-left-color:{border_color};color:{text_color};">'
                f'🤖 <b>Best Price estimé par le modèle IA :</b> '
                f'<b style="color:{text_color};font-size:22px;">{est_price:.2f} €</b>'
                f'</div>',
                unsafe_allow_html=True
            )
            try:
                train_pool = get_ai_training_pool(designation or "", "sac")
                if not train_pool.empty:
                    train_show = train_pool[[c for c in [
                        COL_REF, COL_DESIGNATION, COL_SUPPLIER, COL_PRICE,
                        COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT, COL_THICKNESS
                    ] if c in train_pool.columns]].copy()
                    if COL_PRICE in train_show.columns:
                        train_show = train_show.sort_values(COL_PRICE).reset_index(drop=True)
                    rename_map = {COL_REF: "Référence", COL_DESIGNATION: "Composant",
                                  COL_SUPPLIER: "Fournisseur", COL_PRICE: "Prix (€)",
                                  COL_LENGTH: "L (mm)", COL_WIDTH: "l (mm)",
                                  COL_HEIGHT: "H (mm)", COL_WEIGHT: "Poids (g)",
                                  COL_THICKNESS: "Épaisseur (mm)"}
                    train_show = train_show.rename(columns={k: v for k, v in rename_map.items()
                                                             if k in train_show.columns})
                    n_train = len(train_show)
                    st.markdown(
                        f"<h4 style='color:#0f172a;font-weight:800;margin:14px 0 6px 0;font-size:15px;'>"
                        f"📚 Références qui ont aidé à calculer cette estimation "
                        f"<span style='color:#64748b;font-weight:500;font-size:12px;'>"
                        f"— {n_train} composant(s) du même type</span></h4>",
                        unsafe_allow_html=True
                    )
                    st.dataframe(train_show, use_container_width=True, hide_index=True,
                                 height=min(380, 38 + 35 * min(n_train, 10)),
                                 column_config={
                                     "Prix (€)": st.column_config.NumberColumn(format="%.2f"),
                                     "L (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "l (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "H (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "Poids (g)": st.column_config.NumberColumn(format="%.0f"),
                                     "Épaisseur (mm)": st.column_config.NumberColumn(format="%g"),
                                 })
            except Exception:
                pass
        else:
            st.markdown('<div class="info-box" style="background:#f1f5f9;border-left-color:#94a3b8;">'
                        '⚪ <b>Pas de comparatif disponible</b> — données insuffisantes pour entraîner le modèle IA.</div>',
                        unsafe_allow_html=True)
        return

    # 6) Variables prix (gardées pour annotations chart, bloc Meilleur prix retiré)
    best_price = float(df_exact[COL_PRICE].min())
    median_price = float(df_exact[COL_PRICE].median())  # Stat de cohérence (médiane = robuste aux outliers)

    # 6) Filtre fournisseur AVANT les stats (pour recalcul min/médiane/max)
    sup_opts = sorted(df_exact[COL_SUPPLIER].dropna().unique().tolist()) \
               if COL_SUPPLIER in df_exact.columns else []
    if sup_opts:
        sup_sel = st.selectbox("Filtrer par Fournisseur",
                               ["Tous les Fournisseurs"] + sup_opts,
                               index=0, key="sac_sup_filter")
    else:
        sup_sel = "Tous les Fournisseurs"

    df_show = df_exact.copy()
    if sup_sel and sup_sel != "Tous les Fournisseurs" and COL_SUPPLIER in df_show.columns:
        df_show = df_show[df_show[COL_SUPPLIER] == sup_sel]

    # 7) Stats sur le pool filtré
    st.markdown("<br>", unsafe_allow_html=True)
    _lbl = "Mêmes type + dimensions + poids" + (f" · {sup_sel}" if sup_sel != "Tous les Fournisseurs" else "")
    stat_cards_from_df(df_show, COL_PRICE, _lbl)

    # 8) Courbes — Sélection intelligente (1 graphique si dims+poids identiques, sinon 2 onglets)
    if not df_show.empty:
        df_chart = df_show.copy()
        if "Surface_mm2" not in df_chart.columns and all(c in df_chart.columns for c in [COL_LENGTH, COL_WIDTH, COL_HEIGHT]):
            Lc = df_chart[COL_LENGTH].fillna(0); Wc = df_chart[COL_WIDTH].fillna(0); Hc = df_chart[COL_HEIGHT].fillna(0)
            df_chart["Surface_mm2"] = 2*(Lc*Wc + (Lc+Wc)*Hc)
        if "weight_kg" not in df_chart.columns and COL_WEIGHT in df_chart.columns:
            df_chart["weight_kg"] = df_chart[COL_WEIGHT] / 1000.0

        # Toujours un graphique simple par fournisseur
        st.markdown("<h3 style='color:#0f172a;font-weight:800;margin-bottom:12px;'>"
                    "Graphique d'Analyse</h3>", unsafe_allow_html=True)
        fig = build_scatter_by_reference(
            df_chart, COL_PRICE, "Prix (€)",
            "Prix par fournisseur (mêmes dimensions, même poids)",
            ref_price=price, best_price=best_price,
            target_supplier=sup_r,
        )
        st.plotly_chart(fig, use_container_width=True)

    # 9) Tableau
    st.markdown("<br>", unsafe_allow_html=True)
    show_table(df_show.sort_values(COL_PRICE) if COL_PRICE in df_show.columns else df_show,
               f"Sacs Similaires — {sup_sel}")


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSE — PALETTE
# ═════════════════════════════════════════════════════════════════════════════

def _render_palette_ui(df_pal):
    cl,cr=st.columns([2.2,1.4],gap="large")
    designation = ""
    with cr:
        st.markdown('<div class="config-section"><h3>⚙️ Configuration</h3>',unsafe_allow_html=True)
        mode=st.radio("Mode de Saisie",["Saisie Manuelle","Par Référence"],horizontal=False, key="mode_pal")
        if mode=="Saisie Manuelle":
            L=st.number_input("Longueur (mm)",value=1200.,step=50.,min_value=1., key="L_pal")
            W=st.number_input("Largeur (mm)", value=1000.,step=50.,min_value=1., key="W_pal")
            H=st.number_input("Hauteur (mm)", value=150., step=10.,min_value=1., key="H_pal")
            P=st.number_input("Poids (kg)",   value=25.,  step=1., min_value=0., key="P_pal")
            thickness_in=st.number_input("Épaisseur (mm)", value=0., step=1., min_value=0., key="T_pal",
                                         help="0 = ne pas filtrer par épaisseur")
            P_g = P*1000
        else:
            refs=sorted(df_pal[COL_REF].unique().tolist()) if not df_pal.empty else ["—"]
            rs=st.selectbox("Référence",refs, key="ref_pal"); rr=df_pal[df_pal[COL_REF]==rs]
            if not rr.empty:
                r=rr.iloc[0]
                L=_safe_float(r.get(COL_LENGTH,1200)) or 1200.
                W=_safe_float(r.get(COL_WIDTH,1000))  or 1000.
                H=_safe_float(r.get(COL_HEIGHT,150))  or 150.
                P_g=_safe_float(r.get(COL_WEIGHT,0))
                P=P_g/1000
                designation = str(r.get(COL_DESIGNATION,""))
            else: L,W,H,P,P_g=1200.,1000.,150.,25.,25000.
        vm=(L*W*H)/1e9
        st.markdown(f'<div class="dim-display">{L:.0f} × {W:.0f} × {H:.0f} mm<br>'
                    f'<span style="font-size:12px;opacity:.75;">Poids : {P:.1f} kg &nbsp;|&nbsp; '
                    f'Volume : {vm*1000:.2f} dm³</span></div>',unsafe_allow_html=True)
        if st.button("🚀 Lancer l'Analyse",use_container_width=True,type="primary", key="go_pal"):
            st.session_state.run_analysis=True
        st.markdown('</div>', unsafe_allow_html=True)

    with cl:
        if not st.session_state.run_analysis:
            st.markdown('<div class="info-box">Saisissez les caractéristiques puis cliquez sur '
                        "<b>Lancer l'Analyse</b>.</div>", unsafe_allow_html=True)
            return
        df_pool = filter_same_component(df_pal, designation, L, W, H, P_g, "palette") if designation \
                  else _filter_dims_only(df_pal, L, W, H, P_g)
        if df_pool.empty:
            df_pool = df_pal.copy()
        _run_palette_analysis(df_pool if not df_pool.empty else df_pal,
                              df_pal, L, W, H, P, vm, designation=designation,
                              forced_thickness=float(st.session_state.get("T_pal", 0.0) or 0.0))


def _run_palette_analysis(df_pool, df_all, L, W, H, P_kg, vol_m3,
                          forced_price=None, forced_ref=None, forced_sup=None, designation=None,
                          forced_thickness=0.0, forced_parts_qty=None,
                          country_fab="", country_dispatch=""):
    """Analyse PALETTE — pool unique : type strict + mêmes dims + même poids + même pays."""
    P_g = P_kg * 1000.0

    # 1) Prix affiché
    is_real = False
    if forced_price is not None and forced_price > 0:
        price = forced_price; ref_r = forced_ref; sup_r = forced_sup; is_real = True
    else:
        df_exact_search = filter_exact_dims(df_all, designation or "", L, W, H, P_g, "palette",
                                            thickness=forced_thickness)
        if not df_exact_search.empty:
            r = df_exact_search.iloc[0]
            price = float(r[COL_PRICE]); ref_r = r[COL_REF]; sup_r = r[COL_SUPPLIER]
            is_real = True
        else:
            est_p, _, _ = estimate_price_by_type(designation or "", "palette", L, W, H, P_g,
                                                  thickness=forced_thickness)
            price = est_p if est_p > 0 else 0.
            ref_r = None; sup_r = "Estimation IA"

    # 2) Pool unique (+ même pays)
    df_exact = filter_exact_dims(df_all, designation or "", L, W, H, P_g, "palette",
                                 thickness=forced_thickness,
                                 country_fab=country_fab, country_dispatch=country_dispatch)
    df_exact = _dedup_pool(df_exact, COL_PRICE)
    has_real_comparison = (not df_exact.empty and len(df_exact) >= 2)

    # 3) ESTIMATION IA (toujours calculée)
    est_price, est_method, est_n_sample = estimate_price_by_type(
        designation or "", "palette", L, W, H, P_g, thickness=forced_thickness
    )

    # 4) Statut : comparatif réel sinon ML
    if has_real_comparison:
        status, _ = eval_status_dim(price, designation or "", L, W, H, P_g, "palette",
                                    thickness=forced_thickness)
    elif est_price > 0 and price > 0:
        status, _ = evaluate_price_vs_estimate(price, est_price)
    else:
        status = "neutre"

    kpi_label = (f"Prix pour {forced_parts_qty:.0f} pièces"
                 if forced_parts_qty is not None and forced_parts_qty > 0
                 else ("Prix Trouvé" if is_real else "Prix Estimé"))
    _med_kpi = None; _gain_kpi = None
    try:
        if has_real_comparison and not df_exact.empty and COL_PRICE in df_exact.columns:
            _vp = df_exact[COL_PRICE].dropna(); _vp = _vp[_vp > 0]
            if not _vp.empty:
                _med_kpi = float(pd.Series(sorted(_vp.unique())).median())
                if price > _med_kpi:
                    if forced_parts_qty and forced_parts_qty > 0:
                        _gain_kpi = f"{(price - _med_kpi)/forced_parts_qty:.2f} €/pièce"
                    else:
                        _gain_kpi = f"{price - _med_kpi:.2f} €"
    except Exception:
        _med_kpi = None
    kpi_block(price, kpi_label, is_real,
              kpi_class=kpi_class_for_status(status), status=status,
              median_val=_med_kpi, gain_text=_gain_kpi)

    # 5) Mode estimation IA si pas de comparatif réel
    if not has_real_comparison:
        if est_price > 0:
            method_label = (f"XGBoost — {est_n_sample} palettes du même type"
                            if est_method == "ml_type"
                            else f"Médiane du type — {est_n_sample} palettes")
            if status == "bon":
                status_msg = "✅ Votre prix est <b>bon</b> : au niveau ou en-dessous de la prédiction IA."
                bg_color = "linear-gradient(135deg,#ecfdf5,#d1fae5)"; border_color = "#10b981"; text_color = "#065f46"
            elif status == "moyen":
                status_msg = "🟡 Votre prix est <b>légèrement au-dessus</b> de la prédiction IA (écart moyen)."
                bg_color = "linear-gradient(135deg,#fef9c3,#fef3c7)"; border_color = "#eab308"; text_color = "#854d0e"
            elif status == "élevé":
                status_msg = "🔴 Votre prix est <b>significativement au-dessus</b> de la prédiction IA (écart élevé)."
                bg_color = "linear-gradient(135deg,#fef2f2,#fee2e2)"; border_color = "#ef4444"; text_color = "#991b1b"
            else:
                status_msg = ""
                bg_color = "linear-gradient(135deg,#eff6ff,#dbeafe)"; border_color = "#3b82f6"; text_color = "#1e3a8a"

            # Best Price estimé — message simplifié + tableau des références d'apprentissage
            st.markdown(
                f'<div class="info-box" style="background:{bg_color};'
                f'border-left-color:{border_color};color:{text_color};">'
                f'🤖 <b>Best Price estimé par le modèle IA :</b> '
                f'<b style="color:{text_color};font-size:22px;">{est_price:.2f} €</b>'
                f'</div>',
                unsafe_allow_html=True
            )
            try:
                train_pool = get_ai_training_pool(designation or "", "palette")
                if not train_pool.empty:
                    train_show = train_pool[[c for c in [
                        COL_REF, COL_DESIGNATION, COL_SUPPLIER, COL_PRICE,
                        COL_LENGTH, COL_WIDTH, COL_HEIGHT, COL_WEIGHT, COL_THICKNESS
                    ] if c in train_pool.columns]].copy()
                    if COL_PRICE in train_show.columns:
                        train_show = train_show.sort_values(COL_PRICE).reset_index(drop=True)
                    rename_map = {COL_REF: "Référence", COL_DESIGNATION: "Composant",
                                  COL_SUPPLIER: "Fournisseur", COL_PRICE: "Prix (€)",
                                  COL_LENGTH: "L (mm)", COL_WIDTH: "l (mm)",
                                  COL_HEIGHT: "H (mm)", COL_WEIGHT: "Poids (g)",
                                  COL_THICKNESS: "Épaisseur (mm)"}
                    train_show = train_show.rename(columns={k: v for k, v in rename_map.items()
                                                             if k in train_show.columns})
                    n_train = len(train_show)
                    st.markdown(
                        f"<h4 style='color:#0f172a;font-weight:800;margin:14px 0 6px 0;font-size:15px;'>"
                        f"📚 Références qui ont aidé à calculer cette estimation "
                        f"<span style='color:#64748b;font-weight:500;font-size:12px;'>"
                        f"— {n_train} composant(s) du même type</span></h4>",
                        unsafe_allow_html=True
                    )
                    st.dataframe(train_show, use_container_width=True, hide_index=True,
                                 height=min(380, 38 + 35 * min(n_train, 10)),
                                 column_config={
                                     "Prix (€)": st.column_config.NumberColumn(format="%.2f"),
                                     "L (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "l (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "H (mm)":   st.column_config.NumberColumn(format="%.0f"),
                                     "Poids (g)": st.column_config.NumberColumn(format="%.0f"),
                                     "Épaisseur (mm)": st.column_config.NumberColumn(format="%g"),
                                 })
            except Exception:
                pass
        else:
            st.markdown('<div class="info-box" style="background:#f1f5f9;border-left-color:#94a3b8;">'
                        '⚪ <b>Pas de comparatif disponible</b> — données insuffisantes pour entraîner le modèle IA.</div>',
                        unsafe_allow_html=True)
        return

    # 6) Variables prix (gardées pour annotations chart, bloc Meilleur prix retiré)
    best_price = float(df_exact[COL_PRICE].min())
    median_price = float(df_exact[COL_PRICE].median())  # Stat de cohérence (médiane = robuste aux outliers)

    # 6) Filtre fournisseur AVANT les stats
    sup_opts = sorted(df_exact[COL_SUPPLIER].dropna().unique().tolist()) \
               if COL_SUPPLIER in df_exact.columns else []
    if sup_opts:
        sup_sel = st.selectbox("Filtrer par Fournisseur",
                               ["Tous les Fournisseurs"] + sup_opts,
                               index=0, key="palette_sup_filter")
    else:
        sup_sel = "Tous les Fournisseurs"

    df_show = df_exact.copy()
    if sup_sel and sup_sel != "Tous les Fournisseurs" and COL_SUPPLIER in df_show.columns:
        df_show = df_show[df_show[COL_SUPPLIER] == sup_sel]

    # 7) Stats sur le pool filtré
    st.markdown("<br>", unsafe_allow_html=True)
    _lbl = "Mêmes type + dimensions + poids" + (f" · {sup_sel}" if sup_sel != "Tous les Fournisseurs" else "")
    stat_cards_from_df(df_show, COL_PRICE, _lbl)

    # 8) Courbes — Sélection intelligente (1 graphique si dims+poids identiques, sinon 2 onglets)
    if not df_show.empty:
        df_chart = df_show.copy()
        if "Volume_m3" not in df_chart.columns and all(c in df_chart.columns for c in [COL_LENGTH, COL_WIDTH, COL_HEIGHT]):
            df_chart["Volume_m3"] = (
                df_chart[COL_LENGTH].fillna(0) *
                df_chart[COL_WIDTH].fillna(0) *
                df_chart[COL_HEIGHT].fillna(0)
            ) / 1e9
        if "weight_kg" not in df_chart.columns and COL_WEIGHT in df_chart.columns:
            df_chart["weight_kg"] = df_chart[COL_WEIGHT] / 1000.0

        # Toujours un graphique simple par fournisseur
        st.markdown("<h3 style='color:#0f172a;font-weight:800;margin-bottom:12px;'>"
                    "Graphique d'Analyse</h3>", unsafe_allow_html=True)
        fig = build_scatter_by_reference(
            df_chart, COL_PRICE, "Prix (€)",
            "Prix par fournisseur (mêmes dimensions, même poids)",
            ref_price=price, best_price=best_price,
            target_supplier=sup_r,
        )
        st.plotly_chart(fig, use_container_width=True)

    # 9) Tableau
    st.markdown("<br>", unsafe_allow_html=True)
    show_table(df_show.sort_values(COL_PRICE) if COL_PRICE in df_show.columns else df_show,
               f"Palettes Similaires — {sup_sel}")


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSE — MOD  (uniquement le fournisseur de la référence)
# ═════════════════════════════════════════════════════════════════════════════

def _render_mod_ui(df_mod_in):
    """Analyse libre MOD (depuis la page Analyse libre)."""
    cl,cr=st.columns([2.2,1.4],gap="large")
    with cr:
        st.markdown('<div class="config-section"><h3>⚙️ Configuration</h3>',unsafe_allow_html=True)
        liste_sup=sorted(df_mod_in[COL_SUPPLIER].unique().tolist()) if not df_mod_in.empty else []
        sup=st.selectbox("Fournisseur / Site",["Tous"]+liste_sup, key="mod_sup_libre")
        def_val=float(df_mod_in[COL_LABOUR].dropna().median()) if (
            not df_mod_in.empty and COL_LABOUR in df_mod_in.columns and
            df_mod_in[COL_LABOUR].notna().any()) else 25.
        taux=st.number_input("Taux horaire de référence (€/h)",value=def_val,step=1.,min_value=0.,
                             key="mod_taux_libre")
        if st.button("🚀 Lancer l'Analyse",use_container_width=True,type="primary", key="mod_go_libre"):
            st.session_state.run_analysis=True
        st.markdown('</div>', unsafe_allow_html=True)

    with cl:
        if not st.session_state.run_analysis:
            st.markdown('<div class="info-box">Sélectionnez un fournisseur (ou tous) puis '
                        "cliquez sur <b>Lancer l'Analyse</b>.</div>", unsafe_allow_html=True)
            return
        if sup == "Tous":
            _run_mod_general(df_mod_in, taux)
        else:
            _run_mod_for_supplier(df_mod_in, sup, taux)


def _run_mod_for_supplier(df_mod_in: pd.DataFrame, supplier_name: str,
                          taux_ref: float, display_cost: float = None,
                          mfg_site: str = "", dispatch: str = ""):
    """
    Taux MOD du fournisseur ciblé AU MÊME SITE DE PRODUCTION (manufacturing_site).
    Le taux MOD dépend du site de production, PAS du site d'enlèvement — donc on
    ne filtre QUE sur le site de production. (`dispatch` conservé pour compat.)
    """
    if df_mod_in is None or df_mod_in.empty:
        st.markdown('<div class="warning-box">Aucune donnée MOD disponible.</div>',
                    unsafe_allow_html=True)
        return

    if not supplier_name or supplier_name in ("Inconnu", ""):
        st.markdown('<div class="warning-box">Fournisseur inconnu pour cette référence.</div>',
                    unsafe_allow_html=True)
        return

    # Filtrage STRICT : fournisseur + site de production
    df_sup = df_mod_in[df_mod_in[COL_SUPPLIER] == supplier_name].copy()
    if df_sup.empty:
        st.markdown(f'<div class="warning-box">⚠ Aucun taux MOD trouvé pour <b>{supplier_name}</b>.</div>',
                    unsafe_allow_html=True)
        return

    # Filtre site de production — STRICT (pas de fallback)
    sites_info = []
    if mfg_site and mfg_site.strip() and mfg_site.lower() not in ("nan", "none", "") \
            and COL_MFG_SITE in df_sup.columns:
        df_sup_m = df_sup[df_sup[COL_MFG_SITE].astype(str).str.strip() == mfg_site.strip()]
        if df_sup_m.empty:
            st.markdown(
                f'<div class="warning-box">⚠ Aucun taux MOD trouvé pour <b>{supplier_name}</b> '
                f'au site de production <b>{mfg_site}</b>.</div>',
                unsafe_allow_html=True
            )
            return
        df_sup = df_sup_m
        sites_info.append(f"Site de production : <b>{mfg_site}</b>")

    # Dédoublonnage : on retire les lignes parfaitement identiques
    dedup_cols = [c for c in [COL_REF, COL_SUPPLIER, COL_MFG_SITE, COL_DISPATCH, COL_LABOUR]
                  if c in df_sup.columns]
    if dedup_cols:
        df_sup = df_sup.drop_duplicates(subset=dedup_cols, keep="first").copy()

    rates = df_sup[COL_LABOUR].dropna()
    rates = rates[rates > 0]
    if rates.empty:
        st.markdown(f'<div class="warning-box">⚠ Pas de taux horaire valide pour {supplier_name}.</div>',
                    unsafe_allow_html=True)
        return

    rates_unique = sorted(rates.unique())
    main_rate = float(taux_ref) if taux_ref > 0 else float(np.median(rates_unique))

    # KPI — HELPER UNIFIÉ (avec sites pour cohérence)
    status, _ = eval_status_mod(main_rate, supplier_name,
                                mfg_site=mfg_site, dispatch=dispatch)
    kpi_block(main_rate,
              f"Taux MOD — {supplier_name}",
              True,
              kpi_class=kpi_class_for_status(status),
              status=status if len(rates_unique) > 1 else "neutre")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Graphique : BARRES HORIZONTALES (une barre par ligne, triées par taux)
    df_chart = df_sup[[COL_REF, COL_LABOUR]].dropna().copy()
    df_chart = df_chart[df_chart[COL_LABOUR] > 0].sort_values(COL_LABOUR)
    if not df_chart.empty:
        # Scatter simple : un point par référence, couleur selon écart
        colors = []
        for v in df_chart[COL_LABOUR]:
            dev = (v - main_rate) / main_rate * 100 if main_rate > 0 else 0
            if dev <= -10:   colors.append("#10b981")
            elif dev >= 10:  colors.append("#ef4444")
            else:            colors.append("#3b82f6")

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_chart[COL_LABOUR],
            y=df_chart[COL_REF].astype(str),
            mode="markers",
            marker=dict(size=12, color=colors,
                        line=dict(color="white", width=1.5)),
            hovertemplate="<b>%{y}</b><br>Taux : %{x:.2f} €/h<extra></extra>",
            showlegend=False,
        ))
        fig.add_vline(x=main_rate, line_dash="dash", line_color="#64748b", line_width=1.5)
        # Resserrer l'axe X autour des données pour éliminer l'espace vide
        _xmin = float(df_chart[COL_LABOUR].min())
        _xmax = float(df_chart[COL_LABOUR].max())
        _xmin = min(_xmin, main_rate)
        _xmax = max(_xmax, main_rate)
        _pad = max((_xmax - _xmin) * 0.12, _xmax * 0.02, 0.5)
        fig.update_layout(
            xaxis_title="Taux horaire (€/h)",
            yaxis_title="",
            height=max(220, 40 + 26 * len(df_chart)),
            template="plotly_white",
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(family="DM Sans", size=12),
            margin=dict(l=100, r=40, t=10, b=45),
            yaxis=dict(automargin=True, tickfont=dict(size=11), showgrid=False),
            xaxis=dict(gridcolor="#f1f5f9", zeroline=False,
                       range=[_xmin - _pad, _xmax + _pad]),
        )
        st.plotly_chart(fig, use_container_width=True)
    st.markdown("<br>", unsafe_allow_html=True)

    stat_cards_from_df(df_sup, COL_LABOUR, f"Taux MOD — {supplier_name}")

    # Tableau : on affiche le SITE DE PRODUCTION (manufacturing_site) uniquement.
    # Le site d'enlèvement (dispatch_site) n'est PAS affiché pour la MOD.
    cols_mod = [COL_SUPPLIER, COL_REF]
    if COL_MFG_SITE in df_sup.columns:
        cols_mod.append(COL_MFG_SITE)
    cols_mod.append(COL_LABOUR)
    if COL_ECON_DATE in df_sup.columns:
        cols_mod.append(COL_ECON_DATE)
    df_table = (df_sup[cols_mod].dropna(subset=[COL_LABOUR]).copy())
    df_table = df_table[df_table[COL_LABOUR] > 0].sort_values(COL_LABOUR)
    if COL_ECON_DATE in df_table.columns:
        df_table[COL_ECON_DATE] = parse_econ_date(df_table[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")

    # Renommage pour affichage
    rename_map = {
        COL_SUPPLIER: "Fournisseur",
        COL_REF: "Référence",
        COL_MFG_SITE: "Site de production",
        COL_LABOUR: "Taux horaire (€/h)",
        COL_ECON_DATE: "Date validation",
    }
    df_table_display = df_table.rename(columns=rename_map)
    if "Taux horaire (€/h)" in df_table_display.columns:
        df_table_display["Taux horaire (€/h)"] = pd.to_numeric(
            df_table_display["Taux horaire (€/h)"], errors="coerce"
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f'<div class="table-section"><h3>Détail Taux MOD — {supplier_name}</h3>',
                unsafe_allow_html=True)
    col_config_mod = {
        "Taux horaire (€/h)": st.column_config.NumberColumn(format="%.2f"),
    }
    if not df_table_display.empty:
        st.dataframe(df_table_display, use_container_width=True, hide_index=True,
                     column_config=col_config_mod)
    else:
        st.markdown('<div class="info-box">Aucune donnée à afficher.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


def _run_mod_general(df_mod_in: pd.DataFrame, taux_ref: float = None):
    """Analyse MOD libre tous fournisseurs (depuis la page Analyse libre)."""
    if df_mod_in is None or df_mod_in.empty:
        st.markdown('<div class="warning-box">Aucune donnée MOD disponible.</div>',
                    unsafe_allow_html=True)
        return

    if taux_ref is None or taux_ref <= 0:
        taux_ref = float(df_mod_in[COL_LABOUR].dropna().median())

    status, _ = evaluate_price(taux_ref, df_mod_in, COL_LABOUR)
    kpi_block(taux_ref, "Taux MOD de Référence (€/h)", True,
              kpi_class=kpi_class_for_status(status), status=status)

    st.markdown("<br>", unsafe_allow_html=True)

    valid = df_mod_in.dropna(subset=[COL_LABOUR]).copy()
    valid = valid[valid[COL_LABOUR] > 0]
    if not valid.empty and COL_SUPPLIER in valid.columns:
        # Scatter simple : un point par fournisseur (médiane), couleur selon écart
        med_by_sup = (valid.groupby(COL_SUPPLIER)[COL_LABOUR]
                      .median()
                      .sort_values()
                      .reset_index())
        colors = []
        for v in med_by_sup[COL_LABOUR]:
            dev = (v - taux_ref) / taux_ref * 100 if taux_ref > 0 else 0
            if dev <= -10:   colors.append("#10b981")
            elif dev >= 10:  colors.append("#ef4444")
            else:            colors.append("#3b82f6")

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=med_by_sup[COL_LABOUR],
            y=med_by_sup[COL_SUPPLIER].astype(str),
            mode="markers",
            marker=dict(size=12, color=colors,
                        line=dict(color="white", width=1.5)),
            hovertemplate="<b>%{y}</b><br>Taux médian : %{x:.2f} €/h<extra></extra>",
            showlegend=False,
        ))
        fig.add_vline(x=taux_ref, line_dash="dash", line_color="#64748b", line_width=1.5)
        fig.update_layout(
            xaxis_title="Taux horaire (€/h)",
            yaxis_title="",
            height=max(280, 50 + 22 * len(med_by_sup)),
            template="plotly_white",
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(family="DM Sans", size=12),
            margin=dict(l=160, r=40, t=20, b=50),
            yaxis=dict(automargin=True, tickfont=dict(size=11), showgrid=False),
            xaxis=dict(gridcolor="#f1f5f9", zeroline=False),
        )
        st.plotly_chart(fig, use_container_width=True)
    st.markdown("<br>", unsafe_allow_html=True)

    stat_cards_from_df(df_mod_in, COL_LABOUR, "Taux MOD — tous sites")

    if not df_mod_in.empty and COL_LABOUR in df_mod_in.columns:
        dfs_mod = (df_mod_in[[COL_SUPPLIER, COL_LABOUR]].dropna().copy())
        dfs_mod = dfs_mod[dfs_mod[COL_LABOUR] > 0]
        if COL_SUPPLIER in dfs_mod.columns:
            dfs_mod = (dfs_mod.groupby(COL_SUPPLIER, as_index=False)[COL_LABOUR]
                       .median()
                       .sort_values(COL_LABOUR))
        show_table(dfs_mod, "Taux Horaires MOD par Fournisseur (€/h)",
                   as_df=True, table_type="mod")


# ═════════════════════════════════════════════════════════════════════════════
# ANALYSE — TRANSPORT
# ═════════════════════════════════════════════════════════════════════════════

def _render_transport_ui(df_trans):
    cl,cr=st.columns([2.2,1.4],gap="large")
    with cr:
        st.markdown('<div class="config-section"><h3>⚙️ Configuration</h3>',unsafe_allow_html=True)
        liste_sup=sorted(df_trans[COL_SUPPLIER].unique().tolist()) if not df_trans.empty else []
        sup=st.selectbox("Transporteur",["Tous"]+liste_sup, key="trans_sup")
        def_val=float(df_trans[COL_SHIPPING].dropna().median()) if (
            not df_trans.empty and COL_SHIPPING in df_trans.columns and
            df_trans[COL_SHIPPING].notna().any()) else 50.
        frais=st.number_input("Frais de transport de référence (€)",value=def_val,step=5.,min_value=0.,
                              key="trans_frais")
        if st.button("🚀 Lancer l'Analyse",use_container_width=True,type="primary", key="trans_go"):
            st.session_state.run_analysis=True
        st.markdown('</div>', unsafe_allow_html=True)

    with cl:
        if not st.session_state.run_analysis:
            st.markdown('<div class="info-box">Sélectionnez un transporteur puis '
                        "cliquez sur <b>Lancer l'Analyse</b>.</div>", unsafe_allow_html=True)
            return
        df_view = df_trans.copy()
        if sup != "Tous" and COL_SUPPLIER in df_view.columns:
            df_view = df_view[df_view[COL_SUPPLIER] == sup]
        _run_transport_analysis(df_view, frais)


def _run_transport_analysis(df_pool: pd.DataFrame, frais_ref: float,
                            forced_supplier: str = None,
                            forced_mfg_site: str = None,
                            forced_dispatch: str = None,
                            forced_parts_qty: float = 0.0,
                            forced_client_site: str = None):
    """Analyse Transport — filtre par fournisseur + manufacturing_site + dispatch_site.
    Affichage : barres horizontales (un point par ligne) + table avec sites.
    """
    if df_pool is None or df_pool.empty or COL_SHIPPING not in df_pool.columns:
        st.markdown('<div class="warning-box">Aucune donnée transport disponible.</div>',
                    unsafe_allow_html=True)
        return

    pool = df_pool[df_pool[COL_SHIPPING].notna() & (df_pool[COL_SHIPPING] > 0)].copy()
    if pool.empty:
        st.markdown('<div class="warning-box">Aucun frais de transport valide.</div>',
                    unsafe_allow_html=True)
        return

    # ── Filtrage STRICT : même fournisseur + même site d'enlèvement
    # + même site client + MÊME partsqty packaging.
    if forced_supplier and forced_supplier not in ("", "Inconnu") \
            and COL_SUPPLIER in pool.columns:
        pool_f = pool[pool[COL_SUPPLIER] == forced_supplier]
        if pool_f.empty:
            st.markdown(
                f'<div class="warning-box">⚠ Aucun transport pour <b>{forced_supplier}</b>.</div>',
                unsafe_allow_html=True
            )
            return
        pool = pool_f

    # Site d'enlèvement (dispatch_site)
    if forced_dispatch and forced_dispatch.lower() not in ("", "nan", "none") \
            and COL_DISPATCH in pool.columns:
        pool_f = pool[pool[COL_DISPATCH].astype(str).str.strip() == str(forced_dispatch).strip()]
        if pool_f.empty:
            st.markdown(
                f'<div class="warning-box">⚠ Aucun transport au site d\'enlèvement '
                f'<b>{forced_dispatch}</b>.</div>',
                unsafe_allow_html=True
            )
            return
        pool = pool_f

    # Site client (destination)
    if forced_client_site and str(forced_client_site).lower() not in ("", "nan", "none") \
            and COL_CLIENT_SITE in pool.columns:
        pool_f = pool[pool[COL_CLIENT_SITE].astype(str).str.strip() == str(forced_client_site).strip()]
        if pool_f.empty:
            st.markdown(
                f'<div class="warning-box">⚠ Aucun transport vers le site client '
                f'<b>{forced_client_site}</b>.</div>',
                unsafe_allow_html=True
            )
            return
        pool = pool_f

    # ── Filtre STRICT par partsqty packaging
    if forced_parts_qty and forced_parts_qty > 0 and COL_PARTS_QTY in pool.columns:
        tol_pq = max(forced_parts_qty * 0.001, 0.5)
        pool_f = pool[pool[COL_PARTS_QTY].between(
            forced_parts_qty - tol_pq, forced_parts_qty + tol_pq
        )]
        if pool_f.empty:
            st.markdown(
                f'<div class="warning-box">⚠ Aucun transport avec partsqty = '
                f'<b>{forced_parts_qty:g}</b>.</div>',
                unsafe_allow_html=True
            )
            return
        pool = pool_f

    # ── Dédoublonnage
    dedup_cols = [c for c in [COL_REF, COL_SUPPLIER, COL_SHIPPING, COL_DISPATCH, COL_CLIENT_SITE]
                  if c in pool.columns]
    if dedup_cols:
        pool = pool.drop_duplicates(subset=dedup_cols, keep="first").copy()

    # KPI — HELPER UNIFIÉ (cohérent avec accueil + ref_detail)
    status, _ = eval_status_transport(frais_ref, forced_supplier or "",
                                       pickup_site=forced_dispatch or "",
                                       client_site=forced_client_site or "",
                                       parts_qty=forced_parts_qty or 0)
    kpi_block(frais_ref, "Frais de Transport (€)", True,
              kpi_class=kpi_class_for_status(status), status=status)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Graphique : SCATTER simple (un point par référence)
    df_chart = pool[[COL_REF, COL_SHIPPING]].dropna().copy()
    df_chart = df_chart[df_chart[COL_SHIPPING] > 0].sort_values(COL_SHIPPING)
    if not df_chart.empty:
        colors = []
        for v in df_chart[COL_SHIPPING]:
            dev = (v - frais_ref) / frais_ref * 100 if frais_ref > 0 else 0
            if dev <= -10:   colors.append("#10b981")
            elif dev >= 10:  colors.append("#ef4444")
            else:            colors.append("#3b82f6")

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_chart[COL_SHIPPING],
            y=df_chart[COL_REF].astype(str),
            mode="markers",
            marker=dict(size=12, color=colors,
                        line=dict(color="white", width=1.5)),
            hovertemplate="<b>%{y}</b><br>Frais : %{x:.2f} €<extra></extra>",
            showlegend=False,
        ))
        fig.add_vline(x=frais_ref, line_dash="dash", line_color="#64748b", line_width=1.5)
        fig.update_layout(
            xaxis_title="Frais (€)",
            yaxis_title="",
            height=max(280, 50 + 22 * len(df_chart)),
            template="plotly_white",
            plot_bgcolor="white",
            paper_bgcolor="white",
            font=dict(family="DM Sans", size=12),
            margin=dict(l=100, r=40, t=20, b=50),
            yaxis=dict(automargin=True, tickfont=dict(size=11), showgrid=False),
            xaxis=dict(gridcolor="#f1f5f9", zeroline=False),
        )
        st.plotly_chart(fig, use_container_width=True)
    st.markdown("<br>", unsafe_allow_html=True)

    stat_cards_from_df(pool, COL_SHIPPING, "Transport — sélection")

    # Table avec colonnes étendues : Réf, Fournisseur, Sites, Frais, Date validation
    table_cols = [c for c in [COL_REF, COL_SUPPLIER, COL_MFG_SITE, COL_DISPATCH,
                              COL_SHIPPING, COL_ECON_DATE]
                  if c in pool.columns]
    dfs = (pool[table_cols]
           .dropna(subset=[COL_SHIPPING] if COL_SHIPPING in table_cols else table_cols[:1])
           .copy()
           .sort_values(COL_SHIPPING))
    if COL_ECON_DATE in dfs.columns:
        dfs[COL_ECON_DATE] = parse_econ_date(dfs[COL_ECON_DATE]).dt.strftime("%d/%m/%Y")
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="table-section"><h3>Détail Transport — frais d\'expédition</h3>',
                unsafe_allow_html=True)
    if not dfs.empty:
        # Renommage cosmétique des colonnes pour l'affichage
        rename_map = {
            COL_REF:        "Référence",
            COL_SUPPLIER:   "Fournisseur",
            COL_MFG_SITE:   "Site Manufacturing",
            COL_DISPATCH:   "Site Dispatch",
            COL_SHIPPING:   "Frais (€)",
            COL_ECON_DATE:  "Date validation",
        }
        dfs_display = dfs.rename(columns=rename_map)
        # Forçage typage numérique pour que le tri par clic fonctionne
        if "Frais (€)" in dfs_display.columns:
            dfs_display["Frais (€)"] = pd.to_numeric(dfs_display["Frais (€)"], errors="coerce")
        col_config = {
            "Frais (€)": st.column_config.NumberColumn(format="%.2f"),
        }
        st.dataframe(dfs_display, use_container_width=True, hide_index=True,
                     column_config=col_config)
    else:
        st.markdown('<div class="info-box">Aucune donnée à afficher.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# NOUVELLE PAGE : OPTIMISATION MULTI-RÉFÉRENCES
# L'utilisateur saisit une liste de références (une par ligne ou séparées par
# virgule/point-virgule). L'app calcule pour chacune :
#   • Le coût total
#   • Les composants en écart élevé
#   • L'économie potentielle (somme des (prix - best_price) pour chaque écart élevé)
#   • Le coût optimisé
# Puis affiche la SOMME totale des économies sur toutes les références saisies.
# ═════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def _compute_ref_savings(ref_code: str) -> dict:
    """
    SOURCE UNIQUE DE VÉRITÉ pour l'économie potentielle d'une référence.

    FORMULES :
      Composant packaging en écart ÉLEVÉ :
        économie = (unitprice_packaging - prix_médiane) / partsqty_packaging
      MOD en écart ÉLEVÉ :
        économie = (taux_actuel - taux_médian) × temps_par_pièce
        temps_par_pièce (h) = total_labour_seconds / 3600 / partsqty(packaginglevel=4)
      Transport en écart ÉLEVÉ :
        économie = frais_actuel - frais_médian

      Coût optimisé = Coût total - Σ économies

    GARDE-FOUS :
      • Si total_cost absent → on le calcule = Σ des coûts/pièce des composants
      • L'économie d'un composant est plafonnée à sa contribution au coût total
        (on ne peut pas économiser plus que ce que coûte le composant)
      • Σ économies plafonnée au coût total → coût optimisé jamais négatif
    """
    result = {
        "ref": ref_code, "cout_total": None, "cout_total_estime": False,
        "supplier_label": "\u2014",
        "n_composants": 0, "n_high_ecart": 0,
        "total_savings": 0.0, "cout_optimise": None,
        "details": [], "status": "ok",
    }
    if not ref_code or not str(ref_code).strip():
        result["status"] = "empty"
        return result

    ref_code = str(ref_code).strip()

    # ── Toutes les lignes de la référence (NON filtrées par désignation)
    # → indispensable pour retrouver MOD / transport / packaginglevel=4
    if COL_REF not in df_raw.columns:
        result["status"] = "not_found"
        return result
    ref_norm = ref_code.upper()
    all_rows = df_raw[df_raw[COL_REF].astype(str).str.strip().str.upper() == ref_norm].copy()
    if all_rows.empty:
        result["status"] = "not_found"
        return result

    # ── Composants packaging (avec désignation valide)
    df_ref = get_ref_components(df_raw, ref_code)
    if COL_DESIGNATION in df_ref.columns:
        df_ref = df_ref.drop_duplicates(subset=[COL_DESIGNATION], keep="first").reset_index(drop=True)

    # ── Fournisseur dominant
    if COL_SUPPLIER in all_rows.columns:
        sups = {str(s).strip() for s in all_rows[COL_SUPPLIER].dropna().unique()
                if str(s).strip() and str(s).strip() not in ("Inconnu", "nan", "None")}
        if len(sups) == 1:
            result["supplier_label"] = next(iter(sups))
        elif len(sups) > 1:
            result["supplier_label"] = f"{len(sups)} fournisseurs"

    # ── Coût total : depuis la colonne, sinon calculé = Σ des coûts/pièce
    total_cost = None
    if COL_TOTAL_COST in all_rows.columns:
        vals = all_rows[COL_TOTAL_COST].dropna()
        vals = vals[vals > 0]
        if not vals.empty:
            total_cost = float(vals.iloc[0])
    if total_cost is None:
        # Calcul de remplacement : somme des coûts/pièce (_unit_price) des composants
        if "_unit_price" in df_ref.columns:
            up = df_ref["_unit_price"].dropna()
            up = up[up > 0]
            if not up.empty:
                total_cost = float(up.sum())
                result["cout_total_estime"] = True
    result["cout_total"] = total_cost

    # ── Partsqty au niveau 4 (packaginglevel = 4) pour le calcul du temps MOD
    # On récupère AUSSI total_labour_seconds depuis cette même logique :
    # le temps de main d'œuvre est associé au packaging final (niveau 4).
    parts_qty_level_4 = None
    total_seconds_level_4 = None
    if COL_PACKAGING_LEVEL in all_rows.columns and COL_PARTS_QTY in all_rows.columns:
        try:
            lvl = pd.to_numeric(all_rows[COL_PACKAGING_LEVEL], errors="coerce")
            m4 = (lvl == 4)
            if m4.any():
                rows4 = all_rows.loc[m4]
                pq4 = rows4[COL_PARTS_QTY].dropna()
                pq4 = pq4[pq4 > 0]
                if not pq4.empty:
                    parts_qty_level_4 = float(pq4.iloc[0])
                # total_labour_seconds sur la ligne niveau 4
                if COL_LABOUR_SECONDS in rows4.columns:
                    sec4 = rows4[COL_LABOUR_SECONDS].dropna()
                    sec4 = sec4[sec4 > 0]
                    if not sec4.empty:
                        total_seconds_level_4 = float(sec4.iloc[0])
        except Exception:
            parts_qty_level_4 = None

    total_savings = 0.0
    n_high = 0
    n_components = 0

    # ═══ COMPOSANTS PACKAGING ═══
    for _, row in df_ref.iterrows():
        try:
            cat_raw = row.get("_category", "autre")
            desig = str(row.get(COL_DESIGNATION, "")).strip()
            if not desig:
                continue
            cat_eval = "autre" if (cat_raw == "etiquette" or
                                  (cat_raw == "autre" and is_label_designation(desig))) \
                       else cat_raw
            if cat_eval not in ("carton", "sac", "palette", "cale", "autre"):
                continue

            pv_raw = _safe_float(row.get(COL_PRICE))
            if pv_raw <= 0:
                continue
            n_components += 1

            Lv = _safe_float(row.get(COL_LENGTH, 0))
            Wv = _safe_float(row.get(COL_WIDTH, 0))
            Hv = _safe_float(row.get(COL_HEIGHT, 0))
            Pv = _safe_float(row.get(COL_WEIGHT, 0))

            cfab = str(row.get(COL_COUNTRY_FAB, "")) if COL_COUNTRY_FAB in row.index else ""
            cdis = str(row.get(COL_COUNTRY_DISPATCH, "")) if COL_COUNTRY_DISPATCH in row.index else ""
            status_c, _ = eval_status_dim(pv_raw, desig, Lv, Wv, Hv, Pv, cat_eval,
                                          country_fab=cfab, country_dispatch=cdis)
            # On compte TOUT écart (moyen OU élevé) — l'utilisateur décide ensuite
            if status_c not in ("élevé", "moyen"):
                continue

            # Pool strict → médiane (+ même pays fabrication/expédition)
            pool_full = CAT_DFS.get(cat_eval, df_carton)
            pool_best = filter_exact_dims(pool_full, desig, Lv, Wv, Hv, Pv, cat_eval,
                                          country_fab=cfab, country_dispatch=cdis)
            pool_best = _dedup_pool(pool_best, COL_PRICE)
            if pool_best.empty or COL_PRICE not in pool_best.columns:
                continue
            vp = pool_best[COL_PRICE].dropna()
            vp = vp[vp > 0]
            if vp.empty:
                continue
            median_p = float(vp.median())
            if median_p <= 0 or pv_raw <= median_p:
                continue

            parts_q = _safe_float(row.get(COL_PARTS_QTY, 0))
            if parts_q <= 0:
                continue

            # FORMULE : économie = (prix - médiane) / partsqty
            economie = (pv_raw - median_p) / parts_q
            if economie <= 0:
                continue

            # GARDE-FOU : l'économie ne peut pas dépasser la contribution
            # du composant au coût total (= son _unit_price)
            contribution = _safe_float(row.get("_unit_price", 0))
            if contribution <= 0:
                pack_q = _safe_float(row.get(COL_PACKAGING_QTY, 1))
                if pack_q <= 0:
                    pack_q = 1
                contribution = (pv_raw * pack_q) / parts_q
            if contribution > 0 and economie > contribution:
                economie = contribution  # plafonnement

            # GARDE-FOU 2 : économie aberrante vs coût total → on ignore le composant
            # (cas typique : prix dans une devise non convertie)
            if total_cost is not None and total_cost > 0 and economie > total_cost:
                continue

            total_savings += economie
            n_high += 1
            result["details"].append({
                "composant": desig, "categorie": cat_eval,
                "prix": pv_raw, "mediane": median_p,
                "partsqty": parts_q, "economie": economie,
            })
        except Exception:
            continue

    # ═══ MOD ═══  (lignes lues depuis all_rows, PAS df_ref)
    # Formule : économie = (taux_actuel − taux_médian) × temps_par_pièce
    #   temps_par_pièce (h) = total_labour_seconds / 3600 / partsqty(packaginglevel = 4)
    if COL_LABOUR in all_rows.columns:
        try:
            mod_rows = all_rows[all_rows[COL_LABOUR].notna() & (all_rows[COL_LABOUR] > 0)]
            if not mod_rows.empty:
                mod_row = mod_rows.iloc[0]
                taux = _safe_float(mod_row.get(COL_LABOUR))
                supplier_mod = str(mod_row.get(COL_SUPPLIER, ""))
                mfg = str(mod_row.get(COL_MFG_SITE, "")) if COL_MFG_SITE in mod_row.index else ""

                # Statut + médiane via les helpers UNIFIÉS (même pool partout :
                # même fournisseur + même site de production, PAS le dispatch).
                status_mod, _ = eval_status_mod(taux, supplier_mod, mfg_site=mfg)
                median_rate = mod_median_rate(supplier_mod, mfg)

                # Best price MOD = taux médian du marché (pool en cascade).
                # Dès que le taux actuel DÉPASSE le best price, il y a une économie
                # potentielle — peu importe l'étiquette formelle du statut.
                if median_rate > 0 and taux > median_rate:
                    # temps_total (secondes) : ligne niveau 4 en priorité,
                    # sinon n'importe quelle ligne de la réf qui porte la valeur,
                    # sinon la ligne MOD elle-même.
                    total_seconds = total_seconds_level_4
                    if not total_seconds or total_seconds <= 0:
                        if COL_LABOUR_SECONDS in all_rows.columns:
                            sec_any = all_rows[COL_LABOUR_SECONDS].dropna()
                            sec_any = sec_any[sec_any > 0]
                            if not sec_any.empty:
                                total_seconds = float(sec_any.iloc[0])
                    if not total_seconds or total_seconds <= 0:
                        total_seconds = _safe_float(mod_row.get(COL_LABOUR_SECONDS)) \
                            if COL_LABOUR_SECONDS in mod_row.index else 0

                    # partsqty au niveau 4 en priorité, sinon partsqty de la ligne MOD
                    pq_for_mod = parts_qty_level_4
                    if not pq_for_mod or pq_for_mod <= 0:
                        pq_for_mod = _safe_float(mod_row.get(COL_PARTS_QTY, 0))

                    if pq_for_mod and pq_for_mod > 0 and total_seconds > 0:
                        # temps_par_pièce en HEURES : total_labour_seconds / 3600 / partsqty
                        # (le taux MOD est en €/h, donc le temps doit être en heures)
                        time_per_piece = total_seconds / 3600.0 / pq_for_mod   # h/pièce
                        economie_mod = (taux - median_rate) * time_per_piece
                        # NB : on ne plafonne PAS l'économie MOD au coût total —
                        # la MOD est un poste distinct du coût packaging.
                        if economie_mod > 0:
                            total_savings += economie_mod
                            n_high += 1
                            result["details"].append({
                                "composant": "MOD (Main d'oeuvre directe)",
                                "categorie": "mod",
                                "prix": taux, "mediane": median_rate,
                                "partsqty": pq_for_mod,
                                "temps_total_sec": total_seconds,
                                "temps_par_piece": time_per_piece,
                                "economie": economie_mod,
                            })
        except Exception:
            pass

    # ═══ TRANSPORT ═══  (lignes lues depuis all_rows)
    if COL_SHIPPING in all_rows.columns:
        try:
            trp_rows = all_rows[all_rows[COL_SHIPPING].notna() & (all_rows[COL_SHIPPING] > 0)]
            if not trp_rows.empty:
                trp_row = trp_rows.iloc[0]
                frais_actuel = _safe_float(trp_row.get(COL_SHIPPING))
                supplier_trp = str(trp_row.get(COL_SUPPLIER, ""))
                disp_trp = str(trp_row.get(COL_DISPATCH, "")) if COL_DISPATCH in trp_row.index else ""
                client_trp = str(trp_row.get(COL_CLIENT_SITE, "")) if COL_CLIENT_SITE in trp_row.index else ""
                pq_trp = _safe_float(trp_row.get(COL_PARTS_QTY, 0))

                status_trp, _ = eval_status_transport(
                    frais_actuel, supplier_trp, pickup_site=disp_trp,
                    client_site=client_trp, parts_qty=pq_trp)
                if status_trp in ("élevé", "moyen"):
                    pool_trp = df_transport.copy() if not df_transport.empty else None
                    if pool_trp is not None and supplier_trp and COL_SUPPLIER in pool_trp.columns:
                        pool_trp = pool_trp[pool_trp[COL_SUPPLIER] == supplier_trp]
                    if pool_trp is not None and disp_trp.strip() and COL_DISPATCH in pool_trp.columns:
                        p2 = pool_trp[pool_trp[COL_DISPATCH].astype(str).str.strip() == disp_trp.strip()]
                        if not p2.empty: pool_trp = p2
                    if pool_trp is not None and client_trp.strip() and COL_CLIENT_SITE in pool_trp.columns:
                        p2 = pool_trp[pool_trp[COL_CLIENT_SITE].astype(str).str.strip() == client_trp.strip()]
                        if not p2.empty: pool_trp = p2
                    if pool_trp is not None and pq_trp > 0 and COL_PARTS_QTY in pool_trp.columns:
                        tol_pq = max(pq_trp * 0.001, 0.5)
                        p2 = pool_trp[pool_trp[COL_PARTS_QTY].between(pq_trp - tol_pq, pq_trp + tol_pq)]
                        if not p2.empty: pool_trp = p2
                    if pool_trp is not None:
                        pool_trp = _dedup_pool(pool_trp, COL_SHIPPING)
                        fp = pool_trp[COL_SHIPPING].dropna()
                        fp = fp[fp > 0]
                        if len(fp) >= 2:
                            median_frais = float(fp.median())
                            if frais_actuel > median_frais:
                                economie_trp = frais_actuel - median_frais
                                if total_cost is not None and total_cost > 0 \
                                        and economie_trp > total_cost:
                                    economie_trp = total_cost
                                if economie_trp > 0:
                                    total_savings += economie_trp
                                    n_high += 1
                                    result["details"].append({
                                        "composant": "Transport (frais d'expédition)",
                                        "categorie": "transport",
                                        "prix": frais_actuel, "mediane": median_frais,
                                        "partsqty": pq_trp, "economie": economie_trp,
                                    })
        except Exception:
            pass

    # ── GARDE-FOU FINAL : on ne plafonne PLUS Σ économies au coût total.
    # Les garde-fous par composant (économie ≤ contribution du composant) suffisent
    # à garantir un résultat cohérent. On évite ainsi le cas absurde
    # "économie = coût total exact et coût optimisé = 0".

    result["n_composants"] = n_components
    result["n_high_ecart"] = n_high
    result["total_savings"] = total_savings
    if total_cost is not None and total_savings > 0:
        result["cout_optimise"] = max(0.0, total_cost - total_savings)

    if n_components == 0:
        result["status"] = "no_components"
    return result



def render_batch_optimization():
    """
    Page d'optimisation multi-références — UTILISE LA MÊME LOGIQUE EXACTE que le
    calcul de coût optimisé de la page d'accueil (eval_status_dim + filter_exact_dims),
    appliquée à plusieurs références. Affiche directement le tableau, sans messages
    intermédiaires.
    """
    # ── En-tête
    st.markdown("""
        <div style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#0c4a6e 100%);
                    padding:24px 28px;border-radius:18px;margin-bottom:20px;
                    box-shadow:0 8px 28px rgba(15,23,42,0.25);position:relative;overflow:hidden;">
            <div style="position:absolute;top:0;left:0;right:0;height:4px;
                        background:linear-gradient(90deg,#38bdf8,#10b981,#8b5cf6);"></div>
            <h1 style="margin:0;color:white;font-size:26px;font-weight:900;letter-spacing:-0.5px;">
                💰 Optimisation Multi-Références
            </h1>
            <p style="margin:6px 0 0 0;color:#cbd5e1;font-size:13px;">
                Saisissez une liste de références — l'outil calcule les composants en écart élevé
                et l'économie potentielle totale.
            </p>
        </div>
    """, unsafe_allow_html=True)

    if st.button("← Retour à l'Accueil", key="bk_home_batch"):
        go_page("home"); st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Zone de saisie
    st.markdown(
        "<h3 style='color:#0f172a;font-weight:800;margin-bottom:8px;'>"
        "📋 Liste des références à analyser</h3>"
        "<p style='color:#64748b;font-size:13px;margin:0 0 12px 0;'>"
        "Une référence par ligne, ou séparées par des virgules / points-virgules.</p>",
        unsafe_allow_html=True
    )

    refs_input = st.text_area(
        "Références", height=140, key="batch_refs_input",
        placeholder="Exemple :\nREF-001\nREF-002\nREF-003",
        label_visibility="collapsed",
    )

    col_btn1, col_btn2, _ = st.columns([1, 1, 4])
    with col_btn1:
        analyze_clicked = st.button("🔎 Analyser", use_container_width=True, type="primary",
                                    key="batch_analyze_btn")
    with col_btn2:
        # Callback : modifie session_state AVANT que le widget text_area soit recréé,
        # ce qui évite l'erreur "session_state cannot be modified after widget instantiated"
        def _clear_batch_refs():
            st.session_state["batch_refs_input"] = ""
        st.button("🗑️ Effacer", use_container_width=True, key="batch_clear_btn",
                  on_click=_clear_batch_refs)

    refs_input_clean = (refs_input or "").strip()
    if not refs_input_clean:
        return  # Pas de message — juste pas d'affichage tant qu'aucune saisie

    # ── Parsing robuste : on supporte tous les séparateurs courants
    import re as _re
    raw_refs = _re.split(r"[\n,;|\t]+", refs_input_clean)
    refs_list = [r.strip().strip("'\"") for r in raw_refs if r.strip()]
    seen = set()
    refs_list = [r for r in refs_list if not (r in seen or seen.add(r))]
    if not refs_list:
        return

    # ── Calcul (spinner) — on appelle _compute_ref_savings pour chaque ref
    # Cette fonction utilise les formules CORRECTES :
    #   • Packaging : (prix - médiane) / partsqty
    #   • MOD : (taux - médiane) × (total_labour_seconds / partsqty_niveau_4)
    rows_per_ref = []
    total_savings_all = 0.0
    refs_not_found = []
    n_refs_with_savings = 0

    with st.spinner(f"⏳ Analyse de {len(refs_list)} référence(s) en cours…"):
        for ref_code in refs_list:
            try:
                res = _compute_ref_savings(ref_code)
            except Exception:
                refs_not_found.append(str(ref_code))
                continue
            if res.get("status") == "not_found":
                refs_not_found.append(res["ref"])
                continue
            ref_savings = res.get("total_savings", 0.0) or 0.0
            cout_opt = res.get("cout_optimise")
            total_cost_ref = res.get("cout_total")
            supplier_label = res.get("supplier_label", "—")
            # Si l'économie excède le total_cost (cas anormal du calcul),
            # on n'affiche NI le coût optimisé NI l'écart — pour ne pas montrer
            # des valeurs incohérentes (n'importe quoi) à l'utilisateur.
            calc_unreliable = (
                total_cost_ref is not None and ref_savings > 0
                and ref_savings >= total_cost_ref * 0.999
            )
            if calc_unreliable:
                cout_opt = None
                ref_savings_display = None
            else:
                ref_savings_display = ref_savings if ref_savings > 0 else None
            # On compte quand même comme "avec économie" si savings calculées > 0
            # (même si non affichées) pour le total agrégé
            if ref_savings > 0:
                n_refs_with_savings += 1
            # Mais le total_savings_all ne prend que les VALEURS FIABLES
            if not calc_unreliable:
                total_savings_all += ref_savings
            rows_per_ref.append({
                "Référence":         res["ref"],
                "Fournisseur":       supplier_label,
                "Total Cost (€)":    total_cost_ref,
                "Coût Optimisé (€)": cout_opt,
                "Écart (€)":         ref_savings_display,
                # Composants en écart : pour la colonne "Anomalie sur"
                "_details":          res.get("details", []) if not calc_unreliable else [],
            })

    n_refs_analyzed = len(rows_per_ref)

    # ── Bandeau récap : économie totale = somme des écarts
    if total_savings_all > 0:
        banner_html = (
            f'<div style="background:linear-gradient(135deg,#0f172a 0%,#065f46 100%);'
            f'padding:26px 28px;border-radius:18px;margin:18px 0;'
            f'box-shadow:0 12px 32px rgba(16,185,129,0.25);color:white;">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;gap:24px;flex-wrap:wrap;">'
            f'<div>'
            f'<div style="font-size:12px;color:#86efac;font-weight:700;letter-spacing:0.5px;text-transform:uppercase;">'
            f'💰 Économie potentielle totale</div>'
            f'<div style="font-size:46px;font-weight:900;color:#10b981;line-height:1;margin-top:6px;">'
            f'{total_savings_all:.2f} €</div>'
            f'</div>'
            f'<div style="text-align:right;font-size:13px;color:#cbd5e1;">'
            f'Sur <b style="color:white;font-size:18px;">{n_refs_analyzed}</b> référence(s) analysée(s)<br>'
            f'<b style="color:white;font-size:18px;">{n_refs_with_savings}</b> avec économie potentielle'
            f'</div></div></div>'
        )
        st.markdown(banner_html, unsafe_allow_html=True)

    # Avertissement refs non trouvées
    if refs_not_found:
        st.markdown(
            f'<div class="warning-box">⚠ {len(refs_not_found)} référence(s) introuvable(s) : '
            f'<code style="font-size:11px;">{", ".join(refs_not_found[:8])}'
            f'{("…" if len(refs_not_found) > 8 else "")}</code></div>',
            unsafe_allow_html=True
        )

    # ── TABLEAU : une ligne par référence
    if not rows_per_ref:
        st.markdown(
            '<div class="info-box">Aucune référence valide trouvée dans la base.</div>',
            unsafe_allow_html=True
        )
        return

    # ── Tri : refs avec économie en haut (Écart desc, None en bas)
    rows_sorted = sorted(
        rows_per_ref,
        key=lambda r: (-(r.get("Écart (€)") or -1e9), str(r.get("Référence", ""))),
    )

    # ─────────────────────────────────────────────────────────────────────
    # EXPORT EXCEL : génération en mémoire, bouton de téléchargement
    # ─────────────────────────────────────────────────────────────────────
    try:
        import io
        def _has_anom(r):
            return bool(r.get("_details")) and r.get("Écart (€)") and r.get("Écart (€)") > 0
        df_export = pd.DataFrame([
            {
                "Référence":         r.get("Référence", ""),
                "Fournisseur":       r.get("Fournisseur", "—"),
                "Total Cost (€)":    r.get("Total Cost (€)"),
                "Coût Optimisé (€)": r.get("Coût Optimisé (€)") if _has_anom(r) else "DCL OK",
                "Enjeu (€)":         r.get("Écart (€)") if _has_anom(r) else None,
                "Anomalie sur":      " / ".join(str(d.get("composant", "")) for d in r.get("_details", []))
                                     if _has_anom(r) else "Aucune anomalie",
            }
            for r in rows_sorted
        ])
        # Ligne TOTAL en bas (somme des enjeux quantifiés uniquement)
        enjeu_num = pd.to_numeric(df_export["Enjeu (€)"], errors="coerce")
        df_total = pd.DataFrame([{
            "Référence":         "TOTAL",
            "Fournisseur":       "",
            "Total Cost (€)":    pd.to_numeric(df_export["Total Cost (€)"], errors="coerce").sum(skipna=True),
            "Coût Optimisé (€)": "",
            "Enjeu (€)":         enjeu_num.sum(skipna=True),
            "Anomalie sur":      "",
        }])
        df_export_full = pd.concat([df_export, df_total], ignore_index=True)

        # Génération du buffer xlsx avec formatage propre
        buf = io.BytesIO()
        try:
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_export_full.to_excel(writer, sheet_name="Optimisation", index=False)
                # Formatage : largeurs de colonne + style header + ligne TOTAL
                ws = writer.sheets["Optimisation"]
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                # En-tête : navy fond, blanc texte, gras
                header_fill = PatternFill("solid", fgColor="0F172A")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                center = Alignment(horizontal="center", vertical="center")
                thin = Side(border_style="thin", color="CBD5E1")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = border
                # Largeurs des colonnes
                widths = {"A": 22, "B": 28, "C": 16, "D": 18, "E": 14, "F": 40}
                for col, w in widths.items():
                    ws.column_dimensions[col].width = w
                # Format nombre avec 2 décimales pour les colonnes monétaires
                from openpyxl.styles import numbers
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         min_col=3, max_col=5):
                    for cell in row:
                        cell.number_format = '#,##0.00 "€"'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.border = border
                # Bordure + alignement pour Référence & Fournisseur
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         min_col=1, max_col=2):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                # Mise en forme de la ligne TOTAL (dernière)
                total_row_idx = ws.max_row
                total_fill = PatternFill("solid", fgColor="065F46")
                total_font = Font(color="FFFFFF", bold=True, size=13)
                for cell in ws[total_row_idx]:
                    cell.fill = total_fill
                    cell.font = total_font
            buf.seek(0)
            excel_bytes = buf.getvalue()
            has_export = True
        except Exception:
            # Fallback sans formatage si openpyxl absent / erreur
            buf = io.BytesIO()
            df_export_full.to_excel(buf, index=False)
            buf.seek(0)
            excel_bytes = buf.getvalue()
            has_export = True
    except Exception:
        excel_bytes = None
        has_export = False

    # En-tête de section + bouton de téléchargement
    head_col, dl_col = st.columns([3, 1])
    with head_col:
        st.markdown(
            "<h3 style='color:#0f172a;font-weight:800;margin:18px 0 4px 0;'>"
            "📊 Détail par référence</h3>"
            "<p style='color:#64748b;font-size:12px;margin:0 0 10px 0;'>"
            "Cliquez sur <b>Analyser</b> à droite d'une référence pour voir le détail "
            "des composants en écart élevé.</p>",
            unsafe_allow_html=True
        )
    with dl_col:
        if has_export and excel_bytes:
            st.markdown("<div style='margin-top:24px;'></div>", unsafe_allow_html=True)
            st.download_button(
                label="📥 Exporter en Excel",
                data=excel_bytes,
                file_name="PackPricing_Optimisation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="batch_excel_export",
                use_container_width=True,
            )

    # ── Helper pour formater un nombre avec VIRGULE (format français, copier-coller Excel)
    def _fr(v):
        if v is None or (isinstance(v, float) and (v != v)):  # None ou NaN
            return "—"
        try:
            return f"{float(v):.2f}".replace(".", ",")
        except Exception:
            return str(v)

    # ── En-tête du tableau (7 colonnes — bouton sur 2 unités pour éviter wrap)
    col_w = [1.5, 2.0, 1.4, 1.4, 1.2, 2.4, 1.3]
    header_cols = st.columns(col_w)
    headers = ["Référence", "Fournisseur", "Total Cost (€)",
               "Coût Optimisé (€)", "Enjeu (€)", "Anomalie sur", ""]
    for c, h in zip(header_cols, headers):
        c.markdown(
            f"<div style='background:#0f172a;color:white;padding:8px 10px;"
            f"border-radius:6px;font-weight:700;font-size:12px;text-align:center;'>"
            f"{h}</div>",
            unsafe_allow_html=True
        )

    # ── Lignes : chaque ligne = colonnes + bouton "Voir"
    for idx, r in enumerate(rows_sorted):
        ref_code = r.get("Référence", "")
        sup = r.get("Fournisseur", "—") or "—"
        tc_str = _fr(r.get("Total Cost (€)"))
        co_val = r.get("Coût Optimisé (€)")
        ec_val = r.get("Écart (€)")
        details = r.get("_details", []) or []
        has_anomaly = bool(details) and ec_val and ec_val > 0

        bg = "#ffffff" if idx % 2 == 0 else "#f8fafc"

        if has_anomaly:
            # Référence AVEC anomalie : on montre coût optimisé, enjeu, composants
            co_str = _fr(co_val) if co_val is not None else "—"
            ec_str = _fr(ec_val)
            comp_names = [str(d.get("composant", ""))[:40] for d in details]
            anomaly_text = " / ".join(comp_names[:3])   # séparateur "/"
            if len(comp_names) > 3:
                anomaly_text += f" (+{len(comp_names)-3})"
            co_color, ec_color, anom_color = "#0ea5e9", "#10b981", "#dc2626"
            co_weight, ec_weight, anom_weight = "700", "700", "600"
        else:
            # Référence SANS anomalie : "DCL OK" et rien d'autre
            co_str = "✓ DCL OK"
            ec_str = "—"
            anomaly_text = "✓ Aucune anomalie"
            co_color, ec_color, anom_color = "#10b981", "#94a3b8", "#10b981"
            co_weight, ec_weight, anom_weight = "700", "400", "600"

        row_cols = st.columns(col_w)
        # Référence
        row_cols[0].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:12px;font-weight:700;color:#0f172a;'>"
            f"{ref_code}</div>",
            unsafe_allow_html=True
        )
        # Fournisseur
        row_cols[1].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:12px;color:#334155;'>{sup}</div>",
            unsafe_allow_html=True
        )
        # Total cost
        row_cols[2].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:12px;text-align:right;color:#334155;'>{tc_str}</div>",
            unsafe_allow_html=True
        )
        # Coût optimisé (ou "DCL OK")
        row_cols[3].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:12px;text-align:right;color:{co_color};"
            f"font-weight:{co_weight};'>{co_str}</div>",
            unsafe_allow_html=True
        )
        # Enjeu
        row_cols[4].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:12px;text-align:right;color:{ec_color};"
            f"font-weight:{ec_weight};'>{ec_str}</div>",
            unsafe_allow_html=True
        )
        # Anomalie sur (composants séparés par "/")
        row_cols[5].markdown(
            f"<div style='background:{bg};padding:10px;border:1px solid #e2e8f0;"
            f"border-radius:6px;font-size:11px;color:{anom_color};"
            f"font-weight:{anom_weight};line-height:1.3;'>"
            f"{anomaly_text}</div>",
            unsafe_allow_html=True
        )
        # Bouton Voir → ouvre DIRECTEMENT le détail de la référence sur l'accueil
        with row_cols[6]:
            if st.button("🔎 Voir", key=f"batch_analyze_{idx}_{ref_code}",
                         use_container_width=True,
                         help=f"Voir le détail de {ref_code}"):
                # On prépare tout l'état pour que l'accueil affiche directement
                # le détail (comme si l'utilisateur avait cherché la référence).
                try:
                    df_ref_v = get_ref_components(df_raw, ref_code)
                    summary_v = compute_ref_summary(df_ref_v, df_raw, ref_code)
                    st.session_state.ref_components_summary = summary_v
                    st.session_state.selected_ref = ref_code
                    st.session_state.ref_not_found = False
                except Exception:
                    st.session_state.selected_ref = ref_code
                st.session_state.ref_search_input = ref_code
                go_page("home")
                st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# ROUTAGE
# ═════════════════════════════════════════════════════════════════════════════

PAGE = st.session_state.page

if   PAGE == "home":                 render_home()
elif PAGE == "ref_detail":           render_ref_detail()
elif PAGE == "component_analysis":   render_component_analysis()
elif PAGE == "analysis":             render_analysis()
elif PAGE == "browse_types":         render_browse_types()
elif PAGE == "type_analysis":        render_type_analysis()
elif PAGE == "batch_optimization":   render_batch_optimization()